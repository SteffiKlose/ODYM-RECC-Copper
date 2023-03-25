﻿# -*- coding: utf-8 -*-
"""
Created on February 21, 2020, based on ODYM_RECC_V2_5.py by spauliuk
@authors: sklose
"""

"""
File ODYM_RECC-Copper.py
Contains the ODYM-RECC Copper model for the Sector-level estimates for global future copper demand and the potential for resource efficiency
Global coverage of six sectors.
passenger vehicles and other vehicle types, residential buildings, non-residential buildings, industry, appliances, and infrastructure.
dependencies:
    numpy >= 1.9
    scipy >= 0.14
"""
main = 1
##
if main==1:
#<f
#def main():
    # Import required libraries:
    import os
    import sys
    import logging as log
    import openpyxl
    import numpy as np
    import time
    import datetime
    #import scipy.io
    import pandas as pd
    import shutil   
    import uuid
    import matplotlib.pyplot as plt   
    from matplotlib.lines import Line2D
    import importlib
    import getpass
    from copy import deepcopy
    from tqdm import tqdm
    import scipy.stats
    from scipy.interpolate import interp1d
    from scipy.interpolate import make_interp_spline
    import pylab
    import pickle
    
    import RECC_Paths # Import path file
    
    #import re
    __version__ = str('2.5')
    ##################################
    #    Section 1)  Initialize      #
    ##################################
    # add ODYM module directory to system path
    sys.path.insert(0, os.path.join(os.path.join(RECC_Paths.odym_path,'odym'),'modules'))
    ### 1.1.) Read main script parameters
    # Mylog.info('### 1.1 - Read main script parameters')
    ProjectSpecs_Name_ConFile = 'RECC_Config.xlsx'
    Model_Configfile = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), data_only=True)
    ScriptConfig = {'Model Setting': Model_Configfile['Cover'].cell(4,4).value}
    Model_Configsheet = Model_Configfile[ScriptConfig['Model Setting']]
    #Read debug modus:   
    DebugCounter = 0
    while Model_Configsheet.cell(DebugCounter+1, 3).value != 'Logging_Verbosity':
        DebugCounter += 1
    ScriptConfig['Logging_Verbosity'] = Model_Configsheet.cell(DebugCounter+1,4).value # Read loggin verbosity once entry was reached.    
    # Extract user name from main file
    ProjectSpecs_User_Name     = getpass.getuser()
    
    # import packages whose location is now on the system path:    
    import ODYM_Classes as msc # import the ODYM class file
    importlib.reload(msc)
    import ODYM_Functions as msf  # import the ODYM function file
    importlib.reload(msf)
    import dynamic_stock_model as dsm # import the dynamic stock model library
    importlib.reload(dsm)
    
    Name_Script        = Model_Configsheet.cell(6,4).value
    if Name_Script != 'ODYM_RECC_Main':  # Name of this script must equal the specified name in the Excel config file
        raise AssertionError('Fatal: The name of the current script does not match to the sript name specfied in the project configuration file. Exiting the script.')
    # the model will terminate if the name of the script that is run is not identical to the script name specified in the config file.
    Name_Scenario            = Model_Configsheet.cell(7,4).value # Regional scope as torso for scenario name
    StartTime                = datetime.datetime.now()
    TimeString               = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day) + '__' + str(StartTime.hour) + '_' + str(StartTime.minute) + '_' + str(StartTime.second)
    #DateString               = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day)
    ProjectSpecs_Path_Result = os.path.join(RECC_Paths.results_path, Name_Scenario + '__' + TimeString )
    
    if not os.path.exists(ProjectSpecs_Path_Result): # Create model run results directory.
        os.makedirs(ProjectSpecs_Path_Result)
    # Initialize logger
    if ScriptConfig['Logging_Verbosity'] == 'DEBUG':
        log_verbosity = eval("log.DEBUG")  
    log_filename = Name_Scenario + '__' + TimeString + '.md'
    [Mylog, console_log, file_log] = msf.function_logger(log_filename, ProjectSpecs_Path_Result,
                                                         log_verbosity, log_verbosity)
    # log header and general information
    Time_Start = time.time()
    ScriptConfig['Current_UUID'] = str(uuid.uuid4())
    Mylog.info('# Simulation from ' + time.asctime())
    Mylog.info('Unique ID of scenario run: ' + ScriptConfig['Current_UUID'])
    
    ### 1.2) Read model control parameters
    Mylog.info('### 1.2 - Read model control parameters')
    #Read control and selection parameters into dictionary
    ScriptConfig = msf.ParseModelControl(Model_Configsheet,ScriptConfig)
    
    Mylog.info('Script: ' + Name_Script + '.py')
    Mylog.info('Model script version: ' + __version__)
    Mylog.info('Model functions version: ' + msf.__version__())
    Mylog.info('Model classes version: ' + msc.__version__())
    Mylog.info('Current User: ' + ProjectSpecs_User_Name)
    Mylog.info('Current Scenario: ' + Name_Scenario)
    Mylog.info(ScriptConfig['Description'])
    Mylog.debug('----\n')
    
    ### 1.3) Organize model output folder and logger
    Mylog.info('### 1.3 Organize model output folder and logger')
    #Copy Config file and model script into that folder
    shutil.copy(os.path.join(RECC_Paths.data_path,ProjectSpecs_Name_ConFile), os.path.join(ProjectSpecs_Path_Result, ProjectSpecs_Name_ConFile))
    #shutil.copy(Name_Script + '.py'      , os.path.join(ProjectSpecs_Path_Result, Name_Script + '.py'))
    
    #####################################################
    #     Section 2) Read classifications and data      #
    #####################################################
    Mylog.info('## 2 - Read classification items and define all classifications')
    ### 2.1) # Read model run config data
    Mylog.info('### 2.1 - Read model run config data')
    # Note: This part reads the items directly from the Exel master,
    # will be replaced by reading them from version-managed csv file.
    class_filename       = str(ScriptConfig['Version of master classification']) + '.xlsx'
    Classfile            = openpyxl.load_workbook(os.path.join(RECC_Paths.data_path,class_filename), data_only=True)
    Classsheet           = Classfile['MAIN_Table']
    MasterClassification = msf.ParseClassificationFile_Main(Classsheet,Mylog)
        
    Mylog.info('Read and parse config table, including the model index table, from model config sheet.')
    IT_Aspects,IT_Description,IT_Dimension,IT_Classification,IT_Selector,IT_IndexLetter,PL_Names,PL_Description,PL_Version,PL_IndexStructure,PL_IndexMatch,PL_IndexLayer,PrL_Number,PrL_Name,PrL_Comment,PrL_Type,ScriptConfig = msf.ParseConfigFile(Model_Configsheet,ScriptConfig,Mylog)    
    
    Mylog.info('Define model classifications and select items for model classifications according to information provided by config file.')
    ModelClassification  = {} # Dict of model classifications
    for m in range(0,len(IT_Aspects)):
        ModelClassification[IT_Aspects[m]] = deepcopy(MasterClassification[IT_Classification[m]])
        EvalString = msf.EvalItemSelectString(IT_Selector[m],len(ModelClassification[IT_Aspects[m]].Items))
        if EvalString.find(':') > -1: # range of items is taken
            RangeStart = int(EvalString[0:EvalString.find(':')])
            RangeStop  = int(EvalString[EvalString.find(':')+1::])
            ModelClassification[IT_Aspects[m]].Items = ModelClassification[IT_Aspects[m]].Items[RangeStart:RangeStop]           
        elif EvalString.find('[') > -1: # selected items are taken
            ModelClassification[IT_Aspects[m]].Items = [ModelClassification[IT_Aspects[m]].Items[i] for i in eval(EvalString)]
        elif EvalString == 'all':
            None
        else:
            Mylog.error('Item select error for aspect ' + IT_Aspects[m] + ' were found in datafile.')
            break
        
    ### 2.2) # Define model index table and parameter dictionary
    Mylog.info('### 2.2 - Define model index table and parameter dictionary')
    Model_Time_Start = int(min(ModelClassification['Time'].Items))
    Model_Time_End   = int(max(ModelClassification['Time'].Items))
    Model_Duration   = Model_Time_End - Model_Time_Start + 1
    
    Mylog.info('Define index table dataframe.')
    IndexTable = pd.DataFrame({'Aspect'        : IT_Aspects,  # 'Time' and 'Element' must be present!
                               'Description'   : IT_Description,
                               'Dimension'     : IT_Dimension,
                               'Classification': [ModelClassification[Aspect] for Aspect in IT_Aspects],
                               'IndexLetter'   : IT_IndexLetter})  # Unique one letter (upper or lower case) indices to be used later for calculations.
    
    # Default indexing of IndexTable, other indices are produced on the fly
    IndexTable.set_index('Aspect', inplace=True)
    
    # Add indexSize to IndexTable:
    IndexTable['IndexSize'] = pd.Series([len(IndexTable.Classification[i].Items) for i in range(0, len(IndexTable.IndexLetter))],
                                        index=IndexTable.index)
    
    # list of the classifications used for each indexletter
    IndexTable_ClassificationNames = [IndexTable.Classification[i].Name for i in range(0, len(IndexTable.IndexLetter))]
    
    # 2.3) Define shortcuts for the most important index sizes:
    Nt = len(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items)
    Ne = len(IndexTable.Classification[IndexTable.index.get_loc('Element')].Items)
    Nc = len(IndexTable.Classification[IndexTable.index.get_loc('Cohort')].Items)
    Nr = len(IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items)
    Nl = len(IndexTable.Classification[IndexTable.index.get_loc('Region11')].Items)
    No = len(IndexTable.Classification[IndexTable.index.get_loc('Region1')].Items)
    NG = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('G')].Items)
    Ng = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items)
    Np = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items)
    NB = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items)
    NN = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items) # varies: 24 for region-specific nrb and 4 for aggregated global resolution.
    NI = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items)
    NT = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items)
    Na = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items)
    Nv = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('v')].Items)

    #NA = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('A')].Items)
    NS = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items)
    NR = len(IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    Nw = len(IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items)
    Nm = len(IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items)
    NX = len(IndexTable.Classification[IndexTable.index.get_loc('Extensions')].Items)
    Nn = len(IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items)
    NV = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('V')].Items)
    Ns = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('s')].Items)
    #NT = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items)
    NL = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items)
    NO = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items)    
    Ni = len(IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('i')].Items)     # infrastructure
    
    #IndexTable.loc['t']['Classification'].Items # get classification items
    
    SwitchTime = Nc-Nt+1 # Index of first model year (2016)
    # 2.4) Read model data and parameters.
    Mylog.info('Read model data and parameters.')
    
    ParFileName = os.path.join(RECC_Paths.data_path,'RECC_ParameterDict_' + ScriptConfig['RegionalScope'] + '.dat')
    try: # Load Pickle parameter dict to save processing time
        ParFileObject = open(ParFileName,'rb')  
        ParameterDict = pickle.load(ParFileObject)  
        ParFileObject.close()  
        Mylog.info('Model data and parameters were read from pickled file with pickle file /parameter reading sequence UUID ' + ParameterDict['Checkkey'])
    except:
        ParameterDict = {}
        mo_start = 0 # set mo for re-reading a certain parameter
        for mo in range(mo_start,len(PL_Names)):
            #mo = 76 # set mo for re-reading a certain parameter
            #ParPath = os.path.join(os.path.abspath(os.path.join(ProjectSpecs_Path_Main, '.')), 'ODYM_RECC_Database', PL_Version[mo])
            ParPath = os.path.join(RECC_Paths.data_path, PL_Names[mo] + '_' + PL_Version[mo])
            Mylog.info('Reading parameter ' + PL_Names[mo])
            #MetaData, Values = msf.ReadParameter(ParPath = ParPath,ThisPar = PL_Names[mo], ThisParIx = PL_IndexStructure[mo], IndexMatch = PL_IndexMatch[mo], ThisParLayerSel = PL_IndexLayer[mo], MasterClassification,IndexTable,IndexTable_ClassificationNames,ScriptConfig,Mylog) # Do not change order of parameters handed over to function!
            # Do not change order of parameters handed over to function!
            MetaData, Values = msf.ReadParameterXLSX(ParPath, PL_Names[mo], PL_IndexStructure[mo], PL_IndexMatch[mo],
                                                 PL_IndexLayer[mo], MasterClassification, IndexTable,
                                                 IndexTable_ClassificationNames, ScriptConfig, Mylog, False)
            
            
     #    ParPath, ThisPar, ThisParIx, IndexMatch, ThisParLayerSel, MasterClassification,IndexTable, IndexTable_ClassificationNames, ScriptConfig, Mylog, ParseUncertainty =ParPath, PL_Names[mo], PL_IndexStructure[mo], PL_IndexMatch[mo],PL_IndexLayer[mo], MasterClassification, IndexTable,IndexTable_ClassificationNames, ScriptConfig, Mylog, False
            
            
            ParameterDict[PL_Names[mo]] = msc.Parameter(Name=MetaData['Dataset_Name'], ID=MetaData['Dataset_ID'],
                                                        UUID=MetaData['Dataset_UUID'], P_Res=None, MetaData=MetaData,
                                                        Indices=PL_IndexStructure[mo], Values=Values, Uncert=None,
                                                        Unit=MetaData['Dataset_Unit'])
            Mylog.info('Current parameter file UUID: ' + MetaData['Dataset_UUID'])
            Mylog.info('_')
        Mylog.info('Reading of parameters finished.')
        CheckKey = str(uuid.uuid4()) # generate UUID for this parameter reading sequence.
        Mylog.info('Current parameter reading sequence UUID: ' + CheckKey)
        Mylog.info('Entire parameter set stored under this UUID, will be reloaded for future calculations.')
        ParameterDict['Checkkey'] = CheckKey
        # Save to pickle file for next model run
        ParFileObject = open(ParFileName,'wb') 
        pickle.dump(ParameterDict,ParFileObject)   
        ParFileObject.close()
        
    Mylog.info('_')
    Mylog.info('_')
    ##############################################################
    #     Section 3)  Interpolate missing parameter values:      #
    ##############################################################
    # 0) obtain specific indices and positions:
    # m_reg_o         = 0 # reference region for GHG prices and intensities (Default: 0, which is the first region selected in the config file.)
    LEDindex        = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items.index('LED')
    SSP1index       = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items.index('SSP1')
    SSP2index       = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items.index('SSP2')
    
    SectorList      = eval(ScriptConfig['SectorSelect'])
    if 'nrb' in SectorList and 'nrbg' in SectorList:
        raise AssertionError('Fatal: Non-residential buildings are included both globally (nrbg) and for individual regions (nrb). Double-counting. Exiting the script, check config file.')    
    
    # index location and range of pass. vehs. in product list.
    try:
        Sector_pav_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('passenger vehicles')
    except:
        Sector_pav_loc  = np.nan
    try:
        Sector_pav_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items]
    except:
        if 'pav' in SectorList:
            raise AssertionError('Fatal: All selected items for aspect p must also be selected for aspect g. Exiting the script.')
        else:
            Sector_pav_rge = []
    # index location and range of res. builds. in product list.
    
    try:
        Sector_reb_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('residential buildings')
    except:
        Sector_reb_loc  = np.nan
    try:
        Sector_reb_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items]
    except:
        if 'reb' in SectorList:
            raise AssertionError('Fatal: All selected items for aspect B must also be selected for aspect g. Exiting the script.')
        else:
            Sector_reb_rge = []
        # index location and range of nonres. builds. in product list.
    if 'nrb' in SectorList:
        try:
            Sector_nrb_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('nonresidential buildings r')
        except:
            Sector_nrb_loc  = np.nan
        try:
            Sector_nrb_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items]
        except:
            Sector_nrb_rge  = []
    else:
        Sector_nrb_rge  = []
    # index location and range of nonres. builds. global in product list.    
    if 'nrbg' in SectorList:       
        try:
            Sector_nrbg_loc = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('nonresidential buildings g')
        except:
            Sector_nrbg_loc = np.nan
        try:
            Sector_nrbg_rge = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items]
        except:
            Sector_nrbg_rge = []
    else:
        Sector_nrbg_rge = []        
     # index location and range of industry in product list.    
    
    if 'ind' in SectorList:
        try:
            Sector_ind_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('industry')
        except:
            Sector_ind_loc  = np.nan
        try:
            Sector_ind_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items]
        except:
                raise AssertionError('Fatal: All selected items for aspect I must also be selected for aspect g. Exiting the script.')
    else:
        Sector_ind_rge = []
    # index location and range of appliances in product list. 
    if 'app' in SectorList:
        try:
            Sector_app_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('appliances')
        except:
            Sector_app_loc  = np.nan
        try:
            Sector_app_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items]
        except:
            
                raise AssertionError('Fatal: All selected items for aspect a must also be selected for aspect g. Exiting the script.')
    else:
        Sector_app_rge = []
            
                
    # index location and range of infrastructure in product list.    
    if 'inf' in SectorList:
        try:
            Sector_inf_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('infrastructure')
        except: 
            Sector_inf_loc  = np.nan
        try:
            Sector_inf_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('i')].Items]
        except:
                raise AssertionError('Fatal: All selected items for aspect i must also be selected for aspect g. Exiting the script.')
    else:
        Sector_inf_rge = []
    
                    
    # index location and range of other vehicles in product list.    
    if 'otv' in SectorList:
        try:
            Sector_otv_loc  = IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items.index('other road vehicles, trains, ships, aircrafts')
        except: 
            Sector_otv_loc  = np.nan
        try:
            Sector_otv_rge  = [IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items.index(i) for i in IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('v')].Items]
        except:
                raise AssertionError('Fatal: All selected items for aspect i must also be selected for aspect g. Exiting the script.')
    else:
        Sector_otv_rge = []
        
        
        
     
    Prod_Names = IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items
    Engineering_Materials_m2_List = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items
    Element_List = IndexTable.Classification[IndexTable.index.get_loc('Element')].Items
       
    if 'cement' in Engineering_Materials_m2_List:
        Cement_loc    = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('cement')
    else: Cement_loc = []
    
    if 'concrete' in Engineering_Materials_m2_List:
        Concrete_loc  = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('concrete')
    else: Concrete_loc = []
    
    
    if 'concrete aggregates' in Engineering_Materials_m2_List:
        ConcrAgg_loc  = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('concrete aggregates')
    else: ConcrAgg_loc = []
    
        
    if 'wood and wood products' in Engineering_Materials_m2_List:    
        Wood_loc      = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('wood and wood products')
    else: Wood_loc = []
    
        
    if 'wrought Al' in Engineering_Materials_m2_List:    
        WroughtAl_loc = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('wrought Al')
    else: WroughtAl_loc = []
    
        
    if 'cast Al' in Engineering_Materials_m2_List:    
        CastAl_loc    = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('cast Al')
    else: CastAl_loc = []
    
        
    if 'copper electric grade' in Engineering_Materials_m2_List:    
        Copper_loc    = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('copper electric grade')
    else: Copper_loc = []
    
        
    if 'plastics' in Engineering_Materials_m2_List:    
        Plastics_loc  = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('plastics')
    else: Plastics_loc = []
    
        
    if 'Zinc_loc ' in Engineering_Materials_m2_List:    
       Zinc_loc      = IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items.index('zinc')
    else: Zinc_loc = []
    
        
    Woodwaste_loc = IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items.index('used wood')
    Electric_loc  = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('electricity')
    WoodFuel_loc  = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('fuel wood')
    Hydrogen_loc  = IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items.index('hydrogen')
    
    
    if 'C ' in Element_List:    
       Carbon_loc    = IndexTable.Classification[IndexTable.index.get_loc('Element')].Items.index('C')
    else: Carbon_loc = []
    
    
    
    
    
    
    ClimPolScen   = IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items.index('RCP2.6')
    CO2_loc       = IndexTable.Classification[IndexTable.index.get_loc('Extensions')].Items.index('CO2 emisisons per main output')
    GWP100_loc    = IndexTable.Classification[IndexTable.index.get_loc('Environmental impact/pressure category')].Items.index('GWP100')
    Heating_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Heating')
    Cooling_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Cooling')
    DomstHW_loc   = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('DHW')
    Service_Drivg = IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items.index('Driving')
    Service_Reb   = np.array([Heating_loc,Cooling_loc,DomstHW_loc])
    Ind_2015      = 115 #index of year 2015
    #Ind_2017      = 117 #index of year 2017
    #Ind_2020      = 120 #index of year 2020
    
    
    
    CopperScrap_loc    = IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items.index('#1 Copper scrap') #copper wire scrap
    Cu_loc    = IndexTable.Classification[IndexTable.index.get_loc('Element')].Items.index('Cu')
    Cu_wmg_loc    = IndexTable.Classification[IndexTable.index.get_loc('WasteManagementIndustries')].Items.index('remelting of copper scrap')  #remelting of copper wire scrap
    
    Cu_scrap_loc    = IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items.index('#1 Copper scrap')
    
    
    
    
    
    
    IsClose_Remainder_Small = 1e-15 
    IsClose_Remainder_Large = 1e-7 
    DPI_RES        = ScriptConfig['Plot1Max'] # 50 for overview or 500 for paper plots, defined in ModelConfig_List
    
    # Determine location of the indices of individual sectors in the region-specific list and in the list of all goods
    # indices of sectors with same regional scope in complete goods list
    Sector_11reg_rge    = Sector_ind_rge
    Sector_1reg_rge     = Sector_nrbg_rge + Sector_app_rge + Sector_inf_rge + Sector_otv_rge
    
    
    Sector_32reg_rge =  []

    if 'pav' and 'reb' in SectorList:
        Sector_32reg_rge     = Sector_pav_rge + Sector_reb_rge + Sector_nrb_rge 
    
    elif 'pav' in SectorList:
        Sector_32reg_rge     = Sector_pav_rge 
            
    elif 'reb' in SectorList:
        Sector_32reg_rge = Sector_reb_rge 
    
    
    
    Loc_all_products = Sector_32reg_rge +Sector_11reg_rge + Sector_1reg_rge 
    #indices of individual end-use sectors within regionally separated product lists, check with classification master file!
    Sector_ind_rge_reg  = np.arange(0,21,1)
    Sector_app_rge_reg  = np.arange(4,18,1)
    Sector_nrbg_rge_reg = np.arange(0,4,1)
    Sector_inf_rge_reg = np.arange(18,35,1)
    Sector_pav_rge_reg = np.arange(0,6,1)
    Sector_otv_rge_reg = np.arange(35,47,1)

    
    if 'nrbg' in SectorList:
        Sector_app_rge_reg  = np.arange(4,18,1)

    else: 
        Sector_app_rge_reg  = np.arange(0,14,1)

    
    
    if 'pav' in SectorList: 
        Sector_reb_rge_reg = np.arange(6,19,1)
    
    else: 
        Sector_reb_rge_reg = np.arange(0,13,1)




    if 'inf' in SectorList and 'app' not in SectorList and 'nrbg' not in SectorList: 
        Sector_inf_rge_reg = np.arange(0,17,1)
       
       
    if 'otv' in SectorList and 'inf' not in SectorList and 'app' not in SectorList and 'nrbg' not in SectorList: 
        Sector_otv_rge_reg = np.arange(0,12,1)
           
       
       
       

    if ScriptConfig['Include_Scenario'] == 'ambitious':
        alpha = 0.8
        beta = 0.5
    
    else:
        alpha = 1
        beta = 1
        
        
    OutputDict      = {}  # Dictionary with output variables for entire model run, to export checks and analyses.
    
    # 1a) Material composition of vehicles, will only use historic age-cohorts.
    # Values are given every 5 years, we need all values in between.
    if 'pav' in SectorList:
        index = PL_Names.index('3_MC_RECC_Vehicles')
        MC_Veh_New = np.zeros(ParameterDict[PL_Names[index]].Values.shape)
        Idx_Time = [1980,1985,1990,1995,2000,2005,2010,2015,2020,2025,2030,2035,2040,2045,2050,2055,2060]
        Idx_Time_Rel = [i -1900 for i in Idx_Time]
        tnew = np.linspace(80, 160, num=81, endpoint=True)
        for n in range(0,Nm):
            for o in range(0,Np):
                for p in range(0,Nr):
                    f2 = interp1d(Idx_Time_Rel, ParameterDict[PL_Names[index]].Values[Idx_Time_Rel,n,o,p], kind='linear')
                    MC_Veh_New[80::,n,o,p] = f2(tnew)
        ParameterDict[PL_Names[index]].Values = MC_Veh_New.copy()
    
    # 1b) Material composition of res buildings, will only use historic age-cohorts.
    # Values are given every 5 years, we need all values in between.
    if 'reb' in SectorList:
        index       = PL_Names.index('3_MC_RECC_Buildings')
        index_Ren_A = PL_Names.index('3_MC_RECC_Buildings_Renovation_Absolute')
        index_Ren_R = PL_Names.index('3_MC_RECC_Buildings_Renovation_Relative')
        MC_Bld_New     = np.zeros(ParameterDict[PL_Names[index]].Values.shape)
        MC_Bld_New_Ren_A = np.zeros(ParameterDict[PL_Names[index_Ren_A]].Values.shape)
        MC_Bld_New_Ren_R = np.zeros(ParameterDict[PL_Names[index_Ren_R]].Values.shape)
        Idx_Time = [1900,1910,1920,1930,1940,1950,1960,1970,1980,1985,1990,1995,2000,2005,2010,2015,2020,2025,2030,2035,2040,2045,2050,2055,2060]
        Idx_Time_Rel = [i -1900 for i in Idx_Time]
        tnew = np.linspace(0, 160, num=161, endpoint=True)
        for n in range(0,Nm):
            for o in range(0,NB):
                for p in range(0,Nr):
                    f2 = interp1d(Idx_Time_Rel, ParameterDict[PL_Names[index]].Values[Idx_Time_Rel,n,o,p], kind='linear')
                    MC_Bld_New[:,n,o,p]       = f2(tnew).copy()
                    fA = interp1d(Idx_Time_Rel, ParameterDict[PL_Names[index_Ren_A]].Values[Idx_Time_Rel,n,o,p], kind='linear')
                    MC_Bld_New_Ren_A[:,n,o,p] = fA(tnew).copy()
                    fR = interp1d(Idx_Time_Rel, ParameterDict[PL_Names[index_Ren_R]].Values[Idx_Time_Rel,n,o,p], kind='linear')
                    MC_Bld_New_Ren_R[:,n,o,p] = fR(tnew).copy()
        ParameterDict[PL_Names[index]].Values       = MC_Bld_New.copy()
        ParameterDict[PL_Names[index_Ren_A]].Values = MC_Bld_New_Ren_A.copy()
        ParameterDict[PL_Names[index_Ren_R]].Values = MC_Bld_New_Ren_R.copy()
    
    # 1c) Material composition of nonres buildings, will only use historic age-cohorts.
    # Values are given every 5 years, we need all values in between.
    if 'nrb' in SectorList:
        index = PL_Names.index('3_MC_RECC_NonResBuildings')
        MC_NRB_New = np.zeros(ParameterDict[PL_Names[index]].Values.shape)
        Idx_Time = [1900,1910,1920,1930,1940,1950,1960,1970,1980,1985,1990,1995,2000,2005,2010,2015,2020,2025,2030,2035,2040,2045,2050,2055,2060]
        Idx_Time_Rel = [i -1900 for i in Idx_Time]
        tnew = np.linspace(0, 160, num=161, endpoint=True)
        for n in range(0,Nm):
            for o in range(0,NN):
                for p in range(0,Nr):
                    f2 = interp1d(Idx_Time_Rel, ParameterDict[PL_Names[index]].Values[Idx_Time_Rel,n,o,p], kind='linear')
                    MC_NRB_New[:,n,o,p] = f2(tnew).copy()
        ParameterDict[PL_Names[index]].Values = MC_NRB_New.copy()
    
    # 2a) Determine future energy intensity and material composition of vehicles by mixing archetypes:
    # Check if RE strategies are active and set implementation curves to 2016 value if not.
    if 'pav' in SectorList:
        if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'False': # no additional lightweighting trough material substitution.
            ParameterDict['3_SHA_LightWeighting_Vehicles'].Values  = np.einsum('prS,t->prtS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values[:,:,0,:],np.ones((Nt)))
        DownSizingBuffer = ParameterDict['3_SHA_DownSizing_Vehicles'].Values.copy()
        if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'True': # consider lightweighting trough UsingLessMaterialByDesign (segment shift)
            for nnr in range(0,Nr):
                for nnS in range(0,NS):
                    if ParameterDict['8_FLAG_VehicleDownsizingDirection'].Values[nnr,nnS] == 1:
                        ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,nnr,:,nnS] = np.einsum('s,t->st',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,nnr,0,nnS],np.ones((Nt)))
        else: # no lightweighting trough UsingLessMaterialByDesign.
            # for regions not selected above (high-income regions): Set downsizing to 2016 levels.
            ParameterDict['3_SHA_DownSizing_Vehicles'].Values = np.einsum('srS,t->srtS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,:,0,:],np.ones((Nt)))
            for nnr in range(0,Nr):
                for nnS in range(0,NS):
                    if ParameterDict['8_FLAG_VehicleDownsizingDirection'].Values[nnr,nnS] == 1:
                        ParameterDict['3_SHA_DownSizing_Vehicles'].Values[:,nnr,:,nnS]  = DownSizingBuffer[:,nnr,:,nnS].copy()
        ParameterDict['3_MC_RECC_Vehicles_RECC'] = msc.Parameter(Name='3_MC_RECC_Vehicles_RECC', ID='3_MC_RECC_Vehicles_RECC',
                                                   UUID=None, P_Res=None, MetaData=None,
                                                   Indices='cmprS', Values=np.zeros((Nc,Nm,Np,Nr,NS)), Uncert=None,
                                                   Unit='kg/unit')
        ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[0:SwitchTime-1,:,:,:,:] = np.einsum('cmpr,S->cmprS',ParameterDict['3_MC_RECC_Vehicles'].Values[0:SwitchTime-1,:,:,:],np.ones(NS))
        ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[SwitchTime-1::,:,:,:,:] = \
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[24,28,32,36,40,44],:])) +\
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[25,29,33,37,41,45],:])) +\
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[26,30,34,38,42,46],:])) +\
        np.einsum('prcS,pmrcS->cmprS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[27,31,35,39,43,47],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[0 ,4 ,8 ,12,16,20],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[1 ,5 ,9 ,13,17,21],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[2 ,6 ,10,14,18,22],:])) +\
        np.einsum('prcS,pmrcS->cmprS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pm->pmrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_MC_VehicleArchetypes'].Values[[3 ,7 ,11,15,19,23],:]))
        ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchTime-1::,:,Service_Drivg,:,:,:] = \
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[24,28,32,36,40,44],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[25,29,33,37,41,45],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[26,30,34,38,42,46],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100,     np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[27,31,35,39,43,47],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[0,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[0 ,4 ,8 ,12,16,20],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[1,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[1 ,5 ,9 ,13,17,21],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[2,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[2 ,6 ,10,14,18,22],:])) +\
        np.einsum('prcS,pnrcS->cpnrS',1 - ParameterDict['3_SHA_LightWeighting_Vehicles'].Values/100, np.einsum('rcS,pn->pnrcS',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[3,:,:,:],ParameterDict['3_EI_VehicleArchetypes'].Values[[3 ,7 ,11,15,19,23],:]))
    
    # 2b) Determine future energy intensity and material composition of residential buildings by mixing archetypes:
    # Expand building light-weighting split to all building types:
    if 'reb' in SectorList:
        ParameterDict['3_SHA_LightWeighting_Buildings'].Values = np.einsum('B,rtS->BrtS',np.ones(NB),ParameterDict['3_SHA_LightWeighting_Buildings'].Values[Sector_reb_loc,:,:,:]).copy()
        if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'False': # no additional lightweighting trough material substitution.
            ParameterDict['3_SHA_LightWeighting_Buildings'].Values = np.einsum('BrS,t->BrtS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values[:,:,0,:],np.ones((Nt)))
        if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'False': # no lightweighting trough UsingLessMaterialByDesign.
            ParameterDict['3_SHA_DownSizing_Buildings'].Values = np.einsum('urS,t->urtS',ParameterDict['3_SHA_DownSizing_Buildings'].Values[:,:,0,:],np.ones((Nt)))
        ParameterDict['3_MC_RECC_Buildings_RECC'] = msc.Parameter(Name='3_MC_RECC_Buildings_RECC', ID='3_MC_RECC_Buildings_RECC',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='cmBrS', Values=np.zeros((Nc,Nm,NB,Nr,NS)), Uncert=None,
                                                Unit='kg/m2')
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[0:115,:,:,:,:] = np.einsum('cmBr,S->cmBrS',ParameterDict['3_MC_RECC_Buildings'].Values[0:115,:,:,:],np.ones(NS))
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[115::,:,:,:,:] = \
        np.einsum('BrcS,BmrcS->cmBrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,Brm->BmrcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_MC_BuildingArchetypes'].Values[[87,88,89,90,91,92,93,94,95,96,97,98,99],:,:])) +\
        np.einsum('BrcS,BmrcS->cmBrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,Brm->BmrcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_MC_BuildingArchetypes'].Values[[61,62,63,64,65,66,67,68,69,70,71,72,73],:,:])) +\
        np.einsum('BrcS,BmrcS->cmBrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,Brm->BmrcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_MC_BuildingArchetypes'].Values[[74,75,76,77,78,79,80,81,82,83,84,85,86],:,:])) +\
        np.einsum('BrcS,BmrcS->cmBrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,Brm->BmrcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_MC_BuildingArchetypes'].Values[[48,49,50,51,52,53,54,55,56,57,58,59,60],:,:]))
        # Replicate values for Al, Cu, Plastics:
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[115::,[WroughtAl_loc,CastAl_loc,Copper_loc,Plastics_loc],:,:,:] = np.einsum('mBr,cS->cmBrS',ParameterDict['3_MC_RECC_Buildings'].Values[110,[WroughtAl_loc,CastAl_loc,Copper_loc,Plastics_loc],:,:].copy(),np.ones((Nt,NS)))
        # Split concrete into cement and aggregates:
        # Cement for buildings remains, as this item refers to cement in mortar, screed, and plaster. Cement in concrete is calculated as 3_MC_CementContentConcrete * concrete and added here. 
        # Concrete aggregates (0.87*concrete) are considered as well.
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,Cement_loc,:,:,:]   = ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,Cement_loc,:,:,:] + ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc] * ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,Concrete_loc,:,:,:].copy()
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,ConcrAgg_loc,:,:,:] = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,Concrete_loc,:,:,:].copy()
        ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,Concrete_loc,:,:,:] = 0
        
        ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values[:,Cement_loc,:,:]   = ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values[:,Cement_loc,:,:] + ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc] * ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values[:,Concrete_loc,:,:].copy()
        ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values[:,ConcrAgg_loc,:,:] = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values[:,Concrete_loc,:,:].copy()
        ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values[:,Concrete_loc,:,:] = 0
        
        ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[115::,:,:,:,:,:] = \
        np.einsum('BrcS,BnrVcS->cBVnrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,BrVn->BnrVcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_EI_BuildingArchetypes'].Values[[87,88,89,90,91,92,93,94,95,96,97,98,99],:,:,:])) +\
        np.einsum('BrcS,BnrVcS->cBVnrS',ParameterDict['3_SHA_LightWeighting_Buildings'].Values,     np.einsum('urcS,BrVn->BnrVcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_EI_BuildingArchetypes'].Values[[61,62,63,64,65,66,67,68,69,70,71,72,73],:,:,:])) +\
        np.einsum('BrcS,BnrVcS->cBVnrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,BrVn->BnrVcS',ParameterDict['3_SHA_DownSizing_Buildings'].Values,    ParameterDict['3_EI_BuildingArchetypes'].Values[[74,75,76,77,78,79,80,81,82,83,84,85,86],:,:,:])) +\
        np.einsum('BrcS,BnrVcS->cBVnrS',1 - ParameterDict['3_SHA_LightWeighting_Buildings'].Values, np.einsum('urcS,BrVn->BnrVcS',1 - ParameterDict['3_SHA_DownSizing_Buildings'].Values,ParameterDict['3_EI_BuildingArchetypes'].Values[[48,49,50,51,52,53,54,55,56,57,58,59,60],:,:,:]))
        ParameterDict['3_EI_Products_UsePhase_resbuildings_t'] = msc.Parameter(Name='3_EI_Products_UsePhase_resbuildings_t', ID='3_EI_Products_UsePhase_resbuildings_t',
                                                    UUID=None, P_Res=None, MetaData=None,
                                                    Indices='cBVnrt', Values=np.zeros((Nc,NB,NV,Nn,Nr,Nt)), Uncert=None,
                                                    Unit='MJ/m2/yr')
        ParameterDict['3_MC_RECC_Buildings_t'] = msc.Parameter(Name='3_MC_RECC_Buildings_t', ID='3_MC_RECC_Buildings_t',
                                                    UUID=None, P_Res=None, MetaData=None,
                                                    Indices='mBrctS', Values=np.zeros((Nm,NB,Nr,Nc,Nt,NS)), Uncert=None,
                                                    Unit='kg/m2')    
    

    if 'nrbg' in SectorList:
        # Split concrete into cement and aggregates:
        # Cement for buildings remains, as this item refers to cement in mortar, screed, and plaster. Cement in concrete is calculated as ParameterDict['3_MC_CementContentConcrete'].Values * concrete and added here. 
        # Concrete aggregates (0.87*concrete) are considered as well.
        ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Cement_loc,:]   = ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Cement_loc,:] + ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc] * ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Concrete_loc,:].copy()
        ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[ConcrAgg_loc,:] = (1 - ParameterDict['3_MC_CementContentConcrete'].Values[Cement_loc,Concrete_loc]) * ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Concrete_loc,:].copy()
        ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[Concrete_loc,:] = 0
        
    # 3) GHG intensity of energy supply: Change unit from g/MJ to kg/MJ and add backstop electricity factor
    ParameterDict['4_PE_GHGIntensityElectricitySupply_Backstop'].Values = ParameterDict['4_PE_GHGIntensityElectricitySupply_Backstop'].Values/1000 # convert g/MJ to kg/MJ
    ParameterDict['4_PE_GHGIntensityEnergySupply'].Values               = ParameterDict['4_PE_GHGIntensityEnergySupply'].Values/1000 # convert g/MJ to kg/MJ
    ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values         = ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values/1000 # convert g/MJ to kg/MJ
    # replace electricity emissions factors < backstop by backstop technology factor:
    ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,Electric_loc,:,:,:,:]       = np.maximum(ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,Electric_loc,:,:,:,:],      np.einsum('XSRt,r->XSRrt',ParameterDict['4_PE_GHGIntensityElectricitySupply_Backstop'].Values[:,Electric_loc,:,:,:],np.ones(Nr)))
    ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,Electric_loc,:,:,:,:] = np.maximum(ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,Electric_loc,:,:,:,:],np.einsum('XSRt,o->XSRot',ParameterDict['4_PE_GHGIntensityElectricitySupply_Backstop'].Values[:,Electric_loc,:,:,:],np.ones(No)))
    ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,Hydrogen_loc,:,:,:,:]       = np.maximum(ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,Hydrogen_loc,:,:,:,:],      np.einsum('XSRt,r->XSRrt',ParameterDict['4_PE_GHGIntensityElectricitySupply_Backstop'].Values[:,Hydrogen_loc,:,:,:],np.ones(Nr)))
    ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,Hydrogen_loc,:,:,:,:] = np.maximum(ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,Hydrogen_loc,:,:,:,:],np.einsum('XSRt,o->XSRot',ParameterDict['4_PE_GHGIntensityElectricitySupply_Backstop'].Values[:,Hydrogen_loc,:,:,:],np.ones(No)))
    
    # 4) Fabrication yield and fabrication scrap diversion:
    # Extrapolate 2050-2060 as 2015 values
    index = PL_Names.index('4_PY_Manufacturing')
    ParameterDict[PL_Names[index]].Values[:,:,:,:,1::,:] = np.einsum('t,mwgFr->mwgFtr',np.ones(45),ParameterDict[PL_Names[index]].Values[:,:,:,:,0,:])
    if ScriptConfig['Include_REStrategy_FabScrapDiversion'] == 'False':
        ParameterDict['6_PR_FabricationScrapDiversion'].Values = np.zeros((Nm,Nw,No,NS))
    
    # 5) EoL RR, apply world average to all regions
    ParameterDict['4_PY_EoL_RecoveryRate'].Values = np.einsum('gmwW,r->grmwW',ParameterDict['4_PY_EoL_RecoveryRate'].Values[:,0,:,:,:],np.ones((Nr)))
    
    # 6) Energy carrier split of vehicles, replicate fixed values for all regions and age-cohorts etc.
    ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles'].Values = np.einsum('pn,crVS->cprVnS',ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles'].Values[115,:,0,3,:,SSP1index].copy(),np.ones((Nc,Nr,NV,NS)))
    
    # 7) RE strategy potentials for individual countries are replicated from global average:
    ParameterDict['6_PR_ReUse_Bld'].Values                      = np.einsum('mB,r->mBr',ParameterDict['6_PR_ReUse_Bld'].Values[:,:,0],np.ones(Nr))
#    ParameterDict['6_PR_ReUse_nonresBld'].Values                = np.einsum('mN,r->mNr',ParameterDict['6_PR_ReUse_nonresBld'].Values[:,:,0],np.ones(Nr))
    ParameterDict['6_PR_LifeTimeExtension_passvehicles'].Values = np.einsum('pS,r->prS',ParameterDict['6_PR_LifeTimeExtension_passvehicles'].Values[:,0,:],np.ones(Nr))
    
    ParameterDict['6_PR_LifeTimeExtension_industry'].Values = np.einsum('pS,l->plS',ParameterDict['6_PR_LifeTimeExtension_industry'].Values[:,0,:],np.ones(Nl))
    
    
    #ParameterDict['6_PR_EoL_RR_Improvement'].Values             = np.einsum('gmwW,r->grmwW',ParameterDict['6_PR_EoL_RR_Improvement'].Values[:,0,:,:,:],np.ones(Nr))
    
    # 8) Define a multi-regional RE strategy and building renovation scaleup parameter
    ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'] = msc.Parameter(Name='3_SHA_RECC_REStrategyScaleUp_r', ID='3_SHA_RECC_REStrategyScaleUp_r',
                                                      UUID=None, P_Res=None, MetaData=None,
                                                      Indices='trSR', Values=np.zeros((Nt,Nr,NS,NR)), Uncert=None,
                                                      Unit='1')
    ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values        = np.einsum('RtS,r->trSR',ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[:,0,:,:],np.ones(Nr)).copy()
    ParameterDict['3_SHA_BuildingRenovationScaleUp_r'] = msc.Parameter(Name='3_SHA_BuildingRenovationScaleUp_r', ID='3_SHA_BuildingRenovationScaleUp_r',
                                                      UUID=None, P_Res=None, MetaData=None,
                                                      Indices='trSR', Values=np.zeros((Nt,Nr,NS,NR)), Uncert=None,
                                                      Unit='1')
    ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values        = np.einsum('RtS,r->trSR',ParameterDict['3_SHA_BuildingRenovationScaleUp'].Values[:,0,:,:],np.ones(Nr)).copy()
    
    # 9) LED scenario data from proxy scenarios:
    # 2_P_RECC_Population_SSP_32R
    ParameterDict['2_P_RECC_Population_SSP_32R'].Values[:,:,:,LEDindex]                    = ParameterDict['2_P_RECC_Population_SSP_32R'].Values[:,:,:,SSP2index].copy()
    # 3_EI_Products_UsePhase, historic
    ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[0:115,:,:,:,:,LEDindex]    = ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[0:115,:,:,:,:,SSP2index].copy()
    ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,:,:,:,LEDindex]    = ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,:,:,:,SSP2index].copy()
    ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:115,:,:,:,:,LEDindex] = ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:115,:,:,:,:,SSP2index].copy()
    # 3_IO_Buildings_UsePhase
    ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[:,:,:,:,LEDindex]             = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[:,:,:,:,SSP2index].copy()
    
    # 10) Set future vehicle reuse to 2015 levels if strategy is not included:
    # (To reflect that reuse is already happening to some extent.)
    if ScriptConfig['Include_REStrategy_ReUse'] == 'False':
        ParameterDict['6_PR_ReUse_Veh'].Values       = np.einsum('mprS,t->mprtS',ParameterDict['6_PR_ReUse_Veh'].Values[:,:,:,1,:],np.ones(Nt)) # stay at current levels, which are > 0.
        ParameterDict['6_PR_ReUse_Bld'].Values       = np.zeros(ParameterDict['6_PR_ReUse_Bld'].Values.shape) # set to zero, which corresponds to current levels.
        ParameterDict['6_PR_ReUse_nonresBld'].Values = np.zeros(ParameterDict['6_PR_ReUse_nonresBld'].Values.shape) # set to zero, which corresponds to current levels.
        
        
         
        ParameterDict['6_PR_ReUse_app'].Values = np.einsum('maS,t->matS',ParameterDict['6_PR_ReUse_app'].Values[:,:,1,:],np.ones(Nt)) # stay at current levels, which are > 0.
        ParameterDict['6_PR_ReUse_inf'].Values = np.zeros(ParameterDict['6_PR_ReUse_inf'].Values.shape)  
        ParameterDict['6_PR_ReUse_ind'].Values = np.zeros(ParameterDict['6_PR_ReUse_ind'].Values.shape)
        
        
        
    # 11) MODEL CALIBRATION
    # Calibrate vehicle kilometrage: No longer used! VKM is now calibrated in scenario target table process to deliver correct pC stock number for 2015.
    #### ParameterDict['3_IO_Vehicles_UsePhase'].Values[3,:,:,:]                             = ParameterDict['3_IO_Vehicles_UsePhase'].Values[3,:,:,:]                           * np.einsum('r,tS->rtS',ParameterDict['6_PR_Calibration'].Values[0,:],np.ones((Nt,NS)))
    # Calibrate vehicle fuel consumption, cgVnrS    
    ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[0:115,:,3,:,:,:]        = ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[0:115,:,3,:,:,:]      * np.einsum('r,cpnS->cpnrS',ParameterDict['6_PR_Calibration'].Values[1,:],np.ones((115,Np,Nn,NS)))
    # Calibrate res. building energy consumption
    ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,0:3,:,:,:]      = ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:115,:,0:3,:,:,:]    * np.einsum('r,cBVnS->cBVnrS',ParameterDict['6_PR_Calibration'].Values[2,:],np.ones((115,NB,3,Nn,NS)))
    # Calibrate nonres. building energy consumption
    ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:115,:,0:3,:,:,:]   = ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[0:115,:,0:3,:,:,:] * np.einsum('r,cNVnS->cNVnrS',ParameterDict['6_PR_Calibration'].Values[3,:],np.ones((115,NN,3,Nn,NS)))
    
    # 12) No recycling scenario (counterfactual reference)
    if ScriptConfig['IncludeRecycling'] == 'False': # no recycling and remelting
        ParameterDict['4_PY_EoL_RecoveryRate'].Values            = np.zeros(ParameterDict['4_PY_EoL_RecoveryRate'].Values.shape)
        ParameterDict['4_PY_MaterialProductionRemelting'].Values = np.zeros(ParameterDict['4_PY_MaterialProductionRemelting'].Values.shape)
        
    # 13) Currently not used.
      
    # 14) Define parameter for future vehicle stock:
    # a) calculated passenger vehicle stock
    ParameterDict['2_S_RECC_FinalProducts_Future_passvehicles'] = msc.Parameter(Name='2_S_RECC_FinalProducts_Future_passvehicles', ID='2_S_RECC_FinalProducts_Future_passvehicles',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='StGr', Values=np.zeros((NS,Nt,NG,Nr)), Uncert=None,
                                                Unit='cars per person')
    # b) calculated vehicle kilometrage
    ParameterDict['3_IO_Vehicles_UsePhase_eff'] = msc.Parameter(Name='3_IO_Vehicles_UsePhase_eff', ID='3_IO_Vehicles_UsePhase_eff',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VrtS', Values=np.zeros((NV,Nr,Nt,NS)), Uncert=None,
                                                Unit='km per vehicle')
    
    # 15) Define parameter for future building stock:
    # actual future res and nonres building stock
    ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'] = msc.Parameter(Name='2_S_RECC_FinalProducts_Future_resbuildings_act', ID='2_S_RECC_FinalProducts_Future_resbuildings_act',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='StGr', Values=np.zeros((NS,Nt,NG,Nr)), Uncert=None,
                                                Unit='m2 per person')
    ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings_act'] = msc.Parameter(Name='2_S_RECC_FinalProducts_Future_NonResBuildings_act', ID='2_S_RECC_FinalProducts_Future_NonResBuildings_act',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='GrtS', Values=np.zeros((NG,Nr,Nt,NS)), Uncert=None,
                                                Unit='m2 per person')    
    # 3_IO changing over time:
    ParameterDict['3_IO_Buildings_UsePhase'] = msc.Parameter(Name='3_IO_Buildings_UsePhase', ID='3_IO_Buildings_UsePhase',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='tcBVrS', Values=np.zeros((Nt,Nc,NB,NV,Nr,NS)), Uncert=None,
                                                Unit='1')
    # Historic age-cohorts:
    # ParameterDict['3_IO_Buildings_UsePhase_Historic'] is a combination of climate and socioeconomic 3_IO determinants.
    # We single out the former and keep them constant and let the socioeconomic factors change according to the '3_IO_Buildings_UsePhase_Future_...' parameters.
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Heating_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Heating_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_Heating,np.ones(Nt))
    Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW     = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,DomstHW_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,DomstHW_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_DHW,np.ones(Nt))
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling = ParameterDict['3_IO_Buildings_UsePhase_Historic'].Values[0:SwitchTime,:,Cooling_loc,:,:] / np.einsum('rS,cB->cBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Cooling'].Values[Sector_reb_loc,:,0,:],np.ones((SwitchTime,NB))) * 100
    Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling[np.isnan(Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling)] = 0
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,Cooling_loc,:,:] = np.einsum('cBrS,t->tcBrS',Par_3_IO_Buildings_UsePhase_Historic_Climate_Cooling,np.ones(Nt))
    # Correct for if some of the corrections lead to IO > 1 which may be the case when hist. IO data are incomplete and thus set to 1 already.
    ParameterDict['3_IO_Buildings_UsePhase'].Values[ParameterDict['3_IO_Buildings_UsePhase'].Values > 1] = 1
    # Future age-cohorts:
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Heating_loc,:,:] = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,DomstHW_loc,:,:] = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Heating'].Values[Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,Cooling_loc,:,:] = np.einsum('rtS,cB->tcBrS',ParameterDict['3_IO_Buildings_UsePhase_Future_Cooling'].Values[Sector_reb_loc,:,:,:]/100,np.ones((Nc-SwitchTime,NB)))
    
    # expand 3_IO parameter for nonres buildings with 1 for post 2015 years:
    #ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[SwitchTime::,:,:,:,:] = 1
    
    # 16) Compile parameter for building energy conversion efficiency:
    ParameterDict['4_TC_ResidentialEnergyEfficiency'] = msc.Parameter(Name='4_TC_ResidentialEnergyEfficiency', ID='4_TC_ResidentialEnergyEfficiency',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None,
                                                Unit='1')
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values                                   = np.einsum('VRrn,tS->VRrntS',ParameterDict['4_TC_ResidentialEnergyEfficiency_Default'].Values[:,:,:,:,0],np.ones((Nt,NS)))
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Heating_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Heating'].Values[Heating_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[Cooling_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Cooling'].Values[Cooling_loc,:,:,Electric_loc,:,:] / 100
    ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[DomstHW_loc,:,:,Electric_loc,:,:] = ParameterDict['4_TC_ResidentialEnergyEfficiency_Scenario_Heating'].Values[DomstHW_loc,:,:,Electric_loc,:,:] / 100
    
    # 17) No energy efficiency improvements (counterfactual reference)
    # Freeze type split and archetypes at 2015/17/2020 levels:
    if ScriptConfig['No_EE_Improvements'] == 'True':
        SwitchIndex = Ind_2015 # choose between Ind_2015/17/20 (for continuation from 2015/17/20 onwards.
        if 'pav' in SectorList:
            for nnr in range(0,Nr):
                for nnS in range(0,NS):
                    if ParameterDict['8_FLAG_VehicleDownsizingDirection'].Values[nnr,nnS] == 0: # don't consider type split reset for cases, where type split change involves larger vehicles.
                        ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchIndex::,:,:,:,nnr,nnS]    = np.einsum('pVn,c->cpVn',ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchIndex,:,:,:,nnr,nnS],np.ones(Nc-SwitchIndex))
                        ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[SwitchIndex::,:,:,nnr,nnS]  = np.einsum('mp,c->cmp',ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[SwitchIndex,:,:,nnr,nnS],np.ones(Nc-SwitchIndex))
            ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[:,:,:,:,4::]                           = np.einsum('GrRp,t->GrRpt',ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[:,:,:,:,4],np.ones(Nt-4)) # index 4 is year 2020.
        if 'reb' in SectorList:
            ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchIndex::,:,:,:,:,:]    = np.einsum('BVnrS,c->cBVnrS',ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchIndex,:,:,:,:,:],np.ones(Nc-SwitchIndex))
            ParameterDict['3_MC_RECC_Buildings_RECC'].Values[SwitchIndex::,:,:,:,:]                 = np.einsum('mBrS,c->cmBrS',ParameterDict['3_MC_RECC_Buildings_RECC'].Values[SwitchIndex,:,:,:,:],np.ones(Nc-SwitchIndex))
            ParameterDict['3_SHA_TypeSplit_Buildings'].Values[:,:,4::,:]                            = np.einsum('BrS,t->BrtS',ParameterDict['3_SHA_TypeSplit_Buildings'].Values[:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
            ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,:,:,:,4::,:]                 = np.einsum('VRrnS,t->VRrntS',ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,:,:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
        # if 'nrb' in SectorList:            
        #     ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[SwitchIndex::,:,:,:,:,:] = np.einsum('NVnrS,c->cNVnrS',ParameterDict['3_EI_Products_UsePhase_nonresbuildings'].Values[SwitchIndex,:,:,:,:,:],np.ones(Nc-SwitchIndex))
        #     ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[SwitchIndex::,:,:,:,:]           = np.einsum('mNrS,c->cmNrS',ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[SwitchIndex,:,:,:,:],np.ones(Nc-SwitchIndex))
        #     ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values[:,:,4::,:]                      = np.einsum('NrS,t->NrtS',ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values[:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
        #     ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,:,:,:,4::,:]                 = np.einsum('VRrnS,t->VRrntS',ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values[:,:,:,:,4,:],np.ones(Nt-4)) # index 4 is year 2020.
    
    # 18) Derive energy supply multipliers for buildings for future age-cohorts
    # From energy carrier split and conversion efficiency, the multipliers converting 1 MJ of final building energy demand into different energy carriers are determined.
    ParameterDict['3_SHA_EnergySupply_Buildings'] = msc.Parameter(Name='3_SHA_EnergySupply_Buildings', ID='3_SHA_EnergySupply_Buildings',
                                                UUID=None, P_Res=None, MetaData=None,
                                                Indices='VRrntS', Values=np.zeros((NV,NR,Nr,Nn,Nt,NS)), Uncert=None,
                                                Unit='1')
    Divisor = ParameterDict['4_TC_ResidentialEnergyEfficiency'].Values #VRrntS
    ParameterDict['3_SHA_EnergySupply_Buildings'].Values = np.divide(np.einsum('VRrnt,S->VRrntS',ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values, np.ones(NS)), Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
    SHA_EnergySupply_Buildings_Sum_n = np.einsum('VRrntS->VRrtS',ParameterDict['3_SHA_EnergySupply_Buildings'].Values).copy()
    Divisor = np.einsum('VRrtS,n->VRrntS',SHA_EnergySupply_Buildings_Sum_n,np.ones(Nn)) # The following division happens twice! (cf. model docu)
    ParameterDict['3_SHA_EnergySupply_Buildings'].Values = np.divide(ParameterDict['3_SHA_EnergySupply_Buildings'].Values, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
    ParameterDict['3_SHA_EnergySupply_Buildings'].Values = np.divide(ParameterDict['3_SHA_EnergySupply_Buildings'].Values, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
    
    # 19) Make sure that all share parameters are non-negative and add up to 100%:
    # not necessary as data fulfil constraints.
    #ParameterDict['3_SHA_TypeSplit_Buildings'].Values[ParameterDict['3_SHA_TypeSplit_Buildings'].Values < 0] = 0
    #ParameterDict['3_SHA_TypeSplit_Buildings'].Values = ParameterDict['3_SHA_TypeSplit_Buildings'].Values / np.einsum('rtS,B->BrtS',ParameterDict['3_SHA_TypeSplit_Buildings'].Values.sum(axis=0),np.ones(NB))
    #ParameterDict['3_SHA_TypeSplit_Buildings'].Values[np.isnan(ParameterDict['3_SHA_TypeSplit_Buildings'].Values)] = 0
    
    # 20) Extrapolate appliances beyond 2050:
    # for noS in range(0,NS):
    #     for noR in range(0,NR):
    #         for noa in range(0,Na):
    #             if ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,140,noS,noR,noa] != 0:
    #                 growthrate = (ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,150,noS,noR,noa]/ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,140,noS,noR,noa]-1)/10
    #             else:
    #                 growthrate = 0
    #             for noT in range(151,161):
    #                 ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,noT,noS,noR,noa] = ParameterDict['1_F_RECC_FinalProducts_appliances'].Values[0,150,noS,noR,noa] * np.power(1+growthrate,noT-150)
        
    # 21) GWP_bio factor interpolation
    Idx_Time = [1900,1910,1920,1930,1940,1950,1960,1970,1980,1990,2000]
    Idx_Time_Rel = [i -1900 for i in Idx_Time]
    tnew = np.linspace(0, 100, num=101, endpoint=True)
    f2 = interp1d(Idx_Time_Rel, ParameterDict['6_MIP_GWP_Bio'].Values[Idx_Time_Rel].copy(), kind='linear')
    ParameterDict['6_MIP_GWP_Bio'].Values = np.zeros((300))
    ParameterDict['6_MIP_GWP_Bio'].Values[0:101] = f2(tnew).copy()    
    ParameterDict['6_MIP_GWP_Bio'].Values[101::] = -1
        
    # 22) calculate Stocks on 1. Jan 2016:    
    pC_AgeCohortHist           = np.zeros((NG,Nr))
    #pC_FutureStock             = np.zeros((NS,NG,Nr))
    # a) from historic data:
    Stocks_2016_passvehicles   = ParameterDict['2_S_RECC_FinalProducts_2015_passvehicles'].Values[0,:,:,:].sum(axis=0)
    pCStocks_2016_passvehicles = np.einsum('pr,r->rp',Stocks_2016_passvehicles,1/ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,0,:,1]) 
    Stocks_2016_resbuildings   = ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,:].sum(axis=0)
    pCStocks_2016_resbuildings = np.einsum('Br,r->rB',Stocks_2016_resbuildings,1/ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,0,:,1]) 
  #  Stocks_2016_nresbuildings  = ParameterDict['2_S_RECC_FinalProducts_2015_nonresbuildings'].Values[0,:,:,:].sum(axis=0)
  #  pCStocks_2016_nresbuildings= np.einsum('Nr,r->rN',Stocks_2016_nresbuildings,1/ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,0,:,1]) 
    if 'pav' in SectorList:
        pC_AgeCohortHist[Sector_pav_loc, :] = pCStocks_2016_passvehicles.sum(axis =1)
    if 'reb' in SectorList:
        pC_AgeCohortHist[Sector_reb_loc, :] = pCStocks_2016_resbuildings.sum(axis =1)
    # if 'nrb' in SectorList:    
    #     pC_AgeCohortHist[Sector_nrb_loc, :] = pCStocks_2016_nresbuildings.sum(axis =1)
    OutputDict['pC_AgeCohortHist']   = pC_AgeCohortHist.copy()
    # b) from future stock curves:
    #This is done for the individual sector calculations below.
    
    ##########################################################
    #    Section 4) Initialize dynamic MFA model for RECC    #
    ##########################################################
    Mylog.info('Initialize dynamic MFA model for RECC')
    Mylog.info('Define RECC system and processes.')
    
    #Define arrays for result export:
    GWP_System_3579di                = np.zeros((Nt,NS,NR))
    GWP_UsePhase_7d                  = np.zeros((Nt,NS,NR))
    GWP_OtherThanUsePhaseDirect      = np.zeros((Nt,NS,NR))
    GWP_Materials_3di_9di            = np.zeros((Nt,NS,NR)) # all processes and their energy supply chains except for manufacturing and use phase
    GWP_Vehicles_Direct              = np.zeros((Nt,Nr,NS,NR)) # use phase only
    GWP_ReBuildgs_Direct             = np.zeros((Nt,Nr,NS,NR)) # use phase only
    GWP_NRBuildgs_Direct             = np.zeros((Nt,Nr,NS,NR)) # use phase only
    GWP_NRBuildgs_Direct_g           = np.zeros((Nt,NS,NR)) # use phase only
    #GWP_NonResBuildings_Direct       = np.zeros((Nt,Nr,NS,NR)) # use phase only
    GWP_Vehicles_indir               = np.zeros((Nt,NS,NR)) # energy supply only
    GWP_AllBuildings_indir           = np.zeros((Nt,NS,NR)) # energy supply only
    #GWP_NonResBuilding_id            = np.zeros((Nt,NS,NR)) # energy supply only
    GWP_Manufact_5di_all             = np.zeros((Nt,NS,NR))
    GWP_WasteMgt_9di_all             = np.zeros((Nt,NS,NR))
    GWP_PrimaryMaterial_3di          = np.zeros((Nt,NS,NR))
    GWP_PrimaryMaterial_3di_m        = np.zeros((Nt,Nm,NS,NR))
    GWP_SecondaryMetal_di_m          = np.zeros((Nt,Nm,NS,NR))
    GWP_UsePhase_7i_Scope2_El        = np.zeros((Nt,NS,NR))
    GWP_UsePhase_7i_OtherIndir       = np.zeros((Nt,NS,NR))
    GWP_MaterialCycle_5di_9di        = np.zeros((Nt,NS,NR))
    GWP_RecyclingCredit              = np.zeros((Nt,NS,NR))
    GWP_ForestCO2Uptake              = np.zeros((Nt,NS,NR))
    GWP_WoodCycle                    = np.zeros((Nt,NS,NR)) # net GHG impact of wood use: forest uptake + wood-related emissions from waste mgt. Pos sign for flow from system to environment.
    GWP_EnergyRecoveryWasteWood      = np.zeros((Nt,NS,NR))
    GWP_ByEnergyCarrier_UsePhase_d   = np.zeros((Nt,Nr,Nn,NS,NR))
    GWP_ByEnergyCarrier_UsePhase_i   = np.zeros((Nt,Nr,Nn,NS,NR))
    Material_Inflow                  = np.zeros((Nt,Ng,Nm,NS,NR))
    Material_Outflow                 = np.zeros((Nt,Ng,Nm,NS,NR))
    Scrap_Outflow                    = np.zeros((Nt,Nw,NS,NR))
    PrimaryProduction                = np.zeros((Nt,Nm,NS,NR))
    SecondaryProduct                 = np.zeros((Nt,Nm,NS,NR))
    SecondaryProducts_tot            = np.zeros((Nt,Nm,NS,NR))  
                     
    SecondaryExport                  = np.zeros((Nt,Nm,NS,NR))
    RenovationMaterialInflow_7       = np.zeros((Nt,Nm,NS,NR))
    Element_Material_Composition     = np.zeros((Nt,Nm,Ne,NS,NR))
    Element_Material_Composition_raw = np.zeros((Nt,Nm,Ne,NS,NR))
    Element_Material_Composition_con = np.zeros((Nt,Nm,Ne,NS,NR))
    Manufacturing_Output             = np.zeros((Nt,Ng,Nm,NS,NR))
    StockMatch_2015                  = np.zeros((NG,Nr))
    NegInflowFlags                   = np.zeros((NS,NR))
    #NegInflowFlags_After2020         = np.zeros((NS,NR))
    FabricationScrap                 = np.zeros((Nt,Nw,NS,NR))
    EnergyCons_UP_Vh                 = np.zeros((Nt,NS,NR))
    EnergyCons_UP_Bd                 = np.zeros((Nt,NS,NR))
    EnergyCons_Mn                    = np.zeros((Nt,NS,NR))
    EnergyCons_Wm                    = np.zeros((Nt,NS,NR))
    EnergyCons_UP_Service            = np.zeros((Nt,Nr,NV,NS,NR))
    EnergyCons_UP_total              = np.zeros((Nt,Nn,NS,NR))
    EnergyCons_total                 = np.zeros((Nt,Nn,NS,NR))
    StockCurves_Totl                 = np.zeros((Nt,NG,NS,NR))
    StockCurves_Prod                 = np.zeros((Nt,Ng,NS,NR))
    StockCurves_Mat                  = np.zeros((Nt,Nm,NS,NR))
    Inflow_Prod                      = np.zeros((Nt,Ng,NS,NR))
    Inflow_Prod_r                    = np.zeros((Nt,Nr,Ng,NS,NR))
    Outflow_Prod                     = np.zeros((Nt,Ng,NS,NR))
    EoL_Products_for_WasteMgt        = np.zeros((Nt,Ng,NS,NR))
    Outflow_Materials_Usephase_all   = np.zeros((Nt,Nm,NS,NR))
    Outflow_Products_Usephase_all    = np.zeros((Nt,Ng,NS,NR))
    WasteMgtLosses_To_Landfill       = np.zeros((Nt,Ne,NS,NR))
    Population                       = np.zeros((Nt,Nr,NS,NR))
    pCStocksCurves                   = np.zeros((Nt,NG,Nr,NS,NR))
    Passenger_km                     = np.zeros((Nt,NS,NR))
    Vehicle_km                       = np.zeros((Nt,NS,NR))
    Service_IO_ResBuildings          = np.zeros((Nt,NV,NS,NR))
    Service_IO_NonResBuildings       = np.zeros((Nt,NV,NS,NR))
    ReUse_Materials                  = np.zeros((Nt,Nm,NS,NR))
    Carbon_Wood_Inflow               = np.zeros((Nt,NS,NR))
    Carbon_Wood_Outflow              = np.zeros((Nt,NS,NR))
    Carbon_Wood_Stock                = np.zeros((Nt,NS,NR))
    Vehicle_FuelEff                  = np.zeros((Nt,Np,Nr,NS,NR))
    ResBuildng_EnergyCons            = np.zeros((Nt,NB,Nr,NS,NR))
    GWP_bio_Credit                   = np.zeros((Nt,NS,NR))
    EnergyRecovery_WoodCombustion_EL = np.zeros((Nt,NS,NR))
    BiogenicCO2WasteCombustion       = np.zeros((Nt,NS,NR))
    
    StockCurves_UsePhase_appliances = np.zeros((Nc,Na,No))
    StockCurves_UsePhase_appliances_NS_NR = np.zeros((Nc,Na,No,NS,NR))

    StockCurves_UsePhase_industry = np.zeros((Nc,NI,Nl))
    StockCurves_UsePhase_industry_NS_NR = np.zeros((Nc,NI,Nl,NS,NR)) 
    
    ObsoleteStocks                   = np.zeros((Nt,Ng,NS,NR))
    HybernatingStocks                = np.zeros((Nt,Ng,NS,NR))
    SortingLosses                    = np.zeros((Nt,Ng,NS,NR))
    RecoveryLosses                   = np.zeros((Nt,NS,NR))
    
    ObsoleteFlows                  = np.zeros((Nt,Ng,NS,NR))
    HybernatingFlows             = np.zeros((Nt,Ng,NS,NR))
    SortingFlows                  = np.zeros((Nt,Ng,NS,NR))
    RecoveryFlows                   = np.zeros((Nt,NS,NR))
    
    Cu_Stocks_Nr               = np.zeros((Nt,Ng,NS,NR))
    Cu_Stocks_No               = np.zeros((Nt,Ng,NS,NR))
    Cu_Stocks_Nl               = np.zeros((Nt,Ng,NS,NR))


    
    NegInflowFlags                   = np.zeros((NG,NS,NR))
    time_dsm                         = np.arange(0,Nc,1) # time array of [0:Nc) needed for some sectors
    
    ExitFlags = {} # Exit flags for individual model runs
    #  Examples for testing
    #mS = 1
    #mR = 1
    
    # Select and loop over scenarios
    for mS in range(0,NS):
        for mR in range(0,NR):
            SName = IndexTable.loc['Scenario'].Classification.Items[mS]
            RName = IndexTable.loc['Scenario_RCP'].Classification.Items[mR]
            Mylog.info('_')
            Mylog.info('Computing RECC model for SSP scenario ' + SName + ' and RE scenario ' + RName + '.')
            
            # Initialize MFA system
            RECC_System = msc.MFAsystem(Name='RECC_SingleScenario',
                                        Geogr_Scope='19 regions + 1 single country', #IndexTableR.Classification[IndexTableR.set_index('IndexLetter').index.get_loc('r')].Items,
                                        Unit='Mt',
                                        ProcessList=[],
                                        FlowDict={},
                                        StockDict={},
                                        ParameterDict=ParameterDict,
                                        Time_Start=Model_Time_Start,
                                        Time_End=Model_Time_End,
                                        IndexTable=IndexTable,
                                        Elements=IndexTable.loc['Element'].Classification.Items,
                                        Graphical=None)
                                  
            # Check Validity of index tables:
            # returns true if dimensions are OK and time index is present and element list is not empty
            RECC_System.IndexTableCheck()
            
            # Add processes to system
            for m in range(0, len(PrL_Number)):
                RECC_System.ProcessList.append(msc.Process(Name = PrL_Name[m], ID   = PrL_Number[m]))
                
            # Define system variables: Flows.    
            RECC_System.FlowDict['F_0_1']     = msc.Flow(Name='CO2 uptake', P_Start=0, P_End=1,
                                                     Indices='t,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
    
            RECC_System.FlowDict['F_1_2']     = msc.Flow(Name='harvested wood', P_Start=1, P_End=2,
                                                     Indices='t,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
    
            RECC_System.FlowDict['F_2_3']     = msc.Flow(Name='timber consumed by sawmills', P_Start=2, P_End=3,
                                                     Indices='t,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_2_7']     = msc.Flow(Name='wood fuel use', P_Start=2, P_End=7,
                                                     Indices='t,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None) # flow is directly routed to use phase.
    
            RECC_System.FlowDict['F_7_0']     = msc.Flow(Name='wood fuel use direct emissions', P_Start=7, P_End=0,
                                                     Indices='t,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
    
            RECC_System.FlowDict['F_0_3']     = msc.Flow(Name='ore input', P_Start=0, P_End=3,
                                                     Indices='t,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_3_4']     = msc.Flow(Name='primary material production' , P_Start = 3, P_End = 4, 
                                                     Indices = 't,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_4_5']     = msc.Flow(Name='primary material consumption' , P_Start = 4, P_End = 5, 
                                                     Indices = 't,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_5_6']     = msc.Flow(Name='manufacturing output' , P_Start = 5, P_End = 6, 
                                                     Indices = 't,o,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
                
     
      
            RECC_System.FlowDict['F_5_10']    = msc.Flow(Name='new scrap' , P_Start = 5, P_End = 10, 
                                                     Indices = 't,o,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_10_9']    = msc.Flow(Name='scrap use' , P_Start = 10, P_End = 9, 
                                                     Indices = 't,o,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_9_12']    = msc.Flow(Name='secondary material production' , P_Start = 9, P_End = 12, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
            RECC_System.FlowDict['F_10_12']   = msc.Flow(Name='fabscrapdiversion' , P_Start = 10, P_End = 12, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)        
            
            RECC_System.FlowDict['F_12_5']    = msc.Flow(Name='secondary material consumption' , P_Start = 12, P_End = 5, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_12_21']    = msc.Flow(Name='excess secondary material' , P_Start = 12, P_End = 0, 
                                                     Indices = 't,o,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            # RECC_System.FlowDict['F_9_13']     = msc.Flow(Name='waste mgt. and remelting losses' , P_Start = 9, P_End = 0, 
            #                                          Indices = 't,e', Values=None, Uncert=None, 
            #                                          Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_6_7']     = msc.Flow(Name='final consumption', P_Start=6, P_End=7,
                                                     Indices='t,r,g,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
    
            
            RECC_System.FlowDict['F_7_8']     = msc.Flow(Name='EoL products' , P_Start = 7, P_End = 8, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
      
            
            RECC_System.FlowDict['F_8_11']     = msc.Flow(Name='obsolete stock formation' , P_Start = 8, P_End = 11, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)       
    
            RECC_System.FlowDict['F_8_9']     = msc.Flow(Name='waste mgt. input' , P_Start = 8, P_End = 9, 
                                                     Indices = 't,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_17']    = msc.Flow(Name='product re-use in' , P_Start = 8, P_End = 17, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
                   
            RECC_System.FlowDict['F_17_6']    = msc.Flow(Name='product re-use out' , P_Start = 17, P_End = 6, 
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_9_10']    = msc.Flow(Name='old scrap' , P_Start = 9, P_End = 10, 
                                                     Indices = 't,r,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
         
            RECC_System.FlowDict['F_14_13']    = msc.Flow(Name='old scrap' , P_Start = 9, P_End = 13, 
                                                     Indices = 't,g,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
          
            
            RECC_System.FlowDict['F_10_23']    = msc.Flow(Name='Recovery loss formation' , P_Start = 12, P_End = 23, 
                                                     Indices = 't,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
    
    
         #   if 'pav' in SectorList or 'nrb' in SectorList or 'reb' in SectorList:
    
            
            RECC_System.FlowDict['F_6_7_Nr']  = msc.Flow(Name='final consumption No', P_Start=6, P_End=7,
                                                     Indices='t,r,T,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_9_10_Nr'] = msc.Flow(Name='old scrap No' , P_Start = 9, P_End = 10, 
                                                     Indices = 't,r,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_14_13_Nr'] = msc.Flow(Name='old scrap No' , P_Start = 9, P_End = 13, 
                                                     Indices = 't,T,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            
            
            RECC_System.FlowDict['F_17_6_Nr'] = msc.Flow(Name='product re-use out' , P_Start = 17, P_End = 6, 
                                                     Indices = 't,c,r,T,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_17_Nr'] = msc.Flow(Name='product re-use in No' , P_Start = 8, P_End = 17, 
                                                     Indices = 't,c,r,T,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_19_Nr'] = msc.Flow(Name='hybernating stock formation Nr' , P_Start = 8, P_End = 19, 
                                                     Indices = 't,c,r,T,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_9_Nr']  = msc.Flow(Name='waste mgt. input No' , P_Start = 8, P_End = 9, 
                                                     Indices = 't,r,T,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
            RECC_System.FlowDict['F_7_8_Nr']  = msc.Flow(Name='EoL products No' , P_Start = 7, P_End = 8, 
                                                     Indices = 't,c,r,T,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
    
            RECC_System.FlowDict['F_8_11_Nr']  = msc.Flow(Name='obsolete stock formation Nr' , P_Start = 8, P_End = 11, 
                                                     Indices = 't,c,r,T,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
    
               
         #   if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList:
    
            
            RECC_System.FlowDict['F_6_7_No']  = msc.Flow(Name='final consumption No', P_Start=6, P_End=7,
                                                     Indices='t,o,O,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
            
            RECC_System.FlowDict['F_9_10_No'] = msc.Flow(Name='old scrap No' , P_Start = 9, P_End = 10, 
                                                     Indices = 't,o,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_14_13_No'] = msc.Flow(Name='old scrap No' , P_Start = 9, P_End = 13, 
                                                     Indices = 't,O,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            
            
            RECC_System.FlowDict['F_17_6_No'] = msc.Flow(Name='product re-use out' , P_Start = 17, P_End = 6, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_17_No'] = msc.Flow(Name='product re-use in No' , P_Start = 8, P_End = 17, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_9_No']  = msc.Flow(Name='waste mgt. input No' , P_Start = 8, P_End = 9, 
                                                     Indices = 't,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
    
            RECC_System.FlowDict['F_7_8_No']  = msc.Flow(Name='EoL products No' , P_Start = 7, P_End = 8, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
    
            RECC_System.FlowDict['F_8_11_No']  = msc.Flow(Name='obsolete stock formation No' , P_Start = 8, P_End = 11, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_19_No'] = msc.Flow(Name='hybernating stock formation No' , P_Start = 8, P_End = 19, 
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            
            
          #  if 'ind' in SectorList:
    
            RECC_System.FlowDict['F_6_7_Nl']  = msc.Flow(Name='final consumption Nl', P_Start=6, P_End=7,
                                                     Indices='t,l,L,m,e', Values=None, Uncert=None,
                                                     Color=None, ID=None, UUID=None)
                            
            RECC_System.FlowDict['F_7_8_Nl']  = msc.Flow(Name='EoL products Nl' , P_Start = 7, P_End = 8, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
             
            RECC_System.FlowDict['F_8_11_Nl']  = msc.Flow(Name='obsolete stock formation Nl' , P_Start = 8, P_End = 11, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)        
            
            RECC_System.FlowDict['F_9_10_Nl'] = msc.Flow(Name='old scrap Nl' , P_Start = 9, P_End = 10, 
                                                     Indices = 't,l,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            
            RECC_System.FlowDict['F_14_13_Nl'] = msc.Flow(Name='old scrap Nl' , P_Start = 9, P_End = 13, 
                                                     Indices = 't,L,w,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            
            RECC_System.FlowDict['F_17_6_Nl'] = msc.Flow(Name='product re-use out' , P_Start = 17, P_End = 6, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_17_Nl'] = msc.Flow(Name='product re-use in Nl' , P_Start = 8, P_End = 17, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
            RECC_System.FlowDict['F_8_19_Nl'] = msc.Flow(Name='hybernating stock formation Nl' , P_Start = 8, P_End = 19, 
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
            
        
                    
            RECC_System.FlowDict['F_8_9_Nl']  = msc.Flow(Name='waste mgt. input Nl' , P_Start = 8, P_End = 9, 
                                                     Indices = 't,l,L,m,e', Values=None, Uncert=None, 
                                                     Color = None, ID = None, UUID = None)
                    
            # Define system variables: Stocks.
            RECC_System.StockDict['dS_0']     = msc.Stock(Name='System environment stock change', P_Res=0, Type=1,
                                                     Indices = 't,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_1t']    = msc.Stock(Name='Forestry stock change, timber', P_Res=1, Type=1,
                                                     Indices = 't,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_1t']     = msc.Stock(Name='Forestry carbon stock, timber', P_Res=1, Type=0,
                                                     Indices = 't,c,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_1f']    = msc.Stock(Name='Forestry stock change, fuel wood', P_Res=1, Type=1,
                                                     Indices = 't,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_1f']     = msc.Stock(Name='Forestry carbon stock, fuel wood', P_Res=1, Type=0,
                                                     Indices = 't,c,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)        
            #no regional dim anymore    
            RECC_System.StockDict['S_7']      = msc.Stock(Name='In-use stock', P_Res=7, Type=0,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            #no regional dim anymore    
            RECC_System.StockDict['dS_7']     = msc.Stock(Name='In-use stock change', P_Res=7, Type=1,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
    
            
            RECC_System.StockDict['S_10']     = msc.Stock(Name='Fabrication scrap buffer', P_Res=10, Type=0,
                                                     Indices = 't,c,o,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_10']    = msc.Stock(Name='Fabrication scrap buffer change', P_Res=10, Type=1,
                                                     Indices = 't,o,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_12']     = msc.Stock(Name='secondary material buffer', P_Res=12, Type=0,
                                                     Indices = 't,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_12']    = msc.Stock(Name='Secondary material buffer change', P_Res=12, Type=1,
                                                     Indices = 't,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            #no regional dim anymore
            RECC_System.StockDict['S_11']     = msc.Stock(Name='Obsolete Stocks 32 Reg', P_Res=11, Type=0,
                                                     Indices = 't,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            #no regional dim anymore
            RECC_System.StockDict['dS_11']    = msc.Stock(Name='Obsolete Stocks change 32 Reg', P_Res=11, Type=1,
                                                     Indices = 't,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
    
            #no regional dim anymore
            RECC_System.StockDict['S_19']     = msc.Stock(Name='Hybernating Stocks', P_Res=19, Type=0,
                                                     Indices = 't,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            #no regional dim anymore
            RECC_System.StockDict['dS_19']    = msc.Stock(Name='Hybernating Stocks change', P_Res=19, Type=1,
                                                     Indices = 't,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
    
                    
            RECC_System.StockDict['S_13']     = msc.Stock(Name='Sorting losses World', P_Res=13, Type=0,
                                                     Indices = 't,g,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_13']    = msc.Stock(Name='Sorting losses World', P_Res=13, Type=1,
                                                     Indices = 't,g,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
          
                           
            RECC_System.StockDict['S_23']     = msc.Stock(Name='Recovery losses World', P_Res=23, Type=0,
                                                     Indices = 't,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_23']    = msc.Stock(Name='Recovery losses World', P_Res=23, Type=1,
                                                     Indices = 't,w,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
          #
           # if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList:        
            RECC_System.StockDict['dS_7_No']  = msc.Stock(Name='In-use stock change', P_Res=7, Type=1,
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)        
            
            RECC_System.StockDict['dS_11_No']    = msc.Stock(Name='Obsolete Stocks change 1 Reg', P_Res=11, Type=1,
                                                     Indices = 't,O,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_11_No']     = msc.Stock(Name='Obsolete Stocks 1 Reg', P_Res=11, Type=0,
                                                     Indices = 't,O,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
             
            RECC_System.StockDict['S_19_No']     = msc.Stock(Name='Hybernating Stocks 1 Reg', P_Res=19, Type=0,
                                                     Indices = 't,O,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            
            RECC_System.StockDict['dS_19_No']    = msc.Stock(Name='Hybernating Stocks change 1 Reg', P_Res=19, Type=1,
                                                     Indices = 't,O,o,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
    
            RECC_System.StockDict['S_7_No']   = msc.Stock(Name='In-use stock', P_Res=7, Type=0,
                                                     Indices = 't,c,o,O,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
       
           # if 'ind' in SectorList:
    
            RECC_System.StockDict['S_7_Nl']   = msc.Stock(Name='In-use stock', P_Res=7, Type=0,
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
    
            RECC_System.StockDict['S_19_Nl']     = msc.Stock(Name='Hybernating Stocks 11 Reg', P_Res=12, Type=0,
                                                     Indices = 't,L,l,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_19_Nl']    = msc.Stock(Name='Hybernating Stocks change 11 Reg', P_Res=12, Type=1,
                                                     Indices = 't,L,l,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            RECC_System.StockDict['dS_7_Nl']  = msc.Stock(Name='In-use stock change', P_Res=7, Type=1,
                                                     Indices = 't,c,l,L,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)      
            
            RECC_System.StockDict['S_11_Nl']     = msc.Stock(Name='Obsolete Stocks 11 Reg', P_Res=12, Type=0,
                                                     Indices = 't,L,l,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['dS_11_Nl']    = msc.Stock(Name='Obsolete Stocks change 11 Reg', P_Res=12, Type=1,
                                                     Indices = 't,L,l,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            
            
            RECC_System.StockDict['dS_7_Nr']  = msc.Stock(Name='In-use stock change', P_Res=7, Type=1,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)        
            
            RECC_System.StockDict['dS_11_Nr']    = msc.Stock(Name='Obsolete Stocks change 32 Reg', P_Res=11, Type=1,
                                                     Indices = 't,g,r,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            RECC_System.StockDict['S_11_Nr']     = msc.Stock(Name='Obsolete Stocks 32Reg', P_Res=11, Type=0,
                                                     Indices = 't,g,r,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
             
            RECC_System.StockDict['S_19_Nr']     = msc.Stock(Name='Hybernating Stocks 32 Reg', P_Res=19, Type=0,
                                                     Indices = 't,g,r,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
            
            
            RECC_System.StockDict['dS_19_Nr']    = msc.Stock(Name='Hybernating Stocks change 1 Reg', P_Res=19, Type=1,
                                                     Indices = 't,g,r,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
    
            RECC_System.StockDict['S_7_Nr']   = msc.Stock(Name='In-use stock', P_Res=7, Type=0,
                                                     Indices = 't,c,r,g,m,e', Values=None, Uncert=None,
                                                     ID=None, UUID=None)
       
    
            
            
            RECC_System.Initialize_StockValues() # Assign empty arrays to stocks according to dimensions.
            RECC_System.Initialize_FlowValues()  # Assign empty arrays to flows according to dimensions.
            
            ##########################################################
            #    Section 5) Solve dynamic MFA model for RECC         #
            ##########################################################
            Mylog.info('Solve dynamic MFA model for the RECC project for all sectors chosen.')
            # THIS IS WHERE WE EXPAND FROM THE FORMAL MODEL STRUCTURE AND CODE WHATEVER IS NECESSARY TO SOLVE THE MODEL EQUATIONS.
            SwitchTime=Nc - Model_Duration +1 # Year when future modelling horizon starts: 1.1.2016        
    
            # Determine empty result containers for all sectors
            Stock_Detail_UsePhase_p     = np.zeros((Nt,Nc,Np,Nr)) # index structure: tcpr. Unit: million items.
            Outflow_Detail_UsePhase_p   = np.zeros((Nt,Nc,Np,Nr)) # index structure: tcpr. Unit: million items.
            Inflow_Detail_UsePhase_p    = np.zeros((Nt,Np,Nr))    # index structure: tpr.  Unit: million items.
            
            Stock_Detail_UsePhase_B     = np.zeros((Nt,Nc,NB,Nr)) # index structure: tcBr. Unit: million m².
            Outflow_Detail_UsePhase_B   = np.zeros((Nt,Nc,NB,Nr)) # index structure: tcBr. Unit: million m².
            Inflow_Detail_UsePhase_B    = np.zeros((Nt,NB,Nr))    # index structure: tBr.  Unit: million m².
        
            Stock_Detail_UsePhase_N     = np.zeros((Nt,Nc,NN,Nr)) # index structure: tcNr. Unit: million m².
            Outflow_Detail_UsePhase_N   = np.zeros((Nt,Nc,NN,Nr)) # index structure: tcNr. Unit: million m².
            Inflow_Detail_UsePhase_N    = np.zeros((Nt,NN,Nr))    # index structure: tNr.  Unit: million m². 
            
            Stock_Detail_UsePhase_Ng    = np.zeros((Nt,Nc,NN,No)) # index structure: tcNo. Unit: million m².
            Outflow_Detail_UsePhase_Ng  = np.zeros((Nt,Nc,NN,No)) # index structure: tcNo. Unit: million m².
            Inflow_Detail_UsePhase_Ng   = np.zeros((Nt,NN,No))    # index structure: tNo.  Unit: million m².         
            
            Stock_Detail_UsePhase_I     = np.zeros((Nt,Nc,NI,Nl)) # index structure: tcIL. Unit: GW.
            Outflow_Detail_UsePhase_I   = np.zeros((Nt,Nc,NI,Nl)) # index structure: tcIl. Unit: GW.
            Inflow_Detail_UsePhase_I    = np.zeros((Nt,NI,Nl))    # index structure: tIl.  Unit: GW.
        
            Stock_Detail_UsePhase_a     = np.zeros((Nt,Nc,Na,No)) # index structure: tcao. Unit: # of items (1).
            Outflow_Detail_UsePhase_a   = np.zeros((Nt,Nc,Na,No)) # index structure: tcao. Unit: # of items (1).
            Inflow_Detail_UsePhase_a    = np.zeros((Nt,Na,No))    # index structure: tao.  Unit: # of items (1).    
            
            
            Stock_Detail_UsePhase_i    = np.zeros((Nt,Nc,Ni,No)) # index structure: tcNo. Unit: .
            Outflow_Detail_UsePhase_i  = np.zeros((Nt,Nc,Ni,No)) # index structure: tcNo. Unit: .
            Inflow_Detail_UsePhase_i   = np.zeros((Nt,Ni,No)) 
            
            
            Stock_Detail_UsePhase_otv    = np.zeros((Nt,Nc,Nv,No)) # index structure: tcNo. Unit: .
            Outflow_Detail_UsePhase_otv  = np.zeros((Nt,Nc,Nv,No)) # index structure: tcNo. Unit: .
            Inflow_Detail_UsePhase_otv   = np.zeros((Nt,Nv,No)) 
            
            F_6_7_ren                   = np.zeros((Nt,Nc,Nr,Ng,Nm,Ne)) # Indices='t,c,r,g,m,e', # inflow of renovation material, Mt/yr
            F_6_7_new                   = np.zeros((Nt,Nr,Ng,Nm,Ne))    # Indices='t,r,g,m,e',   # inflow of material in new products, Mt/yr
        
            # Sector: Passenger vehicles
            if 'pav' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, passenger vehicles.')
                # 1) Determine kilometrage endogenously and apply stock-driven model
                SF_Array                    = np.zeros((Nc,Nc,Np,Nr)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                
                #Get historic stock at end of 2015 by age-cohort, and covert unit to Vehicles: million.
                TotalStock_UsePhase_Hist_cpr = RECC_System.ParameterDict['2_S_RECC_FinalProducts_2015_passvehicles'].Values[0,:,:,:]
                
                # Determine total future stock, product level. Units: Vehicles: million.
                # Option implemented: By service curve.
                Total_Service_pav_tr_pC                     = np.einsum('rt->tr',RECC_System.ParameterDict['1_F_Function_Future'].Values[Sector_pav_loc,:,:,mS])
                if ScriptConfig['Include_REStrategy_CarSharing'] == 'False': # set carsharing to zero.
                    RECC_System.ParameterDict['6_PR_CarSharingShare'].Values  = np.zeros(RECC_System.ParameterDict['6_PR_CarSharingShare'].Values.shape)
                if ScriptConfig['Include_REStrategy_RideSharing'] == 'False': # set ride-sharing to zero.                
                    RECC_System.ParameterDict['6_PR_RideSharingShare'].Values = np.zeros(RECC_System.ParameterDict['6_PR_RideSharingShare'].Values.shape)
    
                # i) Calculate pc stocks in the four subdivisions: CaS, RiS, CaS+RiS, none:
                Total_Vehicle_km_pav_tr_pC          = np.zeros((Nt,Nr))
                TotalStockCurves_UsePhase_p_pC_test = np.zeros((Nt,Nr)) 
                for nrr in range(0,Nr):
                    for ntt in range(0,Nt):   
                        # convert abs. change in vehicle OR to relative change:
                        MIP_RideSharing_Occupancy_rel = (RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] + RECC_System.ParameterDict['6_MIP_RideSharing_Occupancy'].Values[mS,nrr]) / RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS]
                        # calculate future stock levels:
                        s0        = (1 - RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(1 - RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS])
                        s_CaS     = (RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(1 - RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS]) \
                                  * (RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr])
                        s_RiS     = (1 - RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS]) \
                                  / (MIP_RideSharing_Occupancy_rel)
                        s_CaS_RiS = (RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] / 100)*(RECC_System.ParameterDict['6_PR_RideSharingShare'].Values[Sector_pav_loc,nrr,ntt,mS] / 100) \
                                  * Total_Service_pav_tr_pC[ntt,nrr] /(RECC_System.ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,nrr,ntt,mS] * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS]) \
                                  / (MIP_RideSharing_Occupancy_rel / RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr])
    
                        s_total   = s0.copy() + s_CaS.copy() + s_RiS.copy() + s_CaS_RiS.copy()
                        TotalStockCurves_UsePhase_p_pC_test[ntt,nrr] = s_total.copy()
                        TotalStockCurves_UsePhase_p_pC_test[np.isnan(TotalStockCurves_UsePhase_p_pC_test)] = 0 # ignore drive technologies where there is no stock.
    
                        # ii) Calculate average vehicle kilometrage and average occupancy rate:
                        vkm       = ((s0 + s_RiS) + (s_CaS + s_CaS_RiS) / RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr]) * RECC_System.ParameterDict['3_IO_Vehicles_UsePhase'].Values[Service_Drivg,nrr,ntt,mS].copy() / s_total
                        Total_Vehicle_km_pav_tr_pC[ntt,nrr] = vkm.copy()
                        # Overwrite predefined values by internally calculated vehicle-km:
                        RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values[Service_Drivg,nrr,ntt,mS] = Total_Vehicle_km_pav_tr_pC[ntt,nrr]
                        #ocr       = Total_Service_pav_tr_pC[ntt,nrr] / (s_total * vkm)
                        
                RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values[np.isnan(RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values)] = 0
                # iii) Make sure that for no scenario, stock values are below LED values, which is assumed to be the lowest possible stock level.         
                # This needs to be made sure during the scenario framing process! Here, only the accounting and model equations to convert PKM to VKM and stock are executed, no further checks are made.
                TotalStockCurves_UsePhase_p_pC   = TotalStockCurves_UsePhase_p_pC_test.copy()
                TotalStockCurves_UsePhase_p      = np.einsum('tr,tr->tr',TotalStockCurves_UsePhase_p_pC, RECC_System.ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,:,:,mS])
                RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_passvehicles'].Values[mS,:,Sector_pav_loc,:] = TotalStockCurves_UsePhase_p_pC.copy() 
                
                # iv) adjust vehicle lifetime to new effective value to reflect impact of car-sharing:
                # First, replicate lifetimes for all age-cohorts
                Par_RECC_ProductLifetime_p = np.einsum('c,pr->prc',np.ones((Nc)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_passvehicles'].Values)
                # Second, adjust lifetime if car-sharing is present
                if ScriptConfig['Include_REStrategy_CarSharing'] == 'True': # adjust lifetime of future age-cohorts.
                    for npp in range(0,Np):
                        for nrr in range(0,Nr):
                            for ntt in range(0,Nt):
                                Par_RECC_ProductLifetime_p[npp,nrr,ntt+SwitchTime-1] = (1 - RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS]/100 + RECC_System.ParameterDict['6_PR_CarSharingShare'].Values[Sector_pav_loc,0,ntt,mS] * RECC_System.ParameterDict['6_MIP_CarSharing_Stock'].Values[mS,nrr]/100) * Par_RECC_ProductLifetime_p[npp,nrr,ntt+SwitchTime-1]
                            
                # Include_REStrategy_LifeTimeExtension: Product lifetime extension.
                # Third, change lifetime of future age-cohorts according to lifetime extension parameter
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
                    Par_RECC_ProductLifetime_p[:,:,SwitchTime -1::] = np.einsum('crp,prc->prc',1 + np.einsum('cr,pr->crp',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],beta * RECC_System.ParameterDict['6_PR_LifeTimeExtension_passvehicles'].Values[:,:,mS]),Par_RECC_ProductLifetime_p[:,:,SwitchTime -1::])
                
                # 2) Dynamic stock model
                # Build pdf array from lifetime distribution: Probability of survival.
                for p in tqdm(range(0, Np), unit=' vehicles types'):
                    for r in range(0, Nr):
                        LifeTimes = Par_RECC_ProductLifetime_p[p, r, :]
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.5 * LifeTimes} # flat decline: obsolescence of 16 % in 3 years around mean lifetime.
                        SF_Array[:, :, p, r] = dsm.DynamicStockModel(t=np.arange(0, Nc, 1), lt=lt).compute_sf().copy()
                        np.fill_diagonal(SF_Array[:, :, p, r],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
        
                # Compute evolution of 2015 in-use stocks: initial stock evolution separately from future stock demand and stock-driven model
                for r in range(0,Nr):   
                    FutureStock                 = np.zeros((Nc))
                    FutureStock[SwitchTime::]   = TotalStockCurves_UsePhase_p[1::, r].copy() # Future total stock
                    InitialStock                = TotalStock_UsePhase_Hist_cpr[:,:,r].copy()
                    InitialStocksum             = InitialStock.sum()
                    StockMatch_2015[Sector_pav_loc,r] = TotalStockCurves_UsePhase_p[0, r]/InitialStocksum
                    SFArrayCombined             = SF_Array[:,:,:,r]
                    TypeSplit                   = np.zeros((Nc,Np))
                    TypeSplit[SwitchTime::,:]   = RECC_System.ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[Sector_pav_loc,r,mR,:,1::].transpose() # indices: cp
                    
                    RECC_dsm                    = dsm.DynamicStockModel(t=np.arange(0,Nc,1), s=FutureStock.copy(), lt = lt)  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                    Var_S, Var_O, Var_I, IFlags = RECC_dsm.compute_stock_driven_model_initialstock_typesplit_negativeinflowcorrect(SwitchTime,InitialStock,SFArrayCombined,TypeSplit,NegativeInflowCorrect = True)
                    
                    # For a single-vehicle dynamic LCA, use this code block below to replace actual flows by single vehicle with fixed lifetime:
    #                VType = 4 # 0 for gasoline, 4 for BEV
    #                VLT   = 15 # fixed lifetime in years
    #                VTin  = 120 # year index of production, 120 stands for 2020.
    #                Var_I = np.zeros((Nc,Np))
    #                Var_I[VTin,VType] = 1
    #                Var_S = np.zeros((Nc,Nc,Np))
    #                Var_S[VTin:VTin+VLT,VTin,VType] = 1
    #                Var_O = np.zeros((Nc,Nc,Np))
    #                Var_O[VTin+VLT,VTin,VType] = 1
    #                InitialStock = np.zeros((Nc,Np))
                    
                    # Below, the results are added with += because the different commodity groups (buildings, vehicles) are calculated separately
                    # to introduce the type split for each, but using the product resolution of the full model with all sectors.
                    Stock_Detail_UsePhase_p[0,:,:,r]     += InitialStock.copy() # cgr, needed for correct calculation of mass balance later.
                    Stock_Detail_UsePhase_p[1::,:,:,r]   += Var_S[SwitchTime::,:,:].copy() # tcpr
                    Outflow_Detail_UsePhase_p[1::,:,:,r] += Var_O[SwitchTime::,:,:].copy() # tcpr
                    Inflow_Detail_UsePhase_p[1::,:,r]    += Var_I[SwitchTime::,:].copy() # tpr
                    Inflow_Prod_r[1::,r,Sector_pav_rge,mS,mR] = Var_I[SwitchTime::,:].copy()
                    # Check for negative inflows:
                    if IFlags.sum() != 0:
                        NegInflowFlags[Sector_pav_loc,mS,mR] = 1 # flag this scenario                
                    
                # Here so far: Units: Vehicles: million. for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_pav_loc,mS,mR] = TotalStockCurves_UsePhase_p.sum(axis =1).copy()
                StockCurves_Prod[:,Sector_pav_rge,mS,mR] = np.einsum('tcpr->tp',Stock_Detail_UsePhase_p).copy()
                pCStocksCurves[:,Sector_pav_loc,:,mS,mR] = TotalStockCurves_UsePhase_p_pC.copy()
                Population[:,:,mS,mR]                    = RECC_System.ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,:,:,mS]
                Inflow_Prod[:,Sector_pav_rge,mS,mR]      = np.einsum('tpr->tp',Inflow_Detail_UsePhase_p).copy()
                Outflow_Prod[:,Sector_pav_rge,mS,mR]     = np.einsum('tcpr->tp',Outflow_Detail_UsePhase_p).copy()
    
    
    
            if 'otv' in SectorList:
  
                
                s_otv = RECC_System.ParameterDict['2_S_RECC_FinalProducts_otherVehicles'].Values[:,:,:,:]
                
  
                
                Par_RECC_ProductLifetime_otv = RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_otherVehicles'].Values[:,:,:]
                
                lifetime_otv_c=np.zeros((Nv,No,Nc))
                lifetime_otv_c[:,:,:]=Par_RECC_ProductLifetime_otv[:,:,:]

                for i in range(0,100):
                   lifetime_otv_c[:,:,i]=Par_RECC_ProductLifetime_otv[:,:,101]
                
                SF_Array = np.zeros((Nc,Nc,Nv,No)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.            


                for v in tqdm(range(0, Nv), unit=' other vehicle types'):
                    for o in range(0, No):
                        # First, replicate lifetimes for all age-cohorts
                        ProductLifetime_otv = lifetime_otv_c[v, o, :] # Dimensions: 'voc'
                        lt = {'Type'  : 'Normal',
                              'Mean'  : ProductLifetime_otv, 
                              'StdDev': 0.3 * ProductLifetime_otv}
                # Compute evolution of nrbg in-use stock and related flows with stock-driven model
       
                        RECC_dsm_otv            = dsm.DynamicStockModel(time_dsm, s = s_otv[o,v,mS,mR,:].copy(), lt = lt)     
                        SF_Array[:, :,v, o]     = RECC_dsm_otv.compute_sf().copy()                             
                        np.fill_diagonal(SF_Array[:, :,v, o],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
                        RECC_dsm_otv.sf         = SF_Array[:, :, v, o].copy()
                        
                        otv_sc, otv_oc, otv_i = RECC_dsm_otv.compute_stock_driven_model(NegativeInflowCorrect = False) # Unit: Mm²
                        
                        Stock_Detail_UsePhase_otv[:,:,v,o]        = otv_sc[SwitchTime-1::,:].copy() # index structure: tcNo. Unit: million m².
                        Outflow_Detail_UsePhase_otv[1::,:,v,o]    = otv_oc[SwitchTime::,:].copy()   # index structure: tcNo. Unit: million m².
                        Inflow_Detail_UsePhase_otv[1::,v,o]       = otv_i[SwitchTime::].copy()      # index structure: tNo.  Unit: million m².     
                        
                # Here so far: Units: Buildings: million m2. for stocks, Mm² for flows.
                StockCurves_Totl[:,Sector_otv_loc,mS,mR] = np.einsum('tcNo->t', Stock_Detail_UsePhase_otv).copy()
                StockCurves_Prod[:,Sector_otv_rge,mS,mR] = np.einsum('tcNo->tN',Stock_Detail_UsePhase_otv).copy()
                pCStocksCurves[:,Sector_otv_loc,:,mS,mR] = 0  # pC stocks are not  consideredfor this sector/dataset
                Inflow_Prod[:,Sector_otv_rge,mS,mR]      = np.einsum('tNo->tN',Inflow_Detail_UsePhase_otv).copy()
                Outflow_Prod[:,Sector_otv_rge,mS,mR]     = np.einsum('tcNo->tN',Outflow_Detail_UsePhase_otv).copy()                             
 




            # Sector: Residential buildings
            if 'reb' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, residential buildings.')
                # 1) Determine total stock and apply stock-driven model
                SF_Array                    = np.zeros((Nc,Nc,NB,Nr)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                
                #Get historic stock at end of 2015 by age-cohort, and covert unit to Buildings: million m2.
                TotalStock_UsePhase_Hist_cBr = RECC_System.ParameterDict['2_S_RECC_FinalProducts_2015_resbuildings'].Values[0,:,:,:]
                
                # Determine total future stock, product level. Units: Buildings: million m2.
                TotalStockCurves_UsePhase_B_pC_test = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings'].Values[mS,:,Sector_reb_loc,:]
                    
                # 2) Include (or not) the RE strategies for the use phase:
                # Include_REStrategy_MoreIntenseUse:
                #if ScriptConfig['Include_REStrategy_MoreIntenseUse'] == 'True': 
                    # Calculate counter-factual scenario: X% decrease of stock levels by 2050 compared to scenario reference. X coded in parameter ..._MIUPotential
                    # if SName != 'LED':
                    #     RemainingFraction = 1-RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_MIUPotential'].Values[Sector_reb_loc,0,mS] / 100
                    #     #clamped_spline = make_interp_spline(np.arange(0,Nt,1), MIURamp, bc_type=([(2, 0)], [(1, 0)]))
                    #     clamped_spline = make_interp_spline([0,2,Nt-5,Nt], [1,1,RemainingFraction,RemainingFraction], bc_type=([(2, 0)], [(1, 0)]))
                    #     MIURamp_Spline = clamped_spline(np.arange(0,Nt,1))
                    #     MIURamp_Spline[MIURamp_Spline>1]=1
                    #     MIURamp_Spline[MIURamp_Spline<RemainingFraction]=RemainingFraction
                        
                    #     TotalStockCurves_UsePhase_B_pC_test    = TotalStockCurves_UsePhase_B_pC_test * np.einsum('t,r->tr',MIURamp_Spline,np.ones((Nr)))
                # Make sure that for no scenario, stock values are below LED values, which is assumed to be the lowest possible stock level.
                TotalStockCurves_UsePhase_B_pC_LED_ref = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings'].Values[LEDindex,:,Sector_reb_loc,:]
                TotalStockCurves_UsePhase_B_pC         = np.maximum(TotalStockCurves_UsePhase_B_pC_test,TotalStockCurves_UsePhase_B_pC_LED_ref)
                TotalStockCurves_UsePhase_B            = np.einsum('tr,tr->tr',TotalStockCurves_UsePhase_B_pC,RECC_System.ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,:,:,mS]) 
                RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'].Values[mS,:,Sector_reb_loc,:] = TotalStockCurves_UsePhase_B_pC.copy()
            
                # Include_REStrategy_LifeTimeExtension: Product lifetime extension.
                # First, replicate lifetimes for post 2020 age-cohorts from 2020 values as the parameter file only specifies values up to 2020:
                RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values[:,:,120::] = np.einsum('c,Br->Brc',np.ones((41)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values[:,:,120]).copy()
                Par_RECC_ProductLifetime_B = RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values.copy()
                # Second, change lifetime of future age-cohorts according to lifetime extension parameter
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
                    # option: future age-cohorts only, used in ODYM-RECC v2.2:
                    Par_RECC_ProductLifetime_B[:,:,SwitchTime -1::] = np.einsum('crB,Brc->Brc',1 + np.einsum('cr,Br->crB',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],beta * RECC_System.ParameterDict['6_PR_LifeTimeExtension_resbuildings'].Values[:,:,mS]),Par_RECC_ProductLifetime_B[:,:,SwitchTime -1::])
                  
                    


                # 3) Dynamic stock model, with lifetime depending on age-cohort.
                # Build pdf array from lifetime distribution: Probability of survival.
                for B in tqdm(range(0, NB), unit=' res. building types'):
                    for r in range(0, Nr):
                        LifeTimes = Par_RECC_ProductLifetime_B[B, r, :]
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.3 * LifeTimes}
                        SF_Array[:, :, B, r] = dsm.DynamicStockModel(t=np.arange(0, Nc, 1), lt=lt).compute_sf().copy()
                        np.fill_diagonal(SF_Array[:, :, B, r],1) # no outflows from current year, 
                        # this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
        
                # Compute evolution of 2015 in-use stocks: initial stock evolution separately from future stock demand and stock-driven model
                for r in range(0,Nr):   
                    FutureStock                 = np.zeros((Nc))
                    FutureStock[SwitchTime::]   = TotalStockCurves_UsePhase_B[1::, r].copy()# Future total stock
                    InitialStock                = TotalStock_UsePhase_Hist_cBr[:,:,r].copy()
                    InitialStocksum             = InitialStock.sum()
                    StockMatch_2015[Sector_reb_loc,r] = TotalStockCurves_UsePhase_B[0, r]/InitialStocksum
                    SFArrayCombined             = SF_Array[:,:,:,r]
                    TypeSplit                   = np.zeros((Nc,NB))
                    TypeSplit[SwitchTime::,:]   = RECC_System.ParameterDict['3_SHA_TypeSplit_Buildings'].Values[:,r,1::,mS].transpose() # indices: Bc
                    
                    RECC_dsm                    = dsm.DynamicStockModel(t=np.arange(0,Nc,1), s=FutureStock.copy(), lt = lt)  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                    Var_S, Var_O, Var_I, IFlags = RECC_dsm.compute_stock_driven_model_initialstock_typesplit_negativeinflowcorrect(SwitchTime,InitialStock,SFArrayCombined,TypeSplit,NegativeInflowCorrect = True)
                    
                    # Below, the results are added with += because the different commodity groups (buildings, vehicles) are calculated separately
                    # to introduce the type split for each, but using the product resolution of the full model with all sectors.
                    Stock_Detail_UsePhase_B[0,:,:,r]     += InitialStock.copy() # cgr, needed for correct calculation of mass balance later.
                    Stock_Detail_UsePhase_B[1::,:,:,r]   += Var_S[SwitchTime::,:,:].copy() # tcBr
                    Outflow_Detail_UsePhase_B[1::,:,:,r] += Var_O[SwitchTime::,:,:].copy() # tcBr
                    Inflow_Detail_UsePhase_B[1::,:,r]    += Var_I[SwitchTime::,:].copy() # tBr
                    # Check for negative inflows:
                    if IFlags.sum() != 0:
                        NegInflowFlags[Sector_reb_loc,mS,mR] = 1 # flag this scenario
        
                # Here so far: Units: Buildings: million m². for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_reb_loc,mS,mR] = TotalStockCurves_UsePhase_B.sum(axis =1).copy()
                StockCurves_Prod[:,Sector_reb_rge,mS,mR] = np.einsum('tcBr->tB',Stock_Detail_UsePhase_B).copy()
                pCStocksCurves[:,Sector_reb_loc,:,mS,mR] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'].Values[mS,:,Sector_reb_loc,:].copy()
                Population[:,:,mS,mR]                    = RECC_System.ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,:,:,mS]
                Inflow_Prod[:,Sector_reb_rge,mS,mR]      = np.einsum('tBr->tB',Inflow_Detail_UsePhase_B).copy()
                Outflow_Prod[:,Sector_reb_rge,mS,mR]     = np.einsum('tcBr->tB',Outflow_Detail_UsePhase_B).copy()
                
                
                
                
                
                
                
                
                # Include renovation
                RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,:,:,mS] = np.einsum('cmBr,t->mBrct',RECC_System.ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,:,:,:,mS],np.ones(Nt)) # mBrctS
                if ScriptConfig['Include_Renovation_reb'] == 'True' and ScriptConfig['No_EE_Improvements'] == 'False': 
                    RenPot_E   = np.einsum('rcB,rB->rcB',RECC_System.ParameterDict['3_SHA_MaxRenovationPotential_ResBuildings'].Values[:,0:SwitchTime,:],RECC_System.ParameterDict['3_SHA_EnergySavingsPot_Renovation_ResBuildings'].Values[:,mS,:]) # Unit: 1
                    RenPot_E_t = np.einsum('tr,rcB->trcB',RECC_System.ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values[:,:,mS,mR],RenPot_E) # Unit: 1, Defined as share of stock crB that is renovated by year t * energy saving potential
                    RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,:,:,:] = np.einsum('cBVnr,trcB->cBVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:SwitchTime,:,:,:,:,mS],(np.ones((Nt,Nr,Nc-Nt+1,NB))-RenPot_E_t)) # cBVnrt
                    # Add renovation material intensity to building material intensity:
                    RenPot_M_t = np.einsum('tr,rcB->trcB',RECC_System.ParameterDict['3_SHA_BuildingRenovationScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['3_SHA_MaxRenovationPotential_ResBuildings'].Values[:,0:SwitchTime,:]) # Unit: 1, Defined as share of stock crB that is renovated by year t
                    MC_Ren = RECC_System.ParameterDict['3_MC_RECC_Buildings_RECC'].Values[:,:,:,:,mS]*RECC_System.ParameterDict['3_MC_RECC_Buildings_Renovation_Relative'].Values + RECC_System.ParameterDict['3_MC_RECC_Buildings_Renovation_Absolute'].Values
                    RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,0:SwitchTime,:,mS] += np.einsum('cmBr,trcB->mBrct',MC_Ren[0:SwitchTime,:,:,:],RenPot_M_t)
                else:
                    RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,:,:,:] = np.einsum('cBVnr,trcB->cBVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[0:SwitchTime,:,:,:,:,mS],np.ones((Nt,Nr,Nc-Nt+1,NB))) # cBVnrt
                RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime-1::,:,:,:,:,:]   = np.einsum('cBVnr,t->cBVnrt',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchTime-1::,:,:,:,:,mS],np.ones(Nt)) # future age-cohorts
                
   


                            
            # Sector: Nonresidential buildings, global total
            if 'nrbg' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, nonresidential buildings.')
                SF_Array = np.zeros((Nc,Nc,NN,No)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.            
                s_nrbg   = RECC_System.ParameterDict['2_S_RECC_FinalProducts_nonresbuildings_g'].Values[:,:]  ### dimensions: 'Nt'
    
    

                Par_RECC_ProductLifetime_nrbg = np.zeros((NN,No,Nc))
                
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True': # for future age-cohorts t = c
                    Par_RECC_ProductLifetime_nrbg[:,:,SwitchTime-1:] = np.einsum('aot,aot->aot',RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values[:,:,SwitchTime -1::],1 + np.einsum('ao,ot->aot',beta * RECC_System.ParameterDict['6_PR_LifeTimeExtension_nonresbuildings_g'].Values[:,:],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS]))
            
                else:   
                    Par_RECC_ProductLifetime_nrbg = RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values

                        

                print(RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values[:,:,160])

                for N in tqdm(range(0, NN), unit='Mm²'):
                    for o in range(0, No):
                        # First, replicate lifetimes for all age-cohorts
                        ProductLifetime_nrbg = Par_RECC_ProductLifetime_nrbg[N,o,:] # Dimensions: 'Noc'
                        lt = {'Type'  : 'Normal',
                              'Mean'  : ProductLifetime_nrbg, 
                              'StdDev': 0.3 * ProductLifetime_nrbg}
                # Compute evolution of nrbg in-use stock and related flows with stock-driven model
       
                        RECC_dsm_nrbg            = dsm.DynamicStockModel(time_dsm, s = s_nrbg[N,:].copy(), lt = lt)     
                        SF_Array[:, :, N, o]     = RECC_dsm_nrbg.compute_sf().copy()                             
                        np.fill_diagonal(SF_Array[:, :, N, o],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
                        RECC_dsm_nrbg.sf         = SF_Array[:, :, N, o].copy()
                        
                        nrbg_sc, nrbg_oc, nrbg_i = RECC_dsm_nrbg.compute_stock_driven_model(NegativeInflowCorrect = False) # Unit: Mm²
                        
                        Stock_Detail_UsePhase_Ng[:,:,N,o]        = nrbg_sc[SwitchTime-1::,:].copy() # index structure: tcNo. Unit: million m².
                        Outflow_Detail_UsePhase_Ng[1::,:,N,o]    = nrbg_oc[SwitchTime::,:].copy()   # index structure: tcNo. Unit: million m².
                        Inflow_Detail_UsePhase_Ng[1::,N,o]       = nrbg_i[SwitchTime::].copy()      # index structure: tNo.  Unit: million m².     
                        
                # Here so far: Units: Buildings: million m2. for stocks, Mm² for flows.
                StockCurves_Totl[:,Sector_nrbg_loc,mS,mR] = np.einsum('tcNo->t', Stock_Detail_UsePhase_Ng).copy()
                StockCurves_Prod[:,Sector_nrbg_rge,mS,mR] = np.einsum('tcNo->tN',Stock_Detail_UsePhase_Ng).copy()
                pCStocksCurves[:,Sector_nrbg_loc,:,mS,mR] = 0  # pC stocks are not considered for this sector/dataset
                Inflow_Prod[:,Sector_nrbg_rge,mS,mR]      = np.einsum('tNo->tN',Inflow_Detail_UsePhase_Ng).copy()
                Outflow_Prod[:,Sector_nrbg_rge,mS,mR]     = np.einsum('tcNo->tN',Outflow_Detail_UsePhase_Ng).copy()                    
                
            # Sector: Industry, 11 region and global coverage, will be calculated separately and waste will be added to wast mgt. inflow for 1st region.
            if 'ind' in SectorList:
                Mylog.info('Calculate inflows and outflows for use phase, industry.')
                # 1) Determine total stock and apply stock-driven model
                
                SF_Array                    = np.zeros((Nc,Nc,NI,Nl)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.

                s_ind = RECC_System.ParameterDict['2_S_RECC_FinalProducts_industry'].Values[:,:,:,:,:]                     ### dimensions: rSRpt of TotalFutureInflow_UsePhase_ind
    
                # set lifetime parameter
                # First, Simply replicate lifetimes for all age-cohorts
                Par_RECC_ProductLifetime_ind = np.einsum('cl,I->Ilc',np.ones((Nc,Nl)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_industry'].Values)
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True': # for future age-cohorts t = c
                    Par_RECC_ProductLifetime_ind[:,:,SwitchTime -1::] = np.einsum('Ilt,Ilt->Ilt',Par_RECC_ProductLifetime_ind[:,:,SwitchTime -1::],1 + np.einsum('Il,t->Ilt',beta * RECC_System.ParameterDict['6_PR_LifeTimeExtension_industry'].Values[:,:,mS],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,0,:,mS]))
                    
                # Dynamic stock model
                # Build pdf array from lifetime distribution: Probability of survival.
                RECC_dsm_ind_s_c        = np.zeros((Nl,NS,NR,NI,Nc,Nc))
                RECC_dsm_ind_s_c_o_c    = np.zeros((Nl,NS,NR,NI,Nc,Nc))
                RECC_dsm_ind_o          = np.zeros((Nl,NS,NR,NI,Nc))
                Inflow                  = np.zeros((Nt,Nl,NI))
                
                TotalStockCurves_UsePhase_I = np.zeros((Nt,NI,Nl))
                            
                            
                            
                for I in tqdm(range(0, NI), unit='units'):
                    for l in range(0, Nl):
                                LifeTimes = Par_RECC_ProductLifetime_ind[I, l, :]
                            
                                lt = {'Type'  : 'Normal',
                                      'Mean'  : LifeTimes,
                                      'StdDev': 0.3 * LifeTimes}
                # Compute inflow-driven model
                                RECC_dsm_ind                         = dsm.DynamicStockModel(time_dsm , s = s_ind[l,mS,mR,I,:].copy()  , lt = lt)
                                SF_Array[:, :, I, l]                 = RECC_dsm_ind.compute_sf().copy()  # The lt parameter is not used, the sf array is handed over directly in the next step.   
                                np.fill_diagonal(SF_Array[:, :, I, l],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                                # Those parts of the stock remain in use instead.
                                RECC_dsm_ind.sf         = SF_Array[:, :, I, l].copy()
    
    
                                ind_sc, ind_oc, ind_i = RECC_dsm_ind.compute_stock_driven_model(NegativeInflowCorrect = False) # Unit: Mm²
    
                                print(ind_sc.max())
     
                                Stock_Detail_UsePhase_I[:,:,I,l]     = ind_sc[SwitchTime-1::,:].copy() 
                                Outflow_Detail_UsePhase_I[:,:,I,l]   = ind_oc[SwitchTime-1::,:].copy() 
                                Outflow_Detail_UsePhase_I[0,:,I,l]   = 0 # no flow calculation in first year
                                Inflow_Detail_UsePhase_I[:,I,l]      = ind_i[SwitchTime-1::].copy()  # index structure: tIl
                                Inflow_Detail_UsePhase_I[0,I,l]      = 0 # no flow calculation in first year
      
    
                                TotalStockCurves_UsePhase_I[:,I,l] = Stock_Detail_UsePhase_I[:,:,I,l].sum(axis=1) 
                               
                                
                               
                                
                # Here so far: Units: Electricity: GW. for stocks, X/yr for flows.
                StockCurves_Totl[:,Sector_ind_loc,mS,mR] = TotalStockCurves_UsePhase_I[:,:,:].sum(axis=1).sum(axis=1).copy()
                StockCurves_Prod[:,Sector_ind_rge,mS,mR] = TotalStockCurves_UsePhase_I[:,:,:].sum(axis=2).copy()
                Inflow_Prod[:,Sector_ind_rge,mS,mR]      = np.einsum('tIl->tI',Inflow_Detail_UsePhase_I).copy()
                Outflow_Prod[:,Sector_ind_rge,mS,mR]     = np.einsum('tcIl->tI',Outflow_Detail_UsePhase_I).copy()                      
            
     
               
            # Sector: Appliances, global coverage, will be calculated separately and waste will be added to wast mgt. inflow for 1st region.
            if 'app' in SectorList:            
                Mylog.info('Calculate inflows and outflows for use phase, appliances.')
                
                SF_Array     = np.zeros((Nc,Nc,Na,No)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
          
            
            
            
                s_app = RECC_System.ParameterDict['2_S_RECC_FinalProducts_appliances'].Values[:,:,:,:,:] # dimensions:oaSRc

            
                # set lifetime parameter
                # First, Simply replicate lifetimes for all age-cohorts
                Par_RECC_ProductLifetime_app = np.einsum('co,a->aoc',np.ones((Nc,No)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_appliances'].Values)
                if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True': # for future age-cohorts t = c
                    Par_RECC_ProductLifetime_app[:,:,SwitchTime -1::] = np.einsum('aot,aot->aot',Par_RECC_ProductLifetime_app[:,:,SwitchTime -1::],1 + np.einsum('ao,ot->aot',beta * RECC_System.ParameterDict['6_PR_LifeTimeExtension_appliances'].Values[:,:,mS],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS]))
   

                if ScriptConfig['Include_RES_GoodToService'] == 'True': # for future age-cohorts t = c
                    s_app[:,:,mS,mR,SwitchTime -1::] = np.einsum('oat,aot->oat', s_app[:,:,mS,mR,SwitchTime -1::], (1-np.einsum('ao,ot->aot',RECC_System.ParameterDict['6_PR_Appliance_use_reduction'].Values[:,:,mS],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS])))


                     
            
                # Dynamic stock model
                # Build pdf array from lifetime distribution: Probability of survival.
                RECC_dsm_app_s_c        = np.zeros((No,NS,NR,Na,Nc,Nc))
                RECC_dsm_app_s_c_o_c    = np.zeros((No,NS,NR,Na,Nc,Nc))
                RECC_dsm_app_o          = np.zeros((No,NS,NR,Na,Nc))
                            
                TotalStockCurves_UsePhase_a = np.zeros((Nt,Na,No))
                
            
                
                for a in tqdm(range(0, Na), unit=' App types'):
                    for o in range(0, No):
                        LifeTimes = Par_RECC_ProductLifetime_app[a, o, :]
                    
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.3 * LifeTimes}
      
                # Compute inflow-driven model     
                        RECC_dsm_app                         = dsm.DynamicStockModel(time_dsm , s = s_app[o,a,mS,mR,:].copy()  , lt = lt)  
                        SF_Array[:, :, a, o]                 = RECC_dsm_app.compute_sf().copy() 
                        np.fill_diagonal(SF_Array[:, :, a, o],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
                                
                        RECC_dsm_app.sf                      = SF_Array[:, :, a, o].copy()
                       
                        
                        app_sc, app_oc, app_i      = RECC_dsm_app.compute_stock_driven_model(NegativeInflowCorrect = True)

                        
                       
                        RECC_dsm_app_s_c[o,mS,mR,a,:,:]      = RECC_dsm_app.compute_s_c_inflow_driven()
                        RECC_dsm_app_s_c_o_c[o,mS,mR,a,:,:]  = RECC_dsm_app.compute_o_c_from_s_c()
                        RECC_dsm_app_o[o,mS,mR,a,:]          = RECC_dsm_app.compute_outflow_total()
    
                        Stock_Detail_UsePhase_a[:,:,a,o]     =  app_sc[SwitchTime-1::,:]
                        
                        Outflow_Detail_UsePhase_a[:,:,a,o]   = app_oc[SwitchTime-1::,:]
                        Outflow_Detail_UsePhase_a[0,:,a,o]   = 0 # no flow calculation in first year
                        Inflow_Detail_UsePhase_a[:,a,o]      = app_i[SwitchTime-1::] # index structure: tIl
                        Inflow_Detail_UsePhase_a[0,a,o]      = 0 # no flow calculation in first year
    
     
                # So far: Units: 1 (# items). for stocks, X/yr for flows.

                TotalStockCurves_UsePhase_a[:,:,:]           = Stock_Detail_UsePhase_a[:,:,:,:].sum(axis=1) 
                StockCurves_Totl[:,Sector_app_loc,mS,mR]     = TotalStockCurves_UsePhase_a[:,:,:].sum(axis=1).sum(axis=1).copy()
                StockCurves_Prod[:,Sector_app_rge,mS,mR]     = TotalStockCurves_UsePhase_a[:,:,:].sum(axis=2).copy()
                Inflow_Prod[:,Sector_app_rge,mS,mR]          = np.einsum('tIl->tI',Inflow_Detail_UsePhase_a).copy()
                Outflow_Prod[:,Sector_app_rge,mS,mR]         = np.einsum('tcIl->tI',Outflow_Detail_UsePhase_a).copy()   
    
    
    
    
    

    
    
     # Sector: Infrastructure, based on Schipper et al. (2018)
            if 'inf' in SectorList:            
                Mylog.info('Calculate inflows and outflows for use phase, infrastructure.')
                
           
                
                
                SF_Array     = np.zeros((Nc,Nc,Ni,No)) # survival functions, by year, age-cohort, good, and region. PDFs are stored externally because recreating them with scipy.stats is slow.
                s_inf = RECC_System.ParameterDict['2_S_RECC_FinalProducts_infrastructure'].Values[:,:,:,:,:] # dimensions:SRoic # unit km for train lines
    
   
    
                # set lifetime parameter
                # First, Simply replicate lifetimes for all age-cohorts
                Par_RECC_ProductLifetime_inf = np.einsum('co,i->ioc',np.ones((Nc,No)),RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_infrastructure'].Values)
         
                # Dynamic stock model
                # Build pdf array from lifetime distribution: Probability of survival.
                RECC_dsm_inf_s_c        = np.zeros((No,NS,NR,Ni,Nc,Nc))
                RECC_dsm_inf_s_c_o_c    = np.zeros((No,NS,NR,Ni,Nc,Nc))
                RECC_dsm_inf_o          = np.zeros((No,NS,NR,Ni,Nc))
                            
                TotalStockCurves_UsePhase_i = np.zeros((Nt,Ni,Nl))
                     
                
                for i in tqdm(range(0, Ni), unit=' inf types'):
                    for o in range(0, No):
                        LifeTimes = Par_RECC_ProductLifetime_inf[i, o, :]
                    
                        lt = {'Type'  : 'Normal',
                              'Mean'  : LifeTimes,
                              'StdDev': 0.3 * LifeTimes}
      
                     # Compute evolution of nrbg in-use stock and related flows with stock-driven model
       
                        RECC_dsm_inf            = dsm.DynamicStockModel(time_dsm, s = s_inf[mS,mR,o,i,:].copy(), lt = lt)     
                        SF_Array[:, :, i, o]     = RECC_dsm_inf.compute_sf().copy()                             
                        np.fill_diagonal(SF_Array[:, :, i, o],1) # no outflows from current year, this would break the mass balance in the calculation routine below, as the element composition of the current year is not yet known.
                        # Those parts of the stock remain in use instead.
                        RECC_dsm_inf.sf         = SF_Array[:, :, i, o].copy()
    
                        inf_sc, inf_oc, inf_i      = RECC_dsm_inf.compute_stock_driven_model(NegativeInflowCorrect = True)
                        
                        RECC_dsm_inf_s_c[o,mS,mR,i,:,:]      = RECC_dsm_inf.compute_s_c_inflow_driven()
                        RECC_dsm_inf_s_c_o_c[o,mS,mR,i,:,:]  = RECC_dsm_inf.compute_o_c_from_s_c()
                        RECC_dsm_inf_o[o,mS,mR,i,:]          = RECC_dsm_inf.compute_outflow_total()
                
    
                        Stock_Detail_UsePhase_i[:,:,i,o]     = inf_sc[SwitchTime-1::,:]
                        Outflow_Detail_UsePhase_i[:,:,i,o]   = inf_oc[SwitchTime-1::,:]
                        Outflow_Detail_UsePhase_i[0,:,i,o]   = 0 # no flow calculation in first year
                        Inflow_Detail_UsePhase_i[:,i,o]      = inf_i[SwitchTime-1::] # index structure: tIl
                        Inflow_Detail_UsePhase_i[0,i,o]      = 0 # no flow calculation in first year
                
     
    
                       
                StockCurves_Totl[:,Sector_inf_loc,mS,mR] = np.einsum('tcNo->t', Stock_Detail_UsePhase_i).copy()
                StockCurves_Prod[:,Sector_inf_rge,mS,mR] = np.einsum('tcNo->tN',Stock_Detail_UsePhase_i).copy()
                pCStocksCurves[:,Sector_inf_loc,:,mS,mR] = 0  # pC stocks are not considered for this sector/dataset
                Inflow_Prod[:,Sector_inf_rge,mS,mR]      = np.einsum('tNo->tN',Inflow_Detail_UsePhase_i).copy()
                Outflow_Prod[:,Sector_inf_rge,mS,mR]     = np.einsum('tcNo->tN',Outflow_Detail_UsePhase_i).copy() 
                
                
                
                TotalStockCurves_UsePhase_i = Stock_Detail_UsePhase_i[:,:,:,:].sum(axis=1) 
     
            
            # Archive 2015 pC stock values for future curves:
            pC_FutureStock_2015             = np.zeros((NS,NG,Nr))
            # b) from scenario curves:
            for mSS in range(0,NS):
                if 'pav' in SectorList:
                    pC_FutureStock_2015[mSS, Sector_pav_loc, :] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_passvehicles'].Values[mSS,0,Sector_pav_loc,:]
                if 'reb' in SectorList:
                    pC_FutureStock_2015[mSS, Sector_reb_loc, :] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_resbuildings_act'].Values[mSS,0,Sector_reb_loc,:]
                if 'nrb' in SectorList:
                    pC_FutureStock_2015[mSS, Sector_nrb_loc, :] = RECC_System.ParameterDict['2_S_RECC_FinalProducts_Future_NonResBuildings_act'].Values[Sector_nrb_loc,:,0,mSS]
            OutputDict['pC_FutureStock_2015']   = pC_FutureStock_2015.copy()
    
    
    
            # ABOVE: each sector separate, individual regional resolution. BELOW: all sectors together, global total.
            ## No / o: regional aggregation level =1 (for app, nrbg, inf and otv (part of transpoort)); Nl / l = regional aggregation level = 11 (for ind); Nr / r = regional aggregation level = 20 (for reb and pav (part of transport))
            # Prepare parameters:        
            # include light-weighting in future MC parameter, cmgr
            Par_RECC_MC_Nr = np.zeros((Nc,Nm,Ng,Nr,NS,Nt))  # Unit: vehicles: kg/item, buildings: kg/m².
            Par_RECC_MC_Nl = np.zeros((Nc,Nm,NL,Nl,NS))          # for electricity generation technologies in kt/GW
            Par_RECC_MC_No = np.zeros((Nc,Nm,NO,No,NS))          # for appliances in g/unit, nonres. buildings in kg/m²
    
            if 'pav' in SectorList:
                Par_RECC_MC_Nr[:,:,Sector_pav_rge,:,mS,:]      = np.einsum('cmpr,t->pcmrt',RECC_System.ParameterDict['3_MC_RECC_Vehicles_RECC'].Values[:,:,:,:,mS],np.ones(Nt))
            if 'reb' in SectorList:
                Par_RECC_MC_Nr[:,:,Sector_reb_rge,:,mS,:]      = np.einsum('mBrct->Bcmrt',RECC_System.ParameterDict['3_MC_RECC_Buildings_t'].Values[:,:,:,:,:,mS])
            if 'nrb' in SectorList: 
                Par_RECC_MC_Nr[:,:,Sector_nrb_rge,:,mS,:]      = np.einsum('cmNr,t->Ncmrt',RECC_System.ParameterDict['3_MC_RECC_NonResBuildings_RECC'].Values[:,:,:,:,mS],np.ones(Nt))
            if 'ind' in SectorList: 
                Par_RECC_MC_Nl[:,:,Sector_ind_rge_reg,:,mS]        = np.einsum('lc,Im->Icml',np.ones((Nl,Nc)), RECC_System.ParameterDict['3_MC_RECC_industry'].Values[:,:])       #3_MC_RECC_industry has dimensions Im
            if 'app' in SectorList: 
                Par_RECC_MC_No[:,:,Sector_app_rge_reg,:,mS]        = np.einsum('c,oOm->Ocmo',np.ones((Nc)), RECC_System.ParameterDict['3_MC_RECC_appliances'].Values[:,:,:])      #3_MC_RECC_appliances has dimensions oam
            if 'nrbg' in SectorList:
                Par_RECC_MC_No[:,:,Sector_nrbg_rge_reg,:,mS]       = np.einsum('c,o,mN->Ncmo',np.ones((Nc)), np.ones((No)), RECC_System.ParameterDict['3_MC_RECC_Nonresbuildings_g'].Values[:,:])  #3_MC_RECC_Nonresbuildings_g has dimensions mN
            if 'inf' in SectorList: 
                Par_RECC_MC_No[:,:,Sector_inf_rge_reg,:,mS]       = np.einsum('c,o,im->icmo',np.ones((Nc)), np.ones((No)), RECC_System.ParameterDict['3_MC_RECC_infrastructure'].Values[:,:])  #3_MC_RECC_infrastructure
            if 'otv' in SectorList: 
                Par_RECC_MC_No[:,:,Sector_otv_rge_reg,:,mS]       = np.einsum('c,ovm->vcmo',np.ones((Nc)),  RECC_System.ParameterDict['3_MC_RECC_otherVehicles'].Values[:,:,:])  #3_MC_RECC_infrastructure

             

          # Units: Vehicles: kg/unit, Buildings: kg/m2  
            
            # historic element composition of materials:
            Par_Element_Composition_of_Materials_m   = np.zeros((Nc,Nm,Ne)) # Unit: 1. Aspects: cme, produced in age-cohort c. Applies to new manufactured goods.
            Par_Element_Composition_of_Materials_m[0:Nc-Nt+1,:,:] = np.einsum('c,me->cme',np.ones(Nc-Nt+1),RECC_System.ParameterDict['3_MC_Elements_Materials_ExistingStock'].Values)
            # For future age-cohorts, the total is known but the element breakdown of this parameter will be updated year by year in the loop below.
            Par_Element_Composition_of_Materials_m[:,:,0] = 1 # element 0 is 'all', for which the mass share is always 100%.
            
            # future element composition of materials inflow use phase (mix new and reused products)
            Par_Element_Composition_of_Materials_c   = np.zeros((Nt,Nm,Ne)) # cme, produced in age-cohort c. Applies to new manufactured goods.
            
            # Element composition of material in the use phase
            Par_Element_Composition_of_Materials_u   = Par_Element_Composition_of_Materials_m.copy() # cme
            
            # Manufacturing yield and other improvements
            # Reduce cement content by up to percentage indicate in 3_SHA_CementContentReduction parameter:
            if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'True':
                Par_RECC_MC_Nr[115::,Cement_loc,:,:,mS,:] = Par_RECC_MC_Nr[115::,Cement_loc,:,:,mS,:] * (1 - RECC_System.ParameterDict['3_SHA_CementContentReduction'].Values[Cement_loc] * np.einsum('oc,gt->cgot',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS],np.ones((Ng,Nt)))).copy()
    
            Par_FabYieldLoss = np.einsum('mwggto->mwgto',RECC_System.ParameterDict['4_PY_Manufacturing'].Values) # take diagonal of product = manufacturing process
            
            # Consider Fabrication yield improvement and reduction in cement content of concrete and plaster
            if ScriptConfig['Include_REStrategy_FabYieldImprovement'] == 'True':
                Par_FabYieldImprovement = np.einsum('w,tmgo->mwgto',np.ones((Nw)),np.einsum('ot,mgo->tmgo',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS],RECC_System.ParameterDict['6_PR_FabricationYieldImprovement'].Values[:,:,:,mS]))
            else:
                Par_FabYieldImprovement = 0              
                
            Par_FabYieldLoss_Raster = Par_FabYieldLoss > 0    
            Par_FabYieldLoss        = Par_FabYieldLoss - Par_FabYieldLoss_Raster * Par_FabYieldImprovement #mwgto
            Par_FabYieldLoss_total  = np.einsum('mwgto->mgto',Par_FabYieldLoss)
            Divisor                 = 1-Par_FabYieldLoss_total
            Par_FabYield_total_inv  = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # mgto
            # Determine total element composition of products (c: age-cohort), needs to be updated for future age-cohorts, is done below after material cycle computation.
            Par_3_MC_Stock_ByElement_Nr = np.einsum('cmgrt,cme->tcrgme',Par_RECC_MC_Nr[:,:,:,:,mS,:],Par_Element_Composition_of_Materials_m) # Unit: vehicles: kg/item, buildings: kg/m².
            Par_3_MC_Stock_ByElement_Nl = np.einsum('cmLl,cme->clLme',Par_RECC_MC_Nl[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: ind: kt/GW
            Par_3_MC_Stock_ByElement_No = np.einsum('cmOo,cme->coOme',Par_RECC_MC_No[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: app: g/unit, nrbg: kg/m²
            # Consider EoL recovery rate improvement:
            if ScriptConfig['Include_REStrategy_EoL_RR_Improvement'] == 'True':           
                Par_RECC_EoL_RR =  np.einsum('r,t,gmw->trmgw',np.ones((Nr)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[:,0,:,:,0] *0.01) \
                + np.einsum('tr,grmw->trmgw',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_EoL_RR_Improvement'].Values[:,:,:,:,0]*0.01)
            else:    
                Par_RECC_EoL_RR =  np.einsum('t,grmw->trmgw',np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[:,:,:,:,0] *0.01)
            
            # For regional dimension 11 and 1
           
            if 'ind' in SectorList:
                Par_RECC_EoL_RR_Nl = np.einsum('l,t,Lmw->tlmLw',np.ones((Nl)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_11reg_rge,0,:,:,0] *0.01)
              
            if 'inf' or 'app' or 'nrbg' or 'otv' in SectorList: 
                Par_RECC_EoL_RR_No = np.einsum('o,t,Omw->tomOw',np.ones((No)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_1reg_rge,0,:,:,0] *0.01)
           
            if 'pav' or 'reb' in SectorList:
                Par_RECC_EoL_RR_Nr = np.einsum('o,t,Omw->tomOw',np.ones((Nr)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_32reg_rge,0,:,:,0] *0.01)
    
            SHA_REStrategyScaleUp_No  = RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:No,mS,mR] 
    
            SHA_REStrategyScaleUp_Nl  = RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:Nl,mS,mR] 
    
    
    
           # add here sorting rate improvement
            if ScriptConfig['Include_RES_ImproveSortingRate'] == 'True':
               
                  if 'ind' in SectorList:
                      Par_RECC_EoL_RR_Nl = np.einsum('r,t,gmw->trmgw',np.ones((Nl)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_11reg_rge,0,:,:,0] *0.01) \
                          + np.einsum('tr,grmw->trmgw',SHA_REStrategyScaleUp_Nl[:,:],RECC_System.ParameterDict['6_PR_EoL_RR_Improvement'].Values[Sector_11reg_rge,:,:,:,0]*0.01)
              
                    
                      for r in range(0,Nl):
                        for t in range(1,Nt):
                            for m in range(1,Nm):
                                for g in range(1,NI):
                                    for w in range(1,Nw):
                                        if Par_RECC_EoL_RR_Nl[t,r,m,g,w]>1:
                                            Par_RECC_EoL_RR_Nl[t,r,m,g,w]=1
                
                  if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:        
    
                      Par_RECC_EoL_RR_No = np.einsum('r,t,gmw->trmgw',np.ones((No)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_1reg_rge,0,:,:,0] *0.01) \
                          + np.einsum('tr,grmw->trmgw',SHA_REStrategyScaleUp_No[:,:],RECC_System.ParameterDict['6_PR_EoL_RR_Improvement'].Values[Sector_1reg_rge,:,:,:,0]*0.01)
                  
    
                      for r in range(0,No):
                        for t in range(1,Nt):
                            for m in range(1,Nm):
                                for g in range(1,NO):
                                    for w in range(1,Nw):
                                        if Par_RECC_EoL_RR_No[t,r,m,g,w]>1:
                                            Par_RECC_EoL_RR_No[t,r,m,g,w]=1
    
                  
                  if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
                      Par_RECC_EoL_RR_Nr = np.einsum('r,t,gmw->trmgw',np.ones((Nr)),np.ones((Nt)),RECC_System.ParameterDict['4_PY_EoL_RecoveryRate'].Values[Sector_32reg_rge,0,:,:,0] *0.01) \
                          + np.einsum('tr,grmw->trmgw',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_EoL_RR_Improvement'].Values[Sector_32reg_rge,:,:,:,0]*0.01)
                  
      
                      for r in range(0,Nr):
                        for t in range(1,Nt):
                            for m in range(1,Nm):
                                for g in range(1,NT):
                                    for w in range(1,Nw):
                                        if Par_RECC_EoL_RR_Nr[t,r,m,g,w]>1:
                                            Par_RECC_EoL_RR_Nr[t,r,m,g,w]=1
    
                    
    
    
            # Calculate reuse factor
            # For vehicles, scenarios already included in target table output!
            ReUseFactor_tmprS = np.einsum('mprtS->tmprS',RECC_System.ParameterDict['6_PR_ReUse_Veh'].Values/100)
            # For Buildings, scenarios obtained from RES scaleup curves
            ReUseFactor_tmBrS = np.einsum('tmBr,S->tmBrS',np.einsum('tr,mBr->tmBr',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_ReUse_Bld'].Values),np.ones((NS)))
            ReUseFactor_tmNrS = np.einsum('tmNr,S->tmNrS',np.einsum('tr,mNr->tmNr',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp_r'].Values[:,:,mS,mR],RECC_System.ParameterDict['6_PR_ReUse_nonresBld'].Values),np.ones((NS)))
            
               
            ReUseFactor_tmaS = np.einsum('matS->tmaS',RECC_System.ParameterDict['6_PR_ReUse_app'].Values/100)
            ReUseFactor_tmiS = np.einsum('mitS->tmiS',RECC_System.ParameterDict['6_PR_ReUse_inf'].Values/100)
            ReUseFactor_tmIS = np.einsum('mItS->tmIS',RECC_System.ParameterDict['6_PR_ReUse_ind'].Values/100)
            ReUseFactor_tmvS = np.einsum('mvtS->tmvS',RECC_System.ParameterDict['6_PR_ReUse_otv'].Values/100)

    
    
            plt.plot(np.arange(0,46)+2015,ReUseFactor_tmBrS[:,6,:,:,2].mean(axis=1).mean(axis=1))
            plt.plot(np.arange(0,46)+2015,ReUseFactor_tmprS[:,6,:,:,2].mean(axis=1).mean(axis=1))
            plt.plot(np.arange(0,46)+2015,ReUseFactor_tmNrS[:,6,:,:,2].mean(axis=1).mean(axis=1))
            plt.plot(np.arange(0,46)+2015,ReUseFactor_tmaS[:,6,:,2].mean(axis=1))
            plt.plot(np.arange(0,46)+2015,ReUseFactor_tmiS[:,6,:,2].mean(axis=1))
            plt.plot(np.arange(0,46)+2015,ReUseFactor_tmIS[:,6,:,2].mean(axis=1))
            
            Mylog.info('Translate total flows into individual materials and elements, for 2015 and historic age-cohorts.')
            # Determine total element composition of products (c: age-cohort), needs to be updated for future age-cohorts, is done below after material cycle computation.
            Par_3_MC_Stock_ByElement_Nr = np.einsum('cmgrt,cme->tcrgme',Par_RECC_MC_Nr[:,:,:,:,mS,:],Par_Element_Composition_of_Materials_m) # Unit: vehicles: kg/item, buildings: kg/m².
            Par_3_MC_Stock_ByElement_Nl = np.einsum('cmLl,cme->clLme',Par_RECC_MC_Nl[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: ind: kt/GW
            Par_3_MC_Stock_ByElement_No = np.einsum('cmOo,cme->coOme',Par_RECC_MC_No[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: app: g/unit, nrbg: kg/m², inf:kg/kWh
            # Consider EoL recovery rate improvement:
            
            if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
                        # Determine total element composition of products (c: age-cohort), needs to be updated for future age-cohorts, is done below after material cycle computation.
                        Par_3_MC_Stock_ByElement_Nr = np.einsum('cmgrt,cme->tcrgme',Par_RECC_MC_Nr[:,:,:,:,mS,:],Par_Element_Composition_of_Materials_m) # Unit: vehicles: kg/item, buildings: kg/m².
     
            
                        if 'pav' in SectorList:
                            # convert product stocks and flows to material stocks and flows, only for chemical element position 'all':
                            # Stock elemental composition, will be updated for future years:
                            RECC_System.StockDict['S_7_Nr'].Values[:,:,:,Sector_pav_rge_reg,:,:] = \
                            np.einsum('tcrpme,tcpr->tcrpme',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_pav_rge_reg,:,:],Stock_Detail_UsePhase_p)/1000   # Indices='t,c,r,p,m,e'
                            # Outflow, 'all' elements only:
                            RECC_System.FlowDict['F_7_8_Nr'].Values[:,:,:,Sector_pav_rge_reg,:,0] = \
                            np.einsum('pcmrt,tcpr->ptcrm',Par_RECC_MC_Nr[:,:,Sector_pav_rge_reg,:,mS,:],Outflow_Detail_UsePhase_p)/1000 # all elements, Indices='t,c,r,p,m'
                            # Inflow as mass balance, to account for renovation material inflows to other age-cohorts than the current one (t=c).
                            RECC_System.FlowDict['F_6_7_Nr'].Values[1::,:,Sector_pav_rge_reg,:,0]   = \
                            np.einsum('ptcrm->ptrm',np.diff(RECC_System.StockDict['S_7'].Values[:,:,:,Sector_pav_rge,:,0],1,axis=1)) + np.einsum('ptcrm->ptrm',RECC_System.FlowDict['F_7_8_Nr'].Values[1::,:,:,Sector_pav_rge_reg,:,0])
                            # inflow of materials in new products, for checking:
                            for mmt in range(0,Nt):
                                F_6_7_new[mmt,:,Sector_pav_rge_reg,:,0] = np.einsum('pr,pmr->prm',Inflow_Detail_UsePhase_p[mmt,:,:],Par_RECC_MC_Nr[SwitchTime+mmt-1,:,Sector_pav_rge_reg,:,mS,mmt])/1000
                            # Check_pav = (RECC_System.FlowDict['F_6_7_Nr'].Values[1::,0,Sector_pav_rge,:,0] - F_6_7_new[1::,0,Sector_pav_rge,:,0]).sum() # must be 0.
                                
                        if 'reb' in SectorList:        
                            # convert product stocks and flows to material stocks and flows, only for chemical element position 'all':
                            # Stock elemental composition, historic for each element and for future years: 'all' elements only
                            RECC_System.StockDict['S_7_Nr'].Values[:,:,:,Sector_reb_rge_reg,:,:] = \
                            np.einsum('tcrBme,tcBr->tcrBme',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge_reg,:,:],Stock_Detail_UsePhase_B)/1000   # Indices='t,c,r,B,m'
                            # Outflow, 'all' elements only:
                            ####why is this different htan in the other sectors?!?
    
                            RECC_System.FlowDict['F_7_8_Nr'].Values[:,:,:,Sector_reb_rge_reg,:,0] = \
                            np.einsum('Btcrm,tcBr->Btcrm',Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge,:,0],Outflow_Detail_UsePhase_B)/1000                        # Inflow of renovation material as stock multiplied with change in material composition:
                       
                    
                       
                        
                            F_6_7_ren[1::,:,:,Sector_reb_rge,:,0]  = np.einsum('tcBr,Btcrm->Btcrm',Stock_Detail_UsePhase_B[1::,:,:,:],np.diff(Par_3_MC_Stock_ByElement_Nr[:,:,:,Sector_reb_rge,:,0],1,axis=1))/1000
                            # inflow of materials in new products
                            for mmt in range(0,Nt):
                                F_6_7_new[mmt,:,Sector_reb_rge,:,0] = np.einsum('Br,Brm->Brm',Inflow_Detail_UsePhase_B[mmt,:,:],Par_3_MC_Stock_ByElement_Nr[mmt,SwitchTime+mmt-1,:,Sector_reb_rge,:,0])/1000
                            # Check_reb = (RECC_System.FlowDict['F_6_7_Nr'].Values[1::,0,Sector_reb_rge,:,0] - F_6_7_new[1::,0,Sector_reb_rge,:,0] - F_6_7_ren[1::,:,0,Sector_reb_rge,:,0].sum(axis=2)) # must be 0.
                            RECC_System.FlowDict['F_6_7_Nr'].Values[:,:,Sector_reb_rge,:,0]   = np.einsum('Btrm->Btrm',F_6_7_new[:,:,Sector_reb_rge,:,0]) + np.einsum('Btcrm->Btrm',F_6_7_ren[:,:,:,Sector_reb_rge,:,0])
                          
            # 1_Nl_No) Inflow, outflow and stock first year for Nl and No regional aggregation and Sector I and a
     
            if 'ind' in SectorList:
                Par_3_MC_Stock_ByElement_Nl = np.einsum('cmLl,cme->clLme',Par_RECC_MC_Nl[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: ind: kt/GW
    
                RECC_System.FlowDict['F_6_7_Nl'].Values[0,:,Sector_ind_rge_reg,:,:]   = \
                np.einsum('Ilme,Il ->Ilme',Par_3_MC_Stock_ByElement_Nl[SwitchTime-1,:,Sector_ind_rge_reg,:,:],Inflow_Detail_UsePhase_I[0,:,:])/1000 # all elements, Indices='t,l,I,m,e'  
                RECC_System.FlowDict['F_6_7_Nl'].Values[1::,:,Sector_ind_rge_reg,:,0]   = \
                np.einsum('Itlm,tIl->Itlm',Par_3_MC_Stock_ByElement_Nl[SwitchTime::,:,Sector_ind_rge_reg,:,0],Inflow_Detail_UsePhase_I[1::,:,:])/1000 # all elements, Indices='t,l,I,m'  
                RECC_System.FlowDict['F_7_8_Nl'].Values[0,0:SwitchTime,:,Sector_ind_rge_reg,:,:] = \
                np.einsum('clIme,cpl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:SwitchTime,:,Sector_ind_rge_reg,:,:],Outflow_Detail_UsePhase_I[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,l,I,m,e'
                RECC_System.StockDict['S_7_Nl'].Values[0,0:SwitchTime,:,Sector_ind_rge_reg,:,:] = \
                np.einsum('clIme,cIl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:SwitchTime,:,Sector_ind_rge_reg,:,:],Stock_Detail_UsePhase_I[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,l,p,m,e'
      
            
    
    
            if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:        
                Par_3_MC_Stock_ByElement_No = np.einsum('cmOo,cme->coOme',Par_RECC_MC_No[:,:,:,:,mS],    Par_Element_Composition_of_Materials_m) # Unit: app: g/unit, nrbg: kg/m², inf:kg/kWh
                
                if 'app' in SectorList:    
                        RECC_System.FlowDict['F_6_7_No'].Values[0,:,Sector_app_rge_reg,:,:]   = \
                        np.einsum('aome,ao->aome',Par_3_MC_Stock_ByElement_No[SwitchTime-1,:,Sector_app_rge_reg,:,:],Inflow_Detail_UsePhase_a[0,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'  
                        RECC_System.FlowDict['F_6_7_No'].Values[1::,:,Sector_app_rge_reg,:,0]   = \
                        np.einsum('atom,tao->atom',Par_3_MC_Stock_ByElement_No[SwitchTime::,:,Sector_app_rge_reg,:,0],Inflow_Detail_UsePhase_a[1::,:,:])/1000000000000 # all elements, Indices='t,o,a,m'  
                        RECC_System.FlowDict['F_7_8_No'].Values[0,0:SwitchTime,:,Sector_app_rge_reg,:,:] = \
                        np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_app_rge_reg,:,:],Outflow_Detail_UsePhase_a[0,0:SwitchTime,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'
                        RECC_System.StockDict['S_7_No'].Values[0,0:SwitchTime,:,Sector_app_rge_reg,:,:] = \
                        np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_app_rge_reg,:,:],Stock_Detail_UsePhase_a[0,0:SwitchTime,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'
            
                if 'nrbg' in SectorList:
                        RECC_System.FlowDict['F_6_7_No'].Values[0,:,Sector_nrbg_rge_reg,:,:]   = \
                        np.einsum('Nome,No->Nome',Par_3_MC_Stock_ByElement_No[SwitchTime-1,:,Sector_nrbg_rge_reg,:,:],Inflow_Detail_UsePhase_Ng[0,:,:])/1000 # all elements, Indices='t,o,N,m,e'  
                        RECC_System.FlowDict['F_6_7_No'].Values[1::,:,Sector_nrbg_rge_reg,:,0]   = \
                        np.einsum('Ntom,tNo->Ntom',Par_3_MC_Stock_ByElement_No[SwitchTime::,:,Sector_nrbg_rge_reg,:,0],Inflow_Detail_UsePhase_Ng[1::,:,:])/1000 # all elements, Indices='t,o,N,m'              
                        RECC_System.StockDict['S_7_No'].Values[0,0:SwitchTime,:,Sector_nrbg_rge_reg,:,:] = \
                        np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_nrbg_rge_reg,:,:],Stock_Detail_UsePhase_Ng[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,o,N,m,e'
                        RECC_System.FlowDict['F_7_8_No'].Values[0,0:SwitchTime,:,Sector_nrbg_rge_reg,:,:] = \
                        np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_nrbg_rge_reg,:,:],Outflow_Detail_UsePhase_Ng[0,0:SwitchTime,:,:])/1000 # all elements, Indices='t,o,N,m,e'
                   
                        
                if 'inf' in SectorList:
                        RECC_System.FlowDict['F_6_7_No'].Values[0,:,Sector_inf_rge_reg,:,:]   = \
                        np.einsum('iome,io->iome',Par_3_MC_Stock_ByElement_No[SwitchTime-1,:,Sector_inf_rge_reg,:,:],Inflow_Detail_UsePhase_i[0,:,:])      
                        RECC_System.FlowDict['F_6_7_No'].Values[1::,:,Sector_inf_rge_reg,:,0]   = \
                        np.einsum('itom,tio->itom',Par_3_MC_Stock_ByElement_No[SwitchTime::,:,Sector_inf_rge_reg,:,0],Inflow_Detail_UsePhase_i[1::,:,:]) # all elements, Indices='t,o,a,m'  
                        RECC_System.FlowDict['F_7_8_No'].Values[0,0:SwitchTime,:,Sector_inf_rge_reg,:,:] = \
                        np.einsum('coime,cio->icome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_inf_rge_reg,:,:],Outflow_Detail_UsePhase_i[0,0:SwitchTime,:,:]) # all elements, Indices='t,o,a,m,e'
                        RECC_System.StockDict['S_7_No'].Values[0,0:SwitchTime,:,Sector_inf_rge_reg,:,:] = \
                        np.einsum('coime,cio->icome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_inf_rge_reg,:,:],Stock_Detail_UsePhase_i[0,0:SwitchTime,:,:]) # all elements, Indices='t,o,a,m,e'
     
                        
                if 'otv' in SectorList:
                        RECC_System.FlowDict['F_6_7_No'].Values[0,:,Sector_otv_rge_reg,:,:]   = \
                        np.einsum('iome,io->iome',Par_3_MC_Stock_ByElement_No[SwitchTime-1,:,Sector_otv_rge_reg,:,:],Inflow_Detail_UsePhase_otv[0,:,:])/1000000000
                        RECC_System.FlowDict['F_6_7_No'].Values[1::,:,Sector_otv_rge_reg,:,0]   = \
                        np.einsum('itom,tio->itom',Par_3_MC_Stock_ByElement_No[SwitchTime::,:,Sector_otv_rge_reg,:,0],Inflow_Detail_UsePhase_otv[1::,:,:])/1000000000 # all elements, Indices='t,o,a,m'  
                        RECC_System.FlowDict['F_7_8_No'].Values[0,0:SwitchTime,:,Sector_otv_rge_reg,:,:] = \
                        np.einsum('coime,cio->icome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_otv_rge_reg,:,:],Outflow_Detail_UsePhase_otv[0,0:SwitchTime,:,:])/1000000000 # all elements, Indices='t,o,a,m,e'
                        RECC_System.StockDict['S_7_No'].Values[0,0:SwitchTime,:,Sector_otv_rge_reg,:,:] = \
                        np.einsum('coime,cio->icome',Par_3_MC_Stock_ByElement_No[0:SwitchTime,:,Sector_otv_rge_reg,:,:],Stock_Detail_UsePhase_otv[0,0:SwitchTime,:,:])/1000000000 # all elements, Indices='t,o,a,m,e'
     
                
     
        
             
            #Units so far: Mt/yr
            Mylog.info('Calculate material stocks and flows, material cycles, determine elemental composition.')
            # Units: Mt and Mt/yr.
            # This calculation is done year-by-year, and the elemental composition of the materials is in part determined by the scrap flow metal composition
            
          
            Par_RECC_CollectionRate =  np.zeros((Ng,Nt))
            
            
            
            
            if ScriptConfig['Include_RES_ImproveCollectionRate'] == 'True': 
    
            
                    Par_RECC_CollectionRate[:,:] = np.einsum('gt,gt-> gt',RECC_System.ParameterDict['4_PY_CollectionRate'].Values[:,:], 1  + np.einsum('go,ot-> gt',alpha * RECC_System.ParameterDict['6_PR_Collection_Improvement'].Values[:,:,mS],RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,:,mS]))
             
                    for l in range(0,len(RECC_System.ParameterDict['4_PY_CollectionRate'].Values[:,0])):
                        for t in range(1,Nt):
                        
                            if Par_RECC_CollectionRate[l,t]>100:
                                Par_RECC_CollectionRate[l,t]=100
            
            
            else:  
                Par_RECC_CollectionRate  = RECC_System.ParameterDict['4_PY_CollectionRate'].Values[:,:]

            
          
    
            
            
    
    
            
            for t in tqdm(range(1,Nt), unit=' years'):  # 1: 2016
                CohortOffset = t +Nc -Nt # index of current age-cohort.   
                # First, before going down to the material layer, we consider obsolete stock formation and re-use.
                        
                
                # 1) Convert use phase outflow to system variables.
                # Split flows into materials and chemical elements.
                # Calculate use phase outflow and obsolete stock formation
                # ObsStockFormation = ObsStockFormationFactor(t,g,r) * Outflow_Detail_UsePhase(t,c,g,r), currently not implemented. 
                
                if 'pav' in SectorList:
                    RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge_reg,:,:] = \
                    np.einsum('pcrme,cpr->pcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_pav_rge_reg,:,:],Outflow_Detail_UsePhase_p[t,0:CohortOffset,:,:])/1000 # All elements.
                
                    RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge_reg,:,:]= np.einsum('p,pcrme->pcrme',0.01*RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_pav_rge], RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge_reg,:,:] )
              
                    RECC_System.FlowDict['F_8_19_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge_reg,:,:]= np.einsum('p,pcrme->pcrme',0.01*(100-Par_RECC_CollectionRate[Sector_pav_rge_reg,t]), RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge_reg,:,:] -  RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge_reg,:,:])
    
                
                
                if 'reb' in SectorList:
                    RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:] = \
                    np.einsum('Bcrme,cBr->Bcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge_reg,:,:],Outflow_Detail_UsePhase_B[t,0:CohortOffset,:,:])/1000 # All elements.
              
                
                    RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:]= np.einsum('B,Bcrme->Bcrme',0.01*RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_reb_rge], RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:] )
    
                
                    RECC_System.FlowDict['F_8_19_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:]= np.einsum('B,Bcrme->Bcrme',0.01*(100-Par_RECC_CollectionRate[Sector_reb_rge,t]), RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:] -  RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:])
    
                
    #            if 'nrb' in SectorList:
    #                RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,:] = \
    #                np.einsum('Ncrme,cNr->Ncrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_nrb_rge,:,:],Outflow_Detail_UsePhase_N[t,0:CohortOffset,:,:])/1000 # All elements.
    #
    #
    #                RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,:]= np.einsum('N,Ncrme->Ncrme',RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_nrb_rge], RECC_System.FlowDict['F_7_8'].Values[t,0:CohortOffset,:,Sector_nrb_rge,:,:] )
    
    
                # 1_Nl_No)
                if 'ind' in SectorList:          
                    RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:] = \
                    np.einsum('clIme,cIl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:CohortOffset,:,Sector_ind_rge_reg,:,:],Outflow_Detail_UsePhase_I[t,0:CohortOffset,:,:])/1000 # All elements.
                
                
                    RECC_System.FlowDict['F_8_11_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:]= np.einsum('I,Icrme->Icrme',0.01*RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_ind_rge_reg], RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:] )
    
                    RECC_System.FlowDict['F_8_19_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:]= np.einsum('B,Bcrme->Bcrme',0.01*(100-Par_RECC_CollectionRate[Sector_ind_rge_reg,t]), RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:] -  RECC_System.FlowDict['F_8_11_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:])
    
                if 'app' in SectorList:          
                    RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:] = \
                    np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:CohortOffset,:,Sector_app_rge_reg,:,:],Outflow_Detail_UsePhase_a[t,0:CohortOffset,:,:])/1000000000000 # All elements.
               
                    RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:]= np.einsum('a,acrme->acrme',0.01*RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_app_rge_reg], RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:] )
    
                    RECC_System.FlowDict['F_8_19_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:]= np.einsum('a,acrme->acrme',0.01*(100-Par_RECC_CollectionRate[Sector_app_rge_reg,t]),( RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:] -  RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:]))
    
                if 'nrbg' in SectorList:
                    RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:] = \
                    np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:CohortOffset,:,Sector_nrbg_rge_reg,:,:],Outflow_Detail_UsePhase_Ng[t,0:CohortOffset,:,:])/1000 # All elements.
                
                    RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:]= np.einsum('N,Ncrme->Ncrme',0.01*RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_nrbg_rge_reg], RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:] )
    
                    RECC_System.FlowDict['F_8_19_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:]= np.einsum('N,Ncrme->Ncrme',0.01*(100-Par_RECC_CollectionRate[Sector_nrbg_rge_reg,t]), (RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:] -  RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:]))
    
                
                if 'inf' in SectorList:
                    RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,:] = \
                    np.einsum('coime,cio->icome',Par_3_MC_Stock_ByElement_No[0:CohortOffset,:,Sector_inf_rge_reg,:,:],Outflow_Detail_UsePhase_i[t,0:CohortOffset,:,:]) # all elements, Indices='t,o,a,m,e'
        
                    RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,:]= np.einsum('i,icrme->icrme',0.01*RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_inf_rge_reg], RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,:] )
    
                    RECC_System.FlowDict['F_8_19_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,:]= np.einsum('i,icrme->icrme',0.01*(100-Par_RECC_CollectionRate[Sector_inf_rge_reg,t]), (RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,:] -  RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,:]))
                 
                    
                 
                if 'otv' in SectorList:
                    RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,:] = \
                    np.einsum('coime,cio->icome',Par_3_MC_Stock_ByElement_No[0:CohortOffset,:,Sector_otv_rge_reg,:,:],Outflow_Detail_UsePhase_otv[t,0:CohortOffset,:,:]/1000000000) # all elements, Indices='t,o,a,m,e'
        
                    RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,:]= np.einsum('i,icrme->icrme',0.01*RECC_System.ParameterDict['4_PY_ObsoleteStocks'].Values[Sector_otv_rge_reg], RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,:] )
    
                    RECC_System.FlowDict['F_8_19_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,:]= np.einsum('i,icrme->icrme',0.01*(100-Par_RECC_CollectionRate[Sector_otv_rge_reg,t]), (RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,:] -  RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,:]))
      
                            
                # 2) Consider re-use of materials in product groups (via components), as ReUseFactor(m,g,r,R,t) * RECC_System.FlowDict['F_7_8'].Values(t,c,r,g,m,e)
                # Distribute material for re-use onto product groups
                if 'pav' in SectorList:
                    ReUsePotential_Materials_t_m_Veh = np.einsum('mpr,pcrm->m',ReUseFactor_tmprS[t,:,:,:,mS],RECC_System.FlowDict['F_7_8_Nr'].Values[t,:,:,Sector_pav_rge_reg,:,0]) # in Mt
                if 'reb' in SectorList:
                    ReUsePotential_Materials_t_m_Bld = np.einsum('mBr,Bcrm->m',ReUseFactor_tmBrS[t,:,:,:,mS],RECC_System.FlowDict['F_7_8_Nr'].Values[t,:,:,Sector_reb_rge_reg,:,0]) # in Mt
    #            if 'nrb' in SectorList:
    #                ReUsePotential_Materials_t_m_NRB = np.einsum('mNr,Ncrm->m',ReUseFactor_tmNrS[t,:,:,:,mS],RECC_System.FlowDict['F_7_8_Nr'].Values[t,:,:,Sector_nrb_rge,:,0]) # in Mt
    #     
    #        
            #### this need to be replicated for app, inf, nrbg and ind
                if 'app' in SectorList:
                    ReUsePotential_Materials_t_m_app = np.einsum('ma,acrm->m',ReUseFactor_tmaS[t,:,:,mS],RECC_System.FlowDict['F_7_8_No'].Values[t,:,:,Sector_app_rge_reg,:,0]) # in Mt
                if 'inf' in SectorList:
                    ReUsePotential_Materials_t_m_inf = np.einsum('mi,icrm->m',ReUseFactor_tmiS[t,:,:,mS],RECC_System.FlowDict['F_7_8_No'].Values[t,:,:,Sector_inf_rge_reg,:,0]) # in Mt
                if 'nrbg' in SectorList:
                    ReUsePotential_Materials_t_m_nrbg = np.einsum('mB,Bcm->m',ReUseFactor_tmNrS[t,:,:,0,mS],RECC_System.FlowDict['F_7_8_No'].Values[t,:,0,Sector_nrbg_rge_reg,:,0]) # in Mt
                if 'ind' in SectorList:                                                                                      
                    ReUsePotential_Materials_t_m_ind = np.einsum('mI,Icrm->m',ReUseFactor_tmIS[t,:,:,mS],RECC_System.FlowDict['F_7_8_Nl'].Values[t,:,:,Sector_ind_rge_reg,:,0]) # in Mt
                if 'otv' in SectorList:                                                                                      
                    ReUsePotential_Materials_t_m_otv = np.einsum('mI,Icrm->m',ReUseFactor_tmvS[t,:,:,mS],RECC_System.FlowDict['F_7_8_No'].Values[t,:,:,Sector_otv_rge_reg,:,0]) # in Mt
                 
            
            # in the future, re-use will be a region-to-region parameter depicting, e.g., the export of used vehicles from the EU to Africa.
                # check whether inflow is big enough for potential to be used, correct otherwise:
                for mmm in range(0,Nm):
                    # Vehicles
                    if 'pav' in SectorList:
                        if RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_pav_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_Veh[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_Veh[mmm] > 0:
                                ReUsePotential_Materials_t_m_Veh[mmm] = RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_pav_rge,mmm,0].sum()
                    # residential buildings
                    if 'reb' in SectorList:
                        if RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_reb_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_Bld[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_Bld[mmm] > 0:
                                ReUsePotential_Materials_t_m_Bld[mmm] = RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_reb_rge,mmm,0].sum()
                    # nonresidential buildings
    #                if 'nrb' in SectorList:
    #                    if RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_nrb_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_NRB[mmm]: # if re-use potential is larger than new inflow:
    #                        if ReUsePotential_Materials_t_m_NRB[mmm] > 0:
    #                            ReUsePotential_Materials_t_m_NRB[mmm] = RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_nrb_rge,mmm,0].sum()
    #                #appliances
                    if 'app' in SectorList:
                        if RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_app_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_app[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_app[mmm] > 0:
                                ReUsePotential_Materials_t_m_app[mmm] = RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_app_rge,mmm,0].sum()
                               
                        #industry
                    if 'app' in SectorList:
                        if RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_app_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_app[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_app[mmm] > 0:
                                ReUsePotential_Materials_t_m_app[mmm] = RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_app_rge,mmm,0].sum()
                                             
                   #infrastructure
                    if 'inf' in SectorList:
                        if RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_inf_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_inf[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_inf[mmm] > 0:
                                ReUsePotential_Materials_t_m_inf[mmm] = RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_inf_rge,mmm,0].sum()
                                      
                    if 'otv' in SectorList:
                        if RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_otv_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_otv[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_otv[mmm] > 0:
                                ReUsePotential_Materials_t_m_otv[mmm] = RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_otv_rge,mmm,0].sum()
                                
                                
                        #nrbg
                    if 'nrbg' in SectorList:
                        if RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_nrbg_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_nrbg[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_nrbg[mmm] > 0:
                                ReUsePotential_Materials_t_m_nrbg[mmm] = RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_nrbg_rge,mmm,0].sum()
                              
        
                    # industry
                    if 'ind' in SectorList:
                        if RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,Sector_ind_rge,mmm,0].sum() < ReUsePotential_Materials_t_m_ind[mmm]: # if re-use potential is larger than new inflow:
                            if ReUsePotential_Materials_t_m_ind[mmm] > 0:
                                ReUsePotential_Materials_t_m_ind[mmm] = RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,Sector_ind_rge,mmm,0].sum()
                              
        
    
    
    
                # Vehicles
                if 'pav' in SectorList:
                    Divisor = np.einsum('m,crp->pcrm',np.einsum('pcrm->m',RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,0]-np.einsum('pcrme -> pcrm', RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,:]-RECC_System.FlowDict['F_8_19_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,:])),np.ones((CohortOffset,Nr,Np)))
                    MassShareVeh = np.divide(RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: pcrm
                                  
                    
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17_Nr'].Values[t,0:CohortOffset,:,Sector_pav_rge_reg,:,:] = \
                    np.einsum('cme,pcrm->pcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,pcrm->pcrm',ReUsePotential_Materials_t_m_Veh,MassShareVeh))  # All elements.
                # residential Buildings
                if 'reb' in SectorList:
                    Divisor = np.einsum('m,crB->Bcrm',np.einsum('Bcrm->m',RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,0]-np.einsum('Bcrme -> Bcrm', RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:]-RECC_System.FlowDict['F_8_19_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge_reg,:,:])),np.ones((CohortOffset,Nr,NB)))
                    MassShareBld = np.divide(RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Bcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17_Nr'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,:] = \
                    np.einsum('cme,Bcrm->Bcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Bcrm->Bcrm',ReUsePotential_Materials_t_m_Bld,MassShareBld))  # All elements.
                    
      
    
                if 'app' in SectorList:
                    Divisor = np.einsum('m,cra->acrm',np.einsum('acrm->m',RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,0]),np.ones((CohortOffset,No,Na)))
                    MassShareapp = np.divide(RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Bcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17_No'].Values[t,0:CohortOffset,:,Sector_app_rge_reg,:,:] = \
                    np.einsum('cme,Bcrm->Bcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Bcrm->Bcrm',ReUsePotential_Materials_t_m_app,MassShareapp))  # All elements.
    
    
                if 'inf' in SectorList:
                    Divisor = np.einsum('m,cra->acrm',np.einsum('acrm->m',RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,0]),np.ones((CohortOffset,No,Ni)))
                    MassShareinf = np.divide(RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Bcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17_No'].Values[t,0:CohortOffset,:,Sector_inf_rge_reg,:,:] = \
                    np.einsum('cme,Bcrm->Bcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Bcrm->Bcrm',ReUsePotential_Materials_t_m_inf,MassShareinf))  # All elements.
 
    
                if 'otv' in SectorList:
                    Divisor = np.einsum('m,cra->acrm',np.einsum('acrm->m',RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,0]),np.ones((CohortOffset,No,Nv)))
                    MassShareinf = np.divide(RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Bcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17_No'].Values[t,0:CohortOffset,:,Sector_otv_rge_reg,:,:] = \
                    np.einsum('cme,Bcrm->Bcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Bcrm->Bcrm',ReUsePotential_Materials_t_m_otv,MassShareinf))  # All elements.
        

    
                if 'nrbg' in SectorList:
                    Divisor = np.einsum('m,cra->acrm',np.einsum('acrm->m',RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,0]),np.ones((CohortOffset,No,NN)))
                    MassSharenrbg = np.divide(RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Bcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17_No'].Values[t,0:CohortOffset,:,Sector_nrbg_rge_reg,:,:] = \
                    np.einsum('cme,Bcrm->Bcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Bcrm->Bcrm',ReUsePotential_Materials_t_m_nrbg,MassSharenrbg))  # All elements.
    
    
                if 'ind' in SectorList:
                    Divisor = np.einsum('m,cra->acrm',np.einsum('acrm->m',RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,0]),np.ones((CohortOffset,Nl,NI)))
                    MassShareind = np.divide(RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,0], Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) # index: Bcrm
                    # share of combination crg in total mass of m in outflow 7_8
                    RECC_System.FlowDict['F_8_17_Nl'].Values[t,0:CohortOffset,:,Sector_ind_rge_reg,:,:] = \
                    np.einsum('cme,Bcrm->Bcrme', Par_Element_Composition_of_Materials_u[0:CohortOffset,:,:],\
                    np.einsum('m,Bcrm->Bcrm',ReUsePotential_Materials_t_m_ind,MassShareind))  # All elements.
    
    
    
                # reused material mapped to final consumption region and good, proportional to final consumption breakdown into products and regions.
                # can be replaced by region-by-region reuse parameter.             
                    
                if 'reb' in SectorList or 'pav' in SectorList:
                    Divisor = np.einsum('m,rg->rgm',np.einsum('rgm->m',RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,:,:,0]),np.ones((Nr,NT)))
                    InvMass = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
                    RECC_System.FlowDict['F_17_6_Nr'].Values[t,0:CohortOffset,:,:,:,:] = \
                    np.einsum('cme,rgm->crgme',np.einsum('crgme->cme',RECC_System.FlowDict['F_8_17_Nr'].Values[t,0:CohortOffset,:,:,:,:]),\
                    RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,:,:,0]*InvMass)
                    
       
    
    
    
    
                 
                    
                if 'app' in SectorList or 'inf' in SectorList or 'nrbg' in SectorList or 'otv' in SectorList:
                    Divisor = np.einsum('m,rg->rgm',np.einsum('rgm->m',RECC_System.FlowDict['F_6_7_No'].Values[t,:,:,:,0]),np.ones((No,NO)))
                    InvMass = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
                    RECC_System.FlowDict['F_17_6_No'].Values[t,0:CohortOffset,:,:,:,:] = \
                    np.einsum('cme,rgm->crgme',np.einsum('crgme->cme',RECC_System.FlowDict['F_8_17_No'].Values[t,0:CohortOffset,:,:,:,:]),\
                    RECC_System.FlowDict['F_6_7_No'].Values[t,:,:,:,0]*InvMass)
                    
                    
                    
                if 'ind' in SectorList:
                    Divisor = np.einsum('m,rg->rgm',np.einsum('rgm->m',RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,:,:,0]),np.ones((Nl,NL)))
                    InvMass = np.divide(1, Divisor, out=np.zeros_like(Divisor), where=Divisor!=0)
                    RECC_System.FlowDict['F_17_6_Nl'].Values[t,0:CohortOffset,:,:,:,:] = \
                    np.einsum('cme,rgm->crgme',np.einsum('crgme->cme',RECC_System.FlowDict['F_8_17_Nl'].Values[t,0:CohortOffset,:,:,:,:]),\
                    RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,:,:,0]*InvMass)
                  
                # 3) calculate inflow waste mgt as EoL products - obsolete stock formation - re-use
               
                
                if len(Sector_32reg_rge) > 0:
                    RECC_System.FlowDict['F_8_9_Nr'].Values[t,:,:,:,:]           = np.einsum('crgme->rgme',RECC_System.FlowDict['F_7_8_Nr'].Values[t,0:CohortOffset,:,:,:,:]    - RECC_System.FlowDict['F_8_11_Nr'].Values[t,0:CohortOffset,:,:,:,:]    - RECC_System.FlowDict['F_8_17_Nr'].Values[t,0:CohortOffset,:,:,:,:]- RECC_System.FlowDict['F_8_19_Nr'].Values[t,0:CohortOffset,:,:,:,:])
                if len(Sector_11reg_rge) > 0:
                    RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:]    = np.einsum('clLme->lLme',RECC_System.FlowDict['F_7_8_Nl'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_11_Nl'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_17_Nl'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_19_Nl'].Values[t,0:CohortOffset,:,:,:,:])
                if len(Sector_1reg_rge) > 0:
                    RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:]    = np.einsum('coOme->oOme',RECC_System.FlowDict['F_7_8_No'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_11_No'].Values[t,0:CohortOffset,:,:,:,:] - RECC_System.FlowDict['F_8_17_No'].Values[t,0:CohortOffset,:,:,:,:]- RECC_System.FlowDict['F_8_19_No'].Values[t,0:CohortOffset,:,:,:,:])
    
    
    
    
            
                # 4) EoL products to postconsumer scrap: trwe. Add Waste mgt. losses.
                if len(Sector_32reg_rge) > 0:
                    RECC_System.FlowDict['F_9_10_Nr'].Values[t,:,:,:]            = np.einsum('rmgw,rgme->rwe',Par_RECC_EoL_RR_Nr[t,:,:,:,:],RECC_System.FlowDict['F_8_9_Nr'].Values[t,:,:,:,:])
                   # RECC_System.FlowDict['F_14_13_Nr'].Values[t,:,:,:]          =   np.einsum('rmgw,rgme->gwe',(1-Par_RECC_EoL_RR[t,:,:,:,:]),RECC_System.FlowDict['F_8_9_Nr'].Values[t,:,:,:,:])
                    
                  
                    RECC_System.FlowDict['F_14_13_Nr'].Values[t,:,Cu_scrap_loc,0] = (1-Par_RECC_EoL_RR_Nr[t,0,6,:,5])*RECC_System.FlowDict['F_8_9_Nr'].Values[t,0,:,6,0]
    
                    RECC_System.FlowDict['F_14_13'].Values[t,Sector_32reg_rge,:,:]   =   np.einsum('gwe->gwe',RECC_System.FlowDict['F_14_13_Nr'].Values[t,:,:,:])
                    
                    
                    
                    
                if len(Sector_11reg_rge) > 0:                    
                    RECC_System.FlowDict['F_9_10_Nl'].Values[t,:,:,:]     = np.einsum('lmLw,lLme->lwe',Par_RECC_EoL_RR_Nl[t,:,:,:,:],RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:])    
                   # RECC_System.FlowDict['F_14_13_Nl'].Values[t,:,:,:]     = np.einsum('lmLw,lLme->Lwe',(1-Par_RECC_EoL_RR_Nl[t,:,:,:,:]),RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:])    
                   
                    
                    RECC_System.FlowDict['F_14_13_Nl'].Values[t,:,Cu_scrap_loc,0] = (1-Par_RECC_EoL_RR_Nl[t,0,6,:,5])*RECC_System.FlowDict['F_8_9_Nl'].Values[t,0,:,6,0]
    
                    
                    RECC_System.FlowDict['F_14_13'].Values[t,Sector_11reg_rge,:,:]   =   np.einsum('Lwe->Lwe',RECC_System.FlowDict['F_14_13_Nl'].Values[t,:,:,:])
    
                
                
                
                if len(Sector_1reg_rge) > 0:            
                    RECC_System.FlowDict['F_9_10_No'].Values[t,:,:,:]     = np.einsum('omOw,oOme->owe',Par_RECC_EoL_RR_No[t,:,:,:,:],RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:])    
                  #  RECC_System.FlowDict['F_14_13_No'].Values[t,:,:,:]     = np.einsum('omOw,oOme->Owe',(1-Par_RECC_EoL_RR_No[t,:,:,:,:]),RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:])    
    
    
                    RECC_System.FlowDict['F_14_13_No'].Values[t,:,Cu_scrap_loc,0] = (1-Par_RECC_EoL_RR_No[t,0,6,:,5])*RECC_System.FlowDict['F_8_9_No'].Values[t,0,:,6,0]
    
                    RECC_System.FlowDict['F_14_13'].Values[t,Sector_1reg_rge,:,:]   =   np.einsum('Owe->Owe',RECC_System.FlowDict['F_14_13_No'].Values[t,:,:,:])
    
                
                
                
                
                # 5) Add re-use flow to inflow and calculate manufacturing output as final consumption - re-use, in Mt/yr, all elements, trgme, element composition not yet known.
    
                if len(Sector_32reg_rge) > 0:
                    RECC_System.FlowDict['F_5_6'].Values[t,0,Sector_32reg_rge,:,0]                   = np.einsum('rgm->gm',RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,:,:,0])    - np.einsum('crgm->gm',RECC_System.FlowDict['F_17_6_Nr'].Values[t,:,:,:,:,0])     # global total
                if len(Sector_11reg_rge) > 0:            
                    RECC_System.FlowDict['F_5_6'].Values[t,0,Sector_11reg_rge,:,0]    = np.einsum('lLm->Lm',RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,:,:,0]) - np.einsum('clLm->Lm',RECC_System.FlowDict['F_17_6_Nl'].Values[t,:,:,:,:,0])  # global total
                if len(Sector_1reg_rge) > 0:            
                    RECC_System.FlowDict['F_5_6'].Values[t,0,Sector_1reg_rge,:,0]     = np.einsum('oOm->Om',RECC_System.FlowDict['F_6_7_No'].Values[t,:,:,:,0]) - np.einsum('coOm->Om',RECC_System.FlowDict['F_17_6_No'].Values[t,:,:,:,:,0])  # global total
                Manufacturing_Output[t,:,:,mS,mR]                                 = RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0].copy()
                
                # 6) Calculate total manufacturing input and primary production, all elements, element composition not yet known.
                # Add fabrication scrap diversion, new scrap and calculate remelting.
                #Manufacturing_Input_m_ref    = np.einsum('mg,gm->m', Par_FabYield_total_inv[:,:,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]).copy()
                #Manufacturing_Input_gm_ref   = np.einsum('mg,gm->gm',Par_FabYield_total_inv[:,:,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]).copy()
                # Same as above, but the variables below will be adjusted for diverted fab scrap and uses subsequently:
                Manufacturing_Input_m_adj    = np.einsum('mg,gm->m', Par_FabYield_total_inv[:,Loc_all_products,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,Loc_all_products,:,0]).copy()
                Manufacturing_Input_gm_adj   = np.einsum('mg,gm->gm',Par_FabYield_total_inv[:,Loc_all_products,t,0],RECC_System.FlowDict['F_5_6'].Values[t,0,Loc_all_products,:,0]).copy()
                # split manufacturing material input into different products g:
                Manufacturing_Input_Split_gm = np.einsum('gm,m->gm', Manufacturing_Input_gm_adj, np.divide(1, Manufacturing_Input_m_adj, out=np.zeros_like(Manufacturing_Input_m_adj), where=Manufacturing_Input_m_adj!=0))
                            
                # 7) Determine available fabrication scrap diversion and secondary material, total and by element.
                # Determine fabscrapdiversionpotential:
                Fabscrapdiversionpotential_twm                     = np.einsum('wm,ow->wm',np.einsum('o,mwo->wm',RECC_System.ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[mR,:,t,mS],RECC_System.ParameterDict['6_PR_FabricationScrapDiversion'].Values[:,:,:,mS]),RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,0]).copy()
                # break down fabscrapdiversionpotential to e:
                NewScrapElemShares                                 = msf.TableWithFlowsToShares(RECC_System.StockDict['S_10'].Values[t-1,t-1,0,:,1::],axis=1) # element composition of fab scrap
                Fabscrapdiversionpotential_twme                    = np.zeros((Nt,Nw,Nm,Ne))
                Fabscrapdiversionpotential_twme[t,:,:,0]           = Fabscrapdiversionpotential_twm # total mass (all chem. elements)
                Fabscrapdiversionpotential_twme[t,:,:,1::]         = np.einsum('wm,we->wme',Fabscrapdiversionpotential_twm,NewScrapElemShares) # other chemical elements
                Fabscrapdiversionpotential_tme                     = np.einsum('wme->me',Fabscrapdiversionpotential_twme[t,:,:,:])
                RECC_System.FlowDict['F_10_12'].Values[t,:,:,:]    = np.einsum('wme,o->ome',Fabscrapdiversionpotential_twme[t,:,:,:],np.ones(No))
                
                
                if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList: 
                    RECC_System.FlowDict['F_10_9'].Values[t,:,:,:]    =  np.einsum('owe->owe',np.einsum('we,o->owe',RECC_System.FlowDict['F_9_10_Nr'].Values[t,:,:,:].sum(axis=0),np.ones(No)) + RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,:].copy()) - np.einsum('wme,o->owe',Fabscrapdiversionpotential_twme[t,:,:,:],np.ones(No)).copy()
                
                    
                if 'ind' in SectorList: 
                    RECC_System.FlowDict['F_10_9'].Values[t,:,:,:]     = np.einsum('owe->owe',np.einsum('we,o->owe',RECC_System.FlowDict['F_9_10_Nl'].Values[t,:,:,:].sum(axis=0),np.ones(No)) + RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,:].copy()) - np.einsum('wme,o->owe',Fabscrapdiversionpotential_twme[t,:,:,:],np.ones(No)).copy()
                
            
                if 'app' in SectorList or 'inf' in SectorList or 'nrbg' in SectorList or 'otv' in SectorList: 
                    RECC_System.FlowDict['F_10_9'].Values[t,:,:,:]     = np.einsum('owe->owe',np.einsum('we,o->owe',RECC_System.FlowDict['F_9_10_No'].Values[t,:,:,:].sum(axis=0),np.ones(No)) + RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,:].copy()) - np.einsum('wme,o->owe',Fabscrapdiversionpotential_twme[t,:,:,:],np.ones(No)).copy()
              
                
              ####from F_9_12 onwards the calculatins are copper specific
                
                RECC_System.FlowDict['F_9_12'].Values[t,:,Copper_loc,:]     = np.einsum('oe->oe', RECC_System.FlowDict['F_10_9'].Values[t,:,5,:]*RECC_System.ParameterDict['4_PY_MaterialProductionRemelting'].Values[CopperScrap_loc,Copper_loc,Cu_loc,Cu_wmg_loc,0,0])
        
         
                # 8) MARKET BALANCE for secondary materials:
    
                # a) Use diverted fab scrap first, if possible:
                DivFabScrap_to_Manuf = np.zeros((Nm,Ne))
                for mat in range(0,Nm):
                    if Fabscrapdiversionpotential_tme[mat,0] > 0:
                        if Fabscrapdiversionpotential_tme[mat,0] > Manufacturing_Input_m_adj[mat]:
                            DivFabScrap_to_Manuf[mat,:]            = (Fabscrapdiversionpotential_tme[mat,:] * Manufacturing_Input_m_adj[mat] / Fabscrapdiversionpotential_tme[mat,0]).copy()
                        else:
                            DivFabScrap_to_Manuf[mat,:]            = Fabscrapdiversionpotential_tme[mat,:].copy()
                Non_DivFabScrap                                    = Fabscrapdiversionpotential_tme - DivFabScrap_to_Manuf
                RemainingManufactInputDemand_1                     = Manufacturing_Input_m_adj - DivFabScrap_to_Manuf[:,0]    
                
                # b) Use secondary material, if possible:            
                SecondaryMaterialUse = np.zeros((Nm,Ne))
                for mat in range(0,Nm):
                    if RECC_System.FlowDict['F_9_12'].Values[t,0,mat,0] > 0:
                        if RECC_System.FlowDict['F_9_12'].Values[t,0,mat,0] > RemainingManufactInputDemand_1[mat]:
                            SecondaryMaterialUse[mat,:]            = (RECC_System.FlowDict['F_9_12'].Values[t,0,mat,:] * RemainingManufactInputDemand_1[mat] / RECC_System.FlowDict['F_9_12'].Values[t,0,mat,0]).copy()
                        else:
                            SecondaryMaterialUse[mat,:]            = RECC_System.FlowDict['F_9_12'].Values[t,0,mat,:].copy()
                Non_UsedSecMaterial                                = RECC_System.FlowDict['F_9_12'].Values[t,0,:,:] - SecondaryMaterialUse
                RemainingManufactInputDemand_2                     = RemainingManufactInputDemand_1 - SecondaryMaterialUse[:,0]
                
                # c) Use stock-piled secondary material, if available and if possible:
                RECC_System.StockDict['S_12'].Values[t,0,:,:]      = RECC_System.StockDict['S_12'].Values[t-1,0,:,:] # age stock pile by 1 year
                StockPileSecondaryMaterialUse = np.zeros((Nm,Ne))
                for mat in range(0,Nm):
                    if RECC_System.StockDict['S_12'].Values[t,0,mat,0] > 0:
                        if RECC_System.StockDict['S_12'].Values[t,0,mat,0] > RemainingManufactInputDemand_2[mat]:
                            StockPileSecondaryMaterialUse[mat,:]   = (RECC_System.StockDict['S_12'].Values[t,0,mat,:] * RemainingManufactInputDemand_2[mat] / RECC_System.StockDict['S_12'].Values[t,0,mat,0]).copy()
                        else:
                            StockPileSecondaryMaterialUse[mat,:]   = RECC_System.StockDict['S_12'].Values[t,0,mat,:].copy()
                RECC_System.StockDict['S_12'].Values[t,0,:,:]      = RECC_System.StockDict['S_12'].Values[t,0,:,:] - StockPileSecondaryMaterialUse.copy()
                PrimaryProductionDemand = RemainingManufactInputDemand_2 - StockPileSecondaryMaterialUse[:,0]
                
                # d) convert internal calculations to system variables:
                RECC_System.FlowDict['F_12_5'].Values[t,0,:,:]     = DivFabScrap_to_Manuf + SecondaryMaterialUse + StockPileSecondaryMaterialUse
                
                # include copper recovery efficiency from Glöser et al. (2013) RECC_System.ParameterDict['4_PY_MaterialProductionRemelting'].Values[:,:,:,:,0,:] with dim wmePo
                
               # Glöser et al. in RECC data sheet  defined for 
               #w Waste_Scrap_m5 = #1 Copper scrap
               #m Engineering_Materials_m2 = copper electric grade
               #e Chemical_Elements = Cu
               #P Waste_Mgt_Industries_i5 = remelting of copper wire scrap
               
                RECC_System.ParameterDict['4_PY_MaterialProductionRemelting'].Values[CopperScrap_loc,Copper_loc,Cu_loc,Cu_wmg_loc,0,0]
    
    
                RECC_System.FlowDict['F_10_23'].Values[t,CopperScrap_loc,Cu_loc] = RECC_System.FlowDict['F_10_9'].Values[t,:,CopperScrap_loc,Cu_loc].sum(axis=0)*(1-RECC_System.ParameterDict['4_PY_MaterialProductionRemelting'].Values[CopperScrap_loc,Copper_loc,Cu_loc,Cu_wmg_loc,0,0])
    
                RECC_System.FlowDict['F_4_5'].Values[t,:,:]        = np.einsum('m,me->me',PrimaryProductionDemand,RECC_System.ParameterDict['3_MC_Elements_Materials_Primary'].Values)
                if ScriptConfig['ScrapExport'] == 'True':
                    RECC_System.FlowDict['F_12_21'].Values[t,0,:,:] = Non_DivFabScrap.copy() + Non_UsedSecMaterial.copy() 
                else:
                    RECC_System.StockDict['S_12'].Values[t,0,:,:] += Non_DivFabScrap.copy() + Non_UsedSecMaterial.copy()
             
                # e) Element composition of material flows:         
                Manufacturing_Input_me_final                       = RECC_System.FlowDict['F_4_5'].Values[t,:,:] + RECC_System.FlowDict['F_12_5'].Values[t,0,:,:]
                Manufacturing_Input_gme_final                      = np.einsum('gm,me->gme',Manufacturing_Input_Split_gm,Manufacturing_Input_me_final)
                Element_Material_Composition_Manufacturing         = msf.DetermineElementComposition_All_Oth(Manufacturing_Input_me_final)
                Element_Material_Composition_raw[t,:,:,mS,mR]      = Element_Material_Composition_Manufacturing.copy()
                
                Element_Material_Composition[t,:,:,mS,mR]          = Element_Material_Composition_Manufacturing.copy()
                Par_Element_Composition_of_Materials_m[t+115,:,:]  = Element_Material_Composition_Manufacturing.copy()
                Par_Element_Composition_of_Materials_u[t+115,:,:]  = Element_Material_Composition_Manufacturing.copy()
    
                # End of 8)MARKET BALANCE for secondary material.
    
                # 9) Primary production and forestry carbon balance
                RECC_System.FlowDict['F_3_4'].Values[t,:,:]        = RECC_System.FlowDict['F_4_5'].Values[t,:,:]
                RECC_System.FlowDict['F_0_3'].Values[t,:,:]        = RECC_System.FlowDict['F_3_4'].Values[t,:,:] 
                # We don't model processes 1 and 2 for materials other than wood. 
                # Supply chain of primary materials is linked to process 3 via parameter 4_PE_ProcessExtensions instead.
                # Only carbon in wood is passed on to process 2 (wood market)
                RECC_System.FlowDict['F_0_3'].Values[t,Wood_loc,Carbon_loc] = 0
                RECC_System.FlowDict['F_2_3'].Values[t,Wood_loc,Carbon_loc] = RECC_System.FlowDict['F_3_4'].Values[t,Wood_loc,Carbon_loc] # carbon only!
                # Further quantification of processes 1 and 2 is done below, after energy consumption flows are calculated.
                
                # 10) Calculate manufacturing output, at global level only
                RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,:]    = np.einsum('me,gm->gme',Element_Material_Composition_Manufacturing,RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0])
            
                # 10a) Calculate material composition of product consumption
                Throughput_FinalGoods_me                           = RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,:].sum(axis =0) + np.einsum('crgme->me',RECC_System.FlowDict['F_17_6'].Values[t,0:CohortOffset,:,:,:,:])
                Element_Material_Composition_cons                  = msf.DetermineElementComposition_All_Oth(Throughput_FinalGoods_me)
                Element_Material_Composition_con[t,:,:,mS,mR]      = Element_Material_Composition_cons.copy()
                
                Par_Element_Composition_of_Materials_c[t,:,:]      = Element_Material_Composition_cons.copy()
                Par_3_MC_Stock_ByElement_Nl[CohortOffset,:,:,:,:]  = np.einsum('mLl,me->lLme',Par_RECC_MC_Nl[CohortOffset,:,:,:,mS],Par_Element_Composition_of_Materials_c[t,:,:]) # clLme
                Par_3_MC_Stock_ByElement_No[CohortOffset,:,:,:,:]  = np.einsum('mOo,me->oOme',Par_RECC_MC_No[CohortOffset,:,:,:,mS],Par_Element_Composition_of_Materials_c[t,:,:]) # cOome                    
                
                # 11) Calculate manufacturing scrap 
                RECC_System.FlowDict['F_5_10'].Values[t,0,:,:]     = np.einsum('gme,mwg->we',Manufacturing_Input_gme_final,Par_FabYieldLoss[:,:,Loc_all_products,t,0]) 
                # Fabrication scrap, to be recycled next year:
                RECC_System.StockDict['S_10'].Values[t,t,:,:,:]    = RECC_System.FlowDict['F_5_10'].Values[t,:,:,:]
            
                # 12) Calculate element composition of final consumption and latest age-cohort in in-use stock
                if 'pav' in SectorList:
                    # update mat. composition by element for current year and latest age-cohort
                    Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset,:,Sector_pav_rge,:,:] = Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_pav_rge,:,:]
                    Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_pav_rge,:,:]   = np.einsum('me,Bmr->Brme',Par_Element_Composition_of_Materials_c[t,:,:],Par_RECC_MC_Nr[CohortOffset,:,Sector_pav_rge,:,mS,t])
                    RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_pav_rge,:,:]   = \
                    np.einsum('prme,pr->prme',Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_pav_rge,:,:],Inflow_Detail_UsePhase_p[t,:,:])/1000 # all elements, Indices='t,r,p,m,e'
                    RECC_System.StockDict['S_7_Nr'].Values[t,0:CohortOffset+1,:,Sector_pav_rge,:,:] = \
                    np.einsum('pcrme,cpr->pcrme',Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset+1,:,Sector_pav_rge,:,:],Stock_Detail_UsePhase_p[t,0:CohortOffset+1,:,:])/1000 # All elements.
    
                if 'reb' in SectorList:
                    # update mat. composition by element for current year and latest age-cohort
                    Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_reb_rge,:,:]   = np.einsum('me,Bmr->Brme',Par_Element_Composition_of_Materials_c[t,:,:],Par_RECC_MC_Nr[SwitchTime+t-1,:,Sector_reb_rge,:,mS,t])
                    # Determine element breakdown of inflow and renovation material
                    RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_reb_rge,:,:]   = \
                    np.einsum('me,Brm->Brme',Par_Element_Composition_of_Materials_c[t,:,:],RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_reb_rge,:,0]) # all elements, Indices='t,r,B,m,e'
                    F_6_7_ren[t,:,:,Sector_reb_rge,:,:] = np.einsum('me,Bcrm->Bcrme',Par_Element_Composition_of_Materials_c[t,:,:],F_6_7_ren[t,:,:,Sector_reb_rge,:,0]) # all elements, Indices='t,c,r,B,m,e' (c is age-cohort where material flows)
                    # Determine the element material composition at the end of last year, as weighting factor for existing stock
                    Divisor  = np.einsum('Bcrm,e->Bcrme',Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,0],np.ones(Ne))
                    Par_ElementComposition_LastYear = np.divide(Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_reb_rge,:,:],Divisor, out=np.zeros_like(Divisor), where=Divisor!=0) #Bcrme
                    # Compile all materials present in stock broken down by element:
                    StockMat = F_6_7_ren[t,0:CohortOffset,:,Sector_reb_rge,:,:] + np.einsum('Bcrm,Bcrme->Bcrme',RECC_System.StockDict['S_7'].Values[t,0:CohortOffset,:,Sector_reb_rge,:,0] - F_6_7_ren[t,0:CohortOffset,:,Sector_reb_rge,:,0],Par_ElementComposition_LastYear)
                    Divisor  = np.einsum('Bcrm,e->Bcrme',StockMat[:,:,:,:,0],np.ones(Ne))
                    # Caculate product element composition of latest age-cohort from total materials by element:
                 #!   Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset,:,Sector_reb_rge,:,:]  = np.einsum('Bcmr,Bcrme->Bcrme',Par_RECC_MC_Nr[0:CohortOffset,:,Sector_reb_rge,:,mS,t],np.divide(StockMat,Divisor, out=np.zeros_like(Divisor), where=Divisor!=0))
                    # Update stock: break down material into elements:                
                    RECC_System.StockDict['S_7_Nr'].Values[t,0:CohortOffset +1,:,Sector_reb_rge,:,:] = \
                    np.einsum('Bcrme,cBr->Bcrme',Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset +1,:,Sector_reb_rge,:,:],Stock_Detail_UsePhase_B[t,0:CohortOffset +1,:,:])/1000
                    
                if 'nrb' in SectorList:
                    # update mat. composition by element for current year and latest age-cohort
                    Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset,:,Sector_nrb_rge,:,:] = Par_3_MC_Stock_ByElement_Nr[t-1,0:CohortOffset,:,Sector_nrb_rge,:,:]
                    Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_nrb_rge,:,:]   = np.einsum('me,Bmr->Brme',Par_Element_Composition_of_Materials_c[t,:,:],Par_RECC_MC_Nr[CohortOffset,:,Sector_nrb_rge,:,mS,t])
                    RECC_System.FlowDict['F_6_7_Nr'].Values[t,:,Sector_nrb_rge,:,:]   = \
                    np.einsum('Nrme,Nr->Nrme',Par_3_MC_Stock_ByElement_Nr[t,CohortOffset,:,Sector_nrb_rge,:,:],Inflow_Detail_UsePhase_N[t,:,:])/1000 # all elements, Indices='t,r,N,m,e'
                    RECC_System.StockDict['S_7_Nr'].Values[t,0:CohortOffset +1,:,Sector_nrb_rge,:,:] = \
                    np.einsum('Ncrme,cNr->Ncrme',Par_3_MC_Stock_ByElement_Nr[t,0:CohortOffset +1,:,Sector_nrb_rge,:,:],Stock_Detail_UsePhase_N[t,0:CohortOffset +1,:,:])/1000 # All elements.
                    
               
                if 'ind' in SectorList:        
                    RECC_System.FlowDict['F_6_7_Nl'].Values[t,:,:,:,:]   = \
                    np.einsum('lIme,Il->lIme',Par_3_MC_Stock_ByElement_Nl[CohortOffset,:,:,:,:],Inflow_Detail_UsePhase_I[t,:,:])/1000 # all elements, Indices='t,l,I,m,e'                
                    RECC_System.StockDict['S_7_Nl'].Values[t,0:CohortOffset +1,:,Sector_ind_rge_reg,:,:] = \
                    np.einsum('clIme,cIl->Iclme',Par_3_MC_Stock_ByElement_Nl[0:CohortOffset +1,:,Sector_ind_rge_reg,:,:],Stock_Detail_UsePhase_I[t,0:CohortOffset +1,:,:])/1000 # All elements. In Mt
               
                
                if 'app' in SectorList:        
                    RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_app_rge_reg,:,:]   = \
                    np.einsum('aome,ao->aome',Par_3_MC_Stock_ByElement_No[CohortOffset,:,Sector_app_rge_reg,:,:], Inflow_Detail_UsePhase_a[t,:,:])/1000000000000 # all elements, Indices='t,o,a,m,e'                
                    RECC_System.StockDict['S_7_No'].Values[t,0:CohortOffset +1,:,Sector_app_rge_reg,:,:] = \
                    np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:CohortOffset +1,:,Sector_app_rge_reg,:,:],Stock_Detail_UsePhase_a[t,0:CohortOffset +1,:,:])/1000000000000 # All elements.In Mt
               
                if 'otv' in SectorList:        
                    RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_otv_rge_reg,:,:]   = \
                    np.einsum('aome,ao->aome',Par_3_MC_Stock_ByElement_No[CohortOffset,:,Sector_otv_rge_reg,:,:], Inflow_Detail_UsePhase_otv[t,:,:])/1000000000 # all elements, Indices='t,o,v,m,e'                
                    RECC_System.StockDict['S_7_No'].Values[t,0:CohortOffset +1,:,Sector_otv_rge_reg,:,:] = \
                    np.einsum('coame,cao->acome',Par_3_MC_Stock_ByElement_No[0:CohortOffset +1,:,Sector_otv_rge_reg,:,:],Stock_Detail_UsePhase_otv[t,0:CohortOffset +1,:,:])/1000000000 # All elements.In Mt
               
       
            
                if 'nrbg' in SectorList: # results for global region o are on position 0 of r index.
                    RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_nrbg_rge_reg,:,:]   = \
                    np.einsum('Nome,No->Nome',Par_3_MC_Stock_ByElement_No[CohortOffset,:,Sector_nrbg_rge_reg,:,:],Inflow_Detail_UsePhase_Ng[t,:,:])/1000 # all elements, Indices='t,o,N,m,e'                
                    RECC_System.StockDict['S_7_No'].Values[t,0:CohortOffset +1,:,Sector_nrbg_rge_reg,:,:] = \
                    np.einsum('coNme,cNo->Ncome',Par_3_MC_Stock_ByElement_No[0:CohortOffset +1,:,Sector_nrbg_rge_reg,:,:],Stock_Detail_UsePhase_Ng[t,0:CohortOffset +1,:,:])/1000 # All elements.In Mt
                 
                if 'inf' in SectorList: # results for global region o are on position 0 of r index.
                    RECC_System.FlowDict['F_6_7_No'].Values[t,:,Sector_inf_rge_reg,:,:]   = \
                    np.einsum('iome,io->iome',Par_3_MC_Stock_ByElement_No[CohortOffset,:,Sector_inf_rge_reg,:,:],Inflow_Detail_UsePhase_i[t,:,:]) # all elements, Indices='t,o,N,m,e'                
                    RECC_System.StockDict['S_7_No'].Values[t,0:CohortOffset +1,:,Sector_inf_rge_reg,:,:] = \
                    np.einsum('coime,cio->icome',Par_3_MC_Stock_ByElement_No[0:CohortOffset +1,:,Sector_inf_rge_reg,:,:],Stock_Detail_UsePhase_i[t,0:CohortOffset +1,:,:]) # All elements.In Mt
                 
            
                   
        
                # 13) Calculate waste mgt. losses.
                
                
    #           # if 'pav' in SectorList or 'reb'in SectorList or 'nrb' in SectorList:
    #             if 'inf' in SectorList or 'app' in SectorList or 'nrbg' in SectorList: 
    #                     if 'ind' in SectorList: 
    #                         RECC_System.FlowDict['F_9_13'].Values[t,:]          = np.einsum('rgme->e',RECC_System.FlowDict['F_8_9'].Values[t,:,:,:,:])    - np.einsum('rwe->e',RECC_System.FlowDict['F_9_10'].Values[t,:,:,:]) \
    #                                                               + np.einsum('lLme->e',RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:]) - np.einsum('lwe->e',RECC_System.FlowDict['F_9_10_Nl'].Values[t,:,:,:]) \
    #                                                               + np.einsum('oOme->e',RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:]) - np.einsum('owe->e',RECC_System.FlowDict['F_9_10_No'].Values[t,:,:,:]) \
    #                                                               + np.einsum('rwe->e',RECC_System.FlowDict['F_10_9'].Values[t,:,:,:])      - np.einsum('ome->e',RECC_System.FlowDict['F_9_12'].Values[t,:,:,:])            
                                                                  
    #                     else:
    #                         RECC_System.FlowDict['F_9_13'].Values[t,:]          = np.einsum('rgme->e',RECC_System.FlowDict['F_8_9'].Values[t,:,:,:,:])    - np.einsum('rwe->e',RECC_System.FlowDict['F_9_10'].Values[t,:,:,:]) \
    #                                                               + np.einsum('oOme->e',RECC_System.FlowDict['F_8_9_No'].Values[t,:,:,:,:]) - np.einsum('owe->e',RECC_System.FlowDict['F_9_10_No'].Values[t,:,:,:]) \
    #                                                               + np.einsum('rwe->e',RECC_System.FlowDict['F_10_9'].Values[t,:,:,:])      - np.einsum('ome->e',RECC_System.FlowDict['F_9_12'].Values[t,:,:,:])            
                                                                  
    #             else:         
    #                 if 'ind' in SectorList: 
    #                         RECC_System.FlowDict['F_9_13'].Values[t,:]          = np.einsum('rgme->e',RECC_System.FlowDict['F_8_9'].Values[t,:,:,:,:])    - np.einsum('rwe->e',RECC_System.FlowDict['F_9_10'].Values[t,:,:,:]) \
    #                                                               + np.einsum('lLme->e',RECC_System.FlowDict['F_8_9_Nl'].Values[t,:,:,:,:]) - np.einsum('lwe->e',RECC_System.FlowDict['F_9_10_Nl'].Values[t,:,:,:]) \
    #                                                               + np.einsum('rwe->e',RECC_System.FlowDict['F_10_9'].Values[t,:,:,:])      - np.einsum('ome->e',RECC_System.FlowDict['F_9_12'].Values[t,:,:,:])            
    #                 else:                                                     
    #                     RECC_System.FlowDict['F_9_13'].Values[t,:]          = np.einsum('rgme->e',RECC_System.FlowDict['F_8_9'].Values[t,:,:,:,:])    - np.einsum('rwe->e',RECC_System.FlowDict['F_9_10'].Values[t,:,:,:]) \
    #                                                               + np.einsum('rwe->e',RECC_System.FlowDict['F_10_9'].Values[t,:,:,:])      - np.einsum('ome->e',RECC_System.FlowDict['F_9_12'].Values[t,:,:,:])            
                          
    # #                            
                                
                
              #  if 'C ' in Element_List:    
    
                    # # Wood material lost in waste mgt.
                    # WoodMaterialLoss_t                                 = RECC_System.FlowDict['F_9_13'].Values[t,Carbon_loc] / (RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44)
                    
                    # # Energy Recovery: in TJ/yr
                    # EnergyRecovery_WoodCombustion_EL[t,mS,mR]  =  1000 * RECC_System.ParameterDict['4_PE_ElectricityFromWoodCombustion'].Values[Woodwaste_loc,0,0] * WoodMaterialLoss_t
                                                                       
                    # # Biogenic CO2 emissions from waste, Mt:
                    # BiogenicCO2WasteCombustion[t,mS,mR]        =     RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * WoodMaterialLoss_t
                
                # 14) Calculate stock changes
                
                if 'inf' in SectorList or 'app' in SectorList or 'nrbg' in SectorList or 'otv' in SectorList: 
                    RECC_System.StockDict['dS_7_No'].Values[t,:,:,:,:,:]  = RECC_System.StockDict['S_7_No'].Values[t,:,:,:,:,:] - RECC_System.StockDict['S_7_No'].Values[t-1,:,:,:,:,:]            
                    RECC_System.StockDict['dS_11_No'].Values[t,:,:,:]             = np.einsum('coOme->Oome',RECC_System.FlowDict['F_8_11_No'].Values[t,:,:,:,:,:]) 
                    RECC_System.StockDict['S_11_No'].Values[t,:,:,:]             = np.einsum('coOme->Oome',RECC_System.FlowDict['F_8_11_No'].Values[t,:,:,:,:,:]) +  RECC_System.StockDict['S_11_No'].Values[t-1,:,:,:]      
                    RECC_System.StockDict['S_11'].Values[t,Sector_1reg_rge,:,:]        =  np.einsum('grme->gme',RECC_System.StockDict['S_11_No'].Values[t,:,:,:])
    
    
                    RECC_System.StockDict['dS_19_No'].Values[t,:,:,:]             = np.einsum('coOme->Oome',RECC_System.FlowDict['F_8_19_No'].Values[t,:,:,:,:,:]) 
                    RECC_System.StockDict['S_19_No'].Values[t,:,:,:]             = np.einsum('coOme->Oome',RECC_System.FlowDict['F_8_19_No'].Values[t,:,:,:,:,:]) +  RECC_System.StockDict['S_19_No'].Values[t-1,:,:,:]      
                    RECC_System.StockDict['S_19'].Values[t,Sector_1reg_rge,:,:]        =  np.einsum('grme->gme',RECC_System.StockDict['S_19_No'].Values[t,:,:,:])
    
    
    
                    RECC_System.StockDict['dS_13'].Values[t,Sector_1reg_rge,:]        = RECC_System.FlowDict['F_14_13_No'].Values[t,:,Cu_scrap_loc,:]
                    RECC_System.StockDict['S_13'].Values[t,:,:]        =  RECC_System.StockDict['dS_13'].Values[t,:,:] + RECC_System.StockDict['S_13'].Values[t-1,:,:]  
    
                    ObsoleteFlows[:,:,mS,mR]    = RECC_System.FlowDict['F_8_11_No'].Values[:,:,:,:,6,:].sum(axis=1).sum(axis=1).sum(axis=2)
                    HybernatingFlows[:,:,mS,mR] = RECC_System.FlowDict['F_8_19_No'].Values[:,:,:,:,6,:].sum(axis=1).sum(axis=1).sum(axis=2)
                    SortingFlows[:,:,mS,mR]     = RECC_System.FlowDict['F_14_13_No'].Values[:,:,Cu_scrap_loc,:].sum(axis=2)
    
    
                if 'ind' in SectorList:        
                    RECC_System.StockDict['dS_7_Nl'].Values[t,:,:,:,:,:]  = RECC_System.StockDict['S_7_Nl'].Values[t,:,:,:,:,:] - RECC_System.StockDict['S_7_Nl'].Values[t-1,:,:,:,:,:]
                    RECC_System.StockDict['dS_11_Nl'].Values[t,:,:,:]             = np.einsum('clLme->Llme',RECC_System.FlowDict['F_8_11_Nl'].Values[t,:,:,:,:,:]) 
                    RECC_System.StockDict['S_11_Nl'].Values[t,:,:,:]             = np.einsum('clLme->Llme',RECC_System.FlowDict['F_8_11_Nl'].Values[t,:,:,:,:,:]) + RECC_System.StockDict['S_11_Nl'].Values[t-1,:,:,:]  
                    RECC_System.StockDict['S_11'].Values[t,Sector_11reg_rge,:,:]        =  RECC_System.StockDict['S_11_Nl'].Values[t,:,:,:].sum(axis=1)            
    
                    RECC_System.StockDict['dS_19_Nl'].Values[t,:,:,:]             = np.einsum('clLme->Llme',RECC_System.FlowDict['F_8_19_Nl'].Values[t,:,:,:,:,:]) 
                    RECC_System.StockDict['S_19_Nl'].Values[t,:,:,:]             = np.einsum('clLme->Llme',RECC_System.FlowDict['F_8_19_Nl'].Values[t,:,:,:,:,:]) + RECC_System.StockDict['S_19_Nl'].Values[t-1,:,:,:]  
                    RECC_System.StockDict['S_19'].Values[t,Sector_11reg_rge,:,:]        =  RECC_System.StockDict['S_19_Nl'].Values[t,:,:,:].sum(axis=1)            
    
    
                
    
                    RECC_System.StockDict['dS_13'].Values[t,Sector_11reg_rge,:]        = RECC_System.FlowDict['F_14_13_Nl'].Values[t,:,Cu_scrap_loc,:]
                    RECC_System.StockDict['S_13'].Values[t,:,:]        =  RECC_System.StockDict['dS_13'].Values[t,:,:] + RECC_System.StockDict['S_13'].Values[t-1,:,:]  
    
    
                    ObsoleteFlows[:,:,mS,mR]    = RECC_System.FlowDict['F_8_11_Nl'].Values[:,:,:,:,6,:].sum(axis=1).sum(axis=1).sum(axis=2)
                    HybernatingFlows[:,:,mS,mR] = RECC_System.FlowDict['F_8_19_Nl'].Values[:,:,:,:,6,:].sum(axis=1).sum(axis=1).sum(axis=2)
                    SortingFlows[:,:,mS,mR]     = RECC_System.FlowDict['F_14_13_Nl'].Values[:,:,Cu_scrap_loc,:].sum(axis=2)
               
                if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList: 
                    RECC_System.StockDict['dS_7_Nr'].Values[t,:,:,:,:,:]     = RECC_System.StockDict['S_7'].Values[t,:,:,:,:,:] - RECC_System.StockDict['S_7'].Values[t-1,:,:,:,:,:]
                    RECC_System.StockDict['dS_11_Nr'].Values[t,:,:,:]             = np.einsum('crgme->grme',RECC_System.FlowDict['F_8_11_Nr'].Values[t,:,:,:,:,:]) 
                    RECC_System.StockDict['S_11_Nr'].Values[t,:,:,:]             = np.einsum('crgme->grme',RECC_System.FlowDict['F_8_11_Nr'].Values[t,:,:,:,:,:]) +  RECC_System.StockDict['S_11_Nr'].Values[t-1,:,:,:]    
                    RECC_System.StockDict['S_11'].Values[t,Sector_32reg_rge,:,:]        =  np.einsum('grme->gme', RECC_System.StockDict['S_11_Nr'].Values[t,:,:,:])
    
                    RECC_System.StockDict['dS_19_Nr'].Values[t,:,:,:]             = np.einsum('crgme->grme',RECC_System.FlowDict['F_8_19_Nr'].Values[t,:,:,:,:,:]) 
                    RECC_System.StockDict['S_19_Nr'].Values[t,:,:,:]             = np.einsum('crgme->grme',RECC_System.FlowDict['F_8_19_Nr'].Values[t,:,:,:,:,:]) +  RECC_System.StockDict['S_19_Nr'].Values[t-1,:,:,:]    
                    RECC_System.StockDict['S_19'].Values[t,Sector_32reg_rge,:,:]        =  np.einsum('grme->gme', RECC_System.StockDict['S_19_Nr'].Values[t,:,:,:])
    
            
                    RECC_System.StockDict['dS_13'].Values[t,Sector_32reg_rge,:]        = RECC_System.FlowDict['F_14_13_Nr'].Values[t,:,Cu_scrap_loc,:]
                    RECC_System.StockDict['S_13'].Values[t,:,:]        =  RECC_System.StockDict['dS_13'].Values[t,:,:] + RECC_System.StockDict['S_13'].Values[t-1,:,:]  
    
           
    
                    ObsoleteFlows[:,:,mS,mR]    = RECC_System.FlowDict['F_8_11_Nr'].Values[:,:,:,:,6,:].sum(axis=1).sum(axis=1).sum(axis=2)
                    HybernatingFlows[:,:,mS,mR] = RECC_System.FlowDict['F_8_19_Nr'].Values[:,:,:,:,6,:].sum(axis=1).sum(axis=1).sum(axis=2)
                    SortingFlows[:,:,mS,mR]     = RECC_System.FlowDict['F_14_13_Nr'].Values[:,:,Cu_scrap_loc,:].sum(axis=2)
            
            
    
                #Fabrication scrap buffer change
                RECC_System.StockDict['dS_10'].Values[t,:,:,:]        = RECC_System.StockDict['S_10'].Values[t,t,:,:,:]  - RECC_System.StockDict['S_10'].Values[t-1,t-1,:,:,:]
                
                #Secondary material buffer change
                RECC_System.StockDict['dS_12'].Values[t,:,:,:]        = RECC_System.StockDict['S_12'].Values[t,:,:,:]    - RECC_System.StockDict['S_12'].Values[t-1,:,:,:]
                
     
    #        
    
                # Recovery losses
                RECC_System.StockDict['dS_23'].Values[t,:,:]        =  RECC_System.FlowDict['F_10_23'].Values[t,:,:]
                RECC_System.StockDict['S_23'].Values[t,:,:]        =  RECC_System.StockDict['dS_23'].Values[t,:,:] + RECC_System.StockDict['S_23'].Values[t-1,:,:]  
    
            
          
                ObsoleteStocks[:,:,mS,mR]    =  RECC_System.StockDict['S_11'].Values[:,:,6,:].sum(axis=2)
                HybernatingStocks[:,:,mS,mR] =  RECC_System.StockDict['S_19'].Values[:,:,6,:].sum(axis=2)
                SortingLosses[:,:,mS,mR]     =  RECC_System.StockDict['S_13'].Values[:,:,:].sum(axis=2)
                RecoveryLosses[:,mS,mR]      =  RECC_System.StockDict['S_23'].Values[:,Cu_scrap_loc,:].sum(axis=1)
            
            
    
                RecoveryFlows[:,mS,mR]  	= RECC_System.FlowDict['F_10_23'].Values[:,Cu_scrap_loc,:].sum(axis=1)
                if 'reb' in SectorList or 'pav' in SectorList:  
                    Cu_Stocks_Nr[:,:,mS,mR]   = RECC_System.StockDict['S_7_Nr'].Values[:,:,:,:,6,0].sum(axis=1).sum(axis=1)
               
                if 'app' in SectorList or 'inf' in SectorList or'nrbg' in SectorList or 'otv' in SectorList:  
                    Cu_Stocks_No[:,:,mS,mR]   = RECC_System.StockDict['S_7_No'].Values[:,:,:,:,6,0].sum(axis=1).sum(axis=1)
               
                
                if 'ind' in SectorList:  
                    Cu_Stocks_Nl[:,:,mS,mR]   = RECC_System.StockDict['S_7_Nl'].Values[:,:,:,:,6,0].sum(axis=1).sum(axis=1)

    
    
            
            # Diagnostics:
    #         Aa = np.einsum('ptcrm->trm',RECC_System.FlowDict['F_7_8'].Values[:,:,:,Sector_pav_rge,:,0]) # VehiclesOutflowMaterials
    #        Ab = np.einsum('Btcrm->trm',RECC_System.FlowDict['F_7_8'].Values[:,:,:,Sector_reb_rge,:,0]) # BuildingOutflowMaterials
    #        Aa = np.einsum('tcrgm->tmr',RECC_System.FlowDict['F_7_8'].Values[:,:,:,:,:,0])              # outflow use phase
    #        Aa = np.einsum('tcoOm->tOm',RECC_System.FlowDict['F_7_8_No'].Values[:,:,:,:,:,0])           # outflow use phase
    #        Aa = np.einsum('Btcrm->trm',RECC_System.StockDict['S_7'].Values[:,:,:,Sector_reb_rge,:,0])  # BuildingOutflowMaterials
    #        Aa = np.einsum('ptcrm->trm',RECC_System.StockDict['S_7'].Values[:,:,:,Sector_pav_rge,:,0])  # VehiclesOutflowMaterials
    #        Aa = np.einsum('Btrm->trm',RECC_System.FlowDict['F_6_7_Nr'].Values[:,:,Sector_reb_rge,:,0])    # BuildingInflowMaterials
    #        Aa = np.einsum('ptrm->trm',RECC_System.FlowDict['F_6_7_Nr'].Values[:,:,Sector_pav_rge,:,0])    # PassVehsInflowMaterials        
    #
    #        Aa = np.einsum('tcgr->tgr',Outflow_Detail_UsePhase_p)                                       # product outflow use phase
    #        Aa = np.einsum('tgr->tgr',Inflow_Detail_UsePhase_p)                                         # product inflow use phase
    #        Aa = np.einsum('tgr->tg',Inflow_Detail_UsePhase_p)                                          # product inflow use phase, global total
    #        Aa = np.einsum('tcgr->tgr',Stock_Detail_UsePhase_p)                                         # Total stock time series            
    #        Aa = np.einsum('tcgr->tgr',Outflow_Detail_UsePhase_B)                                       # product outflow use phase
    #        Aa = np.einsum('tgr->tgr',Inflow_Detail_UsePhase_B)                                         # product inflow use phase
    #        Aa = np.einsum('tgr->tg',Inflow_Detail_UsePhase_B)                                          # product inflow use phase, global total
    #        Aa = np.einsum('tcgr->tgr',Stock_Detail_UsePhase_B)                                         # Total stock time series            
    #        
    #        # Material composition and flows
    #        Aa = Par_3_MC_Stock_ByElement_Nr[:,0,:,:,0] # indices: cgm.
    #        Aa = RECC_System.FlowDict['F_5_10'].Values[:,0,:,:] # indices: twe
    #        Aa = np.einsum('trw->trw',RECC_System.FlowDict['F_9_10'].Values[:,:,:,0])                   # old scrap
    #        Aa = np.einsum('trgm->tr',RECC_System.FlowDict['F_8_9'].Values[:,:,:,:,0])                  # inflow waste mgt.
    #        Aa = np.einsum('tgm->tgm',RECC_System.FlowDict['F_8_9'].Values[:,0,:,:,0])                  # inflow waste mgt.        
    #        Aa = np.einsum('trm->tm',RECC_System.FlowDict['F_9_12'].Values[:,:,:,0])                    # secondary material
    #        Aa = np.einsum('trm->trm',RECC_System.FlowDict['F_12_5'].Values[:,:,:,0])                   # secondary material use
    #        Aa = np.einsum('tm->tm',RECC_System.FlowDict['F_4_5'].Values[:,:,0])                        # primary material production
    #        Aa = np.einsum('ptrm->trm',RECC_System.FlowDict['F_5_6'].Values[:,:,Sector_pav_rge,:,0])    # materials in manufactured vehicles
    #        Aa = np.einsum('Btrm->trm',RECC_System.FlowDict['F_5_6'].Values[:,:,Sector_reb_rge,:,0])    # materials in manufactured buildings
    #        Aa = Element_Material_Composition[:,:,:,0,0]                                                # indices tme, latter to indices are mS and mR
    #        
    #        # Manufacturing diagnostics
    #        Aa = RECC_System.FlowDict['F_5_6'].Values[:,0,:,:,:].sum(axis=1) #tme
    #        Ab = RECC_System.FlowDict['F_5_10'].Values[:,0,:,:] # twe
    #        Ac = RECC_System.FlowDict['F_4_5'].Values # tme
    #        Ad = RECC_System.FlowDict['F_12_5'].Values[:,0,:,:] # tme
    #        Bal_20 = Aa[20,:,0] - Ac[20,:,0] - Ad[20,:,0]
    #        Bal_30 = Aa[30,:,0] - Ac[30,:,0] - Ad[30,:,0]
    #        Aa = RECC_System.FlowDict['F_5_6'].Values[t,0,:,:,0]
    #        
    #        # ReUse Diagnostics
    #        Aa = np.einsum('tcrgm->tr',RECC_System.FlowDict['F_8_17'].Values[:,:,:,:,:,0])   # reuse        
    #        Aa = np.einsum('tcrgme->tme',RECC_System.FlowDict['F_8_17'].Values)
    #        Aa = np.einsum('tcrgme->tme',RECC_System.FlowDict['F_17_6'].Values)
    #        Aa = Element_Material_Composition_con[:,:,:,mS,mR]
    #        Aa = MassShareVeh[:,0,:,:]
    #        Aa = np.einsum('m,crgm->cgm',ReUsePotential_Materials_t_m_Veh,MassShareVeh)
                
            ##########################################################
            #    Section 6) Post-process RECC model solution         #
            ##########################################################            
                        
            # A) Calculate intensity of operation, by sector
            # Hop over to save computation time:
            # SysVar_StockServiceProvision_UsePhase_pav = np.einsum('Vrt,tcpr->tcprV',  RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[:,:,:,mS],     Stock_Detail_UsePhase_p)
            # SysVar_StockServiceProvision_UsePhase_reb = np.einsum('tcBVr,tcBr->tcBrV',RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],     Stock_Detail_UsePhase_B)
            # SysVar_StockServiceProvision_UsePhase_nrb = np.einsum('cNVr,tcNr->tcNrV', RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[:,:,:,:,mS], Stock_Detail_UsePhase_N)
            # Unit: million km/yr for vehicles, million m2 for buildings by three use types: heating, cooling, and DHW.
            # Aggreated computation of building service for export:
            SysVar_StockServiceProvision_UsePhase_reb_agg = np.einsum('tcBVr,tcBr->tV',RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],     Stock_Detail_UsePhase_B)
         #   SysVar_StockServiceProvision_UsePhase_nrb_agg = np.einsum('cNVr,tcNr->tV' ,RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[:,:,:,:,mS], Stock_Detail_UsePhase_N)
            
            # B) Calculate total operational energy use, by sector
            # Hop over to save computation time:
            # SysVar_EnergyDemand_UsePhase_Total_pav  = np.einsum('cpVnr,tcprV->tcprnV', RECC_System.ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[:,:,:,:,:,mS], SysVar_StockServiceProvision_UsePhase_pav)
            # SysVar_EnergyDemand_UsePhase_Total_reb  = np.einsum('cBVnrt,tcBrV->tcBrnV',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values, SysVar_StockServiceProvision_UsePhase_reb)
            if 'nrb' in SectorList: 
                None 
                # SysVar_EnergyDemand_UsePhase_Total_nrb  = np.einsum('cNVnrt,tcNrV->tcNrnV',RECC_System.ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values, SysVar_StockServiceProvision_UsePhase_nrb)
            if 'nrbg' in SectorList: 
                None
                # SysVar_EnergyDemand_UsePhase_Total_nrbg = np.zeros((Nt,Nc,NN,No,Nn,NV))
            # Unit: TJ/yr for both vehicles and buildings.
            
            # C) Translate 'all' energy carriers to specific ones, use phase, by sector
            #Hop over to save computation time:
            # SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav = np.einsum('cprVn,tcprV->trpn',RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles' ].Values[:,:,:,:,:,mS] ,SysVar_EnergyDemand_UsePhase_Total_pav[:,:,:,:,-1,:]
            # SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb =  np.einsum('Vrnt,tcBrV->trBn', RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values[:,mR,:,:,:],   SysVar_EnergyDemand_UsePhase_Total_reb[:,0:SwitchTime,:,:,-1,:]
            # SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb += np.einsum('Vrnt,tcBrV->trBn', RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,:,:,:,mS],      SysVar_EnergyDemand_UsePhase_Total_reb[:,SwitchTime::,:,:,-1,:]
            # SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all      = np.zeros((Nt,Nn))
            # if 'pav' in SectorList:
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav  = np.einsum('cprVn,cpVr,Vrt,tcpr->trpn',RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles' ].Values[:,:,:,:,:,mS],RECC_System.ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[:,:,:,-1,:,mS],RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[:,:,:,mS],Stock_Detail_UsePhase_p, optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByService_pav        = np.einsum('cprVn,cpVr,Vrt,tcpr->trV', RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Vehicles' ].Values[:,:,:,:,:,mS],RECC_System.ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[:,:,:,-1,:,mS],RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[:,:,:,mS],Stock_Detail_UsePhase_p, optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all += np.einsum('trpn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            # else:
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav  = np.zeros((Nt,Nr,Np,Nn))
            #     SysVar_EnergyDemand_UsePhase_ByService_pav        = np.zeros((Nt,Nr,NV))
            # if 'reb' in SectorList:
            #     # Note that for hist. c, 3_EI_Products_UsePhase_resbuildings_t contains EI for final energy (F_16_7), and for future c, it contains useful energy (F_7_0a).
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb  = np.einsum('Vrnt,cBVrt,tcBVr,tcBr->trBn', RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values[:,mR,:,:,:],ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,:,:,mS],Stock_Detail_UsePhase_B[:,0:SwitchTime,:,:], optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb += np.einsum('Vrnt,cBVrt,tcBVr,tcBr->trBn', RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,:,:,:,mS],ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime::,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,:,:,mS],Stock_Detail_UsePhase_B[:,SwitchTime::,:,:], optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByService_reb        = np.einsum('Vrnt,cBVrt,tcBVr,tcBr->trV',  RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values[:,mR,:,:,:],ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,:,:,mS],Stock_Detail_UsePhase_B[:,0:SwitchTime,:,:], optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByService_reb       += np.einsum('Vrnt,cBVrt,tcBVr,tcBr->trV',  RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,:,:,:,mS],ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime::,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,:,:,mS],Stock_Detail_UsePhase_B[:,SwitchTime::,:,:], optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all += np.einsum('trpn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
            #     # Diagnose variables
    #            Aa_S       = np.einsum('tcBr->t', Stock_Detail_UsePhase_B) # total stock
    #            Aa_S_H     = np.einsum('tcBr,tcBr->t',RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,0,:,mS],Stock_Detail_UsePhase_B) # total heated stock
    #            Aa_S_C     = np.einsum('tcBr,tcBr->t',RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,1,:,mS],Stock_Detail_UsePhase_B) # total cooled stock
    #            Aa_E_usef  = np.einsum('cBVrt,tcBVr,tcBr->t', ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[:,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,:,:,:,:,mS],Stock_Detail_UsePhase_B[:,:,:,:], optimize = True)
    #            Aa_E_U_hist= np.einsum('cBVrt,tcBVr,tcBr->t', ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,:,:,mS],Stock_Detail_UsePhase_B[:,0:SwitchTime,:,:], optimize = True)
    #            Aa_E_U_fut = np.einsum('cBVrt,tcBVr,tcBr->tV', ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime::,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,:,:,mS],Stock_Detail_UsePhase_B[:,SwitchTime::,:,:], optimize = True)
    #            Aa_E_final = np.einsum('trBn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
    #            Aa_E_f_hist= np.einsum('Vrnt,cBVrt,tcBVr,tcBr->t', RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_Buildings'].Values[:,mR,:,:,:],ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[0:SwitchTime,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,0:SwitchTime,:,:,:,mS],Stock_Detail_UsePhase_B[:,0:SwitchTime,:,:], optimize = True)
    #            Aa_E_f_fut = np.einsum('Vrnt,cBVrt,tcBVr,tcBr->tV', RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,:,:,:,mS],ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime::,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,:,:,mS],Stock_Detail_UsePhase_B[:,SwitchTime::,:,:], optimize = True)
    #            # assume that energy carrier split changes with age-cohort rather than with time
    #            Aa_E_f_futc= np.einsum('Vrnc,cBVrt,tcBVr,tcBr->tV', RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,:,:,1::,mS],ParameterDict['3_EI_Products_UsePhase_resbuildings_t'].Values[SwitchTime::,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_Buildings_UsePhase'].Values[:,SwitchTime::,:,:,:,mS],Stock_Detail_UsePhase_B[:,SwitchTime::,:,:], optimize = True)
    #            Aa_sha_fut = np.einsum('Vrnt->Vnt', RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,:,:,:,mS])
    #            Aa_sha     = RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,0,:,:,mS]
    #            Aa_test    = Aa_E_f_fut[:,1] / Aa_E_U_fut[:,1]
            # else:
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb  = np.zeros((Nt,Nr,NB,Nn))
            #     SysVar_EnergyDemand_UsePhase_ByService_reb        = np.zeros((Nt,Nr,NV))
            # if 'nrb' in SectorList: 
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb  = np.einsum('Vrnt,cNVrt,cNVr,tcNr->trNn', RECC_System.ParameterDict['3_SHA_EnergyCarrierSplit_NonResBuildings'].Values[:,mR,:,:,:],ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values[0:SwitchTime,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[0:SwitchTime,:,:,:,mS],Stock_Detail_UsePhase_N[:,0:SwitchTime,:,:], optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb += np.einsum('Vrnt,cNVrt,cNVr,tcNr->trNn', RECC_System.ParameterDict['3_SHA_EnergySupply_Buildings'].Values[:,mR,:,:,:,mS],ParameterDict['3_EI_Products_UsePhase_nonresbuildings_t'].Values[SwitchTime::,:,:,-1,:,:],RECC_System.ParameterDict['3_IO_NonResBuildings_UsePhase'].Values[SwitchTime::,:,:,:,mS],Stock_Detail_UsePhase_N[:,SwitchTime::,:,:], optimize = True)
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all += np.einsum('trpn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb)
            # else:
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb  = np.zeros((Nt,Nr,NN,Nn))
            #     #SysVar_EnergyDemand_UsePhase_ByService_nrb        = np.zeros((Nt,Nr,NV))
            # if 'nrbg' in SectorList:     
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg = np.zeros((Nt,No,NN,Nn)) # Not yet quantified!
            # else:
            #     SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg = np.zeros((Nt,No,NN,Nn))
                
            # # D) Calculate energy demand of the other industries
            # SysVar_EnergyDemand_PrimaryProd    = 1000 * np.einsum('mnt,tm->tmn',RECC_System.ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,:,0,mR],RECC_System.FlowDict['F_3_4'].Values[:,:,0])
            # SysVar_EnergyDemand_Manufacturing  = np.zeros((Nt,Nn))
            # # all in TJ/yr
            # if 'pav' in SectorList:
            #     SysVar_EnergyDemand_Manufacturing += np.einsum('pn,tpr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_pav_rge,:,110,-1],Inflow_Detail_UsePhase_p)        # conversion factor: 1, as MJ/item  = TJ/Million items.
            # if 'reb' in SectorList:
            #     SysVar_EnergyDemand_Manufacturing += np.einsum('Bn,tBr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_reb_rge,:,110,-1],Inflow_Detail_UsePhase_B)        # conversion factor: 1, as MJ/m²    = TJ/Million m².
            # if 'nrb' in SectorList:
            #     SysVar_EnergyDemand_Manufacturing += np.einsum('Nn,tNr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_nrb_rge,:,110,-1],Inflow_Detail_UsePhase_N)        # conversion factor: 1, as MJ/m²    = TJ/Million m².
            # if 'nrbg' in SectorList:
            #     SysVar_EnergyDemand_Manufacturing += np.einsum('Nn,tNr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_nrbg_rge,:,110,-1],Inflow_Detail_UsePhase_N)        # conversion factor: 1, as MJ/m²    = TJ/Million m².                
            # if 'ind' in SectorList:
            #     SysVar_EnergyDemand_Manufacturing += np.einsum('In,tIr->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_ind_rge,:,110,-1],Inflow_Detail_UsePhase_I) * 1e-6 # conversion factor: 1, as TJ/GW    = 10e-6 MJ/GW. 
            # if 'app' in SectorList:
            #     SysVar_EnergyDemand_Manufacturing += np.einsum('an,tar->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_app_rge,:,110,-1],Inflow_Detail_UsePhase_a) * 1e-6 # conversion factor: 1, as TJ/item  = 10e-6 MJ/items. 
            # if 'otv' in SectorList:
            #     SysVar_EnergyDemand_Manufacturing += np.einsum('an,tar->tn',RECC_System.ParameterDict['4_EI_ManufacturingEnergyIntensity'].Values[Sector_otv_rge,:,110,-1],Inflow_Detail_UsePhase_otv) * 1e-9 # conversion factor: 1, as TJ/item  = 10e-6 MJ/items. 
                  
            
            
            
            # SysVar_EnergyDemand_WasteMgt       = 1000 * (np.einsum('wn,trw->tn',RECC_System.ParameterDict['4_EI_WasteMgtEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_10'].Values[:,:,:,0]) +\
            #                                             np.einsum('wn,trw->tn',RECC_System.ParameterDict['4_EI_WasteMgtEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_10_Nl'].Values[:,:,:,0]) +\
            #                                             np.einsum('wn,trw->tn',RECC_System.ParameterDict['4_EI_WasteMgtEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_10_No'].Values[:,:,:,0]))
            # SysVar_EnergyDemand_Remelting      = 1000 * np.einsum('mn,trm->tn',RECC_System.ParameterDict['4_EI_RemeltingEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_12'].Values[:,:,:,0])
            # SysVar_EnergyDemand_Remelting_m    = 1000 * np.einsum('mn,trm->tnm',RECC_System.ParameterDict['4_EI_RemeltingEnergyIntensity'].Values[:,:,110,-1],RECC_System.FlowDict['F_9_12'].Values[:,:,:,0])
            # SysVar_EnergySavings_WasteToEnerg  = np.zeros((Nt,Nn))
            # if  mR == ClimPolScen: # in the climate policy scenario:
            #     SysVar_EnergySavings_WasteToEnerg[:,Electric_loc] = EnergyRecovery_WoodCombustion_EL[:,mS,mR].copy()
            # # Unit: TJ/yr.
            
            # Calculate total energy demand by individual energy carrier
            # SysVar_TotalEnergyDemand = np.einsum('trpn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav) + np.einsum('trBn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb) \
            # + np.einsum('trNn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb) + np.einsum('toNn->tn',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg) \
            # + np.einsum('tmn->tn',SysVar_EnergyDemand_PrimaryProd) + SysVar_EnergyDemand_Manufacturing + SysVar_EnergyDemand_WasteMgt + SysVar_EnergyDemand_Remelting
            # # Calculate total energy demand for all energy carriers together. ONLY CORRECT if 'all' energy carriers are at last location -1!
            # SysVar_TotalEnergyDemand[:,-1]                         = SysVar_TotalEnergyDemand[:,0:-1].sum(axis=1)
            # SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all[:,-1] = SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all[:,0:-1].sum(axis=1)
            # # Unit: TJ/yr.
            
            
            # if 'C ' in Element_List:  
            #     # E) Calculate carbon flows and stocks in forestry.
            #     # a) energy in fuel wood in TJ/yr
            #     SysVar_Energy_FuelWood_t = np.einsum('trp->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav[:,:,:,WoodFuel_loc]) + np.einsum('trB->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb[:,:,:,WoodFuel_loc]) + np.einsum('trN->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb[:,:,:,WoodFuel_loc]) + np.einsum('toN->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg[:,:,:,WoodFuel_loc])
            #     # carbon flows in Mt/yr
            #     RECC_System.FlowDict['F_2_7'].Values[:,Carbon_loc] = SysVar_Energy_FuelWood_t / RECC_System.ParameterDict['3_EI_HeatingValueWoodPerCarbon'].Values[Carbon_loc,WoodFuel_loc] / 1000 # carbon only!
            #     RECC_System.FlowDict['F_7_0'].Values               = RECC_System.FlowDict['F_2_7'].Values.copy() # This is for mass balance only. The CO2 emissions from burning fuel wood in the use phase are accounted for by the direct emissions already.
            #     RECC_System.FlowDict['F_1_2'].Values               = RECC_System.FlowDict['F_2_7'].Values + np.einsum('tme->te',RECC_System.FlowDict['F_2_3'].Values)
            #     # Quantify wood carbon stock change (through harvested wood and its regrowth only, no soil carbon balance, albedo, etc.)
            #     # only the forest carbon pool change relative to the baseline is quantified, not the total forest carbon stock.
            #     CarbonTimberHarvest = np.einsum('crg->c',RECC_System.StockDict['S_7'].Values[0,:,:,:,Wood_loc,Carbon_loc]) # Historic age-cohorts
            #     CarbonTimberHarvest[SwitchTime::] = RECC_System.FlowDict['F_2_3'].Values[1::,Wood_loc,Carbon_loc]
            #     Forest_GrowthTable_timber = np.zeros((Nc,Nc))
            #     np.fill_diagonal(Forest_GrowthTable_timber, -1* CarbonTimberHarvest)
            #     Forest_GrowthTable_fuelwd = np.zeros((Nt,Nt))
            #     np.fill_diagonal(Forest_GrowthTable_fuelwd, -1* RECC_System.FlowDict['F_2_7'].Values[:,Carbon_loc])
            #     # We also take into account for the regrowth of forest attributable to all wood present in the 2015 stock.
            #     RegrowthCurve_Timber = scipy.stats.norm.cdf(np.arange(0,Nc,1),RECC_System.ParameterDict['3_LT_ForestRotationPeriod_Timber'  ].Values[Wood_loc]    /2, RECC_System.ParameterDict['3_LT_ForestRotationPeriod_Timber'  ].Values[Wood_loc]    /4)
            #     RegrowthCurve_FuelWo = scipy.stats.norm.cdf(np.arange(0,Nt,1),RECC_System.ParameterDict['3_LT_ForestRotationPeriod_FuelWood'].Values[WoodFuel_loc]/2, RECC_System.ParameterDict['3_LT_ForestRotationPeriod_FuelWood'].Values[WoodFuel_loc]/4)
            #     # Scale regrowth curves and insert them into growth tables:
            #     for nnc in range(0,Nc):
            #         Forest_GrowthTable_timber[nnc+1::,nnc] = Forest_GrowthTable_timber[nnc,nnc] * (1 - RegrowthCurve_Timber[0:Nc-nnc-1])
            #     for nnt in range(0,Nt):
            #         Forest_GrowthTable_fuelwd[nnt+1::,nnt] = Forest_GrowthTable_fuelwd[nnt,nnt] * (1 - RegrowthCurve_FuelWo[0:Nt-nnt-1])
        
            #     # Assign growth table values to stock and stock changes in process 1:
        
            #     RECC_System.StockDict['S_1t'].Values[:,:,Carbon_loc]              = Forest_GrowthTable_timber[SwitchTime-1::,:]
            #     RECC_System.StockDict['S_1f'].Values[:,SwitchTime-1::,Carbon_loc] = Forest_GrowthTable_fuelwd 
                            
            #     RECC_System.StockDict['dS_1t'].Values[:,Carbon_loc]   = Forest_GrowthTable_timber.sum(axis=1)[SwitchTime-1::]
            #     RECC_System.StockDict['dS_1t'].Values[1::,Carbon_loc] = np.diff(Forest_GrowthTable_timber.sum(axis=1))[SwitchTime-1::]
            #     RECC_System.StockDict['dS_1f'].Values[:,Carbon_loc]   = Forest_GrowthTable_fuelwd.sum(axis=1).copy()
            #     RECC_System.StockDict['dS_1f'].Values[1::,Carbon_loc] = np.diff(Forest_GrowthTable_fuelwd.sum(axis=1)).copy()
        
            #     RECC_System.FlowDict['F_0_1'].Values[:,Carbon_loc]    = RECC_System.FlowDict['F_1_2'].Values[:,Carbon_loc] + RECC_System.StockDict['dS_1t'].Values[:,Carbon_loc] + RECC_System.StockDict['dS_1f'].Values[:,Carbon_loc]
            #     # This flow has a large negative initial value due to the boundary conditions. The value for year 0 is set to 0 in the results of the calculations using this system variable, as year 0 is for initialisation only.
               
              
         
            
    #        
    #        
    #        
            # F) Check whether flow value arrays match their indices, etc.
            RECC_System.Consistency_Check() 
    #    
            # G) Determine Mass Balance
            # Commment out to save computation time:
            #BalAbs = -1 # means that mass bal. computation was commented out to save computation time.
            Bal = RECC_System.MassBalance()
            BalAbs = np.abs(Bal).sum()
            Mylog.info('Total mass balance deviation (np.abs(Bal).sum() for socioeconomic scenario ' + SName + ' and RE scenario ' + RName + ': ' + str(BalAbs) + ' Mt.')                    
    
            
    
            # H) Calculate direct emissions by combustion of energy carriers in processes
    
    
            # if 'C ' in Element_List:  
        
            #     SysVar_DirectEmissions_UsePhase_Vehicles    = 0.001 * np.einsum('Xn,trpn->Xtrp',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            #     SysVar_DirectEmissions_UsePhase_Buildings   = 0.001 * np.einsum('Xn,trBn->XtrB',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
            #     SysVar_DirectEmissions_UsePhase_NRBuildgs   = 0.001 * np.einsum('Xn,trNn->XtrN',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb)
            #     SysVar_DirectEmissions_UsePhase_NRBuildgs_g = 0.001 * np.einsum('Xn,toNn->XtoN',RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg)
            #     SysVar_DirectEmissions_PrimaryProd          = 0.001 * np.einsum('Xn,tmn->Xtm'  ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_PrimaryProd)
            #     SysVar_DirectEmissions_Manufacturing        = 0.001 * np.einsum('Xn,tn->Xt'    ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_Manufacturing)
            #     # SysVar_DirectEmissions_WasteMgt             = 0.001 * np.einsum('Xn,tn->Xt'    ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_WasteMgt)
            #     # SysVar_DirectEmissions_Remelting            = 0.001 * np.einsum('Xn,tn->Xt'    ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_Remelting)
            #     # SysVar_DirectEmissions_Remelting_m          = 0.001 * np.einsum('Xn,tnm->Xtm'  ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_Remelting_m)
            #     # Unit: Mt/yr. 1 kg/MJ = 1kt/TJ
                
            #     # I) Calculate process emissions
            #     SysVar_ProcessEmissions_PrimaryProd         = np.einsum('mXt,tm->Xt'    ,RECC_System.ParameterDict['4_PE_ProcessExtensions'].Values[:,:,0,:,mR,mS],RECC_System.FlowDict['F_3_4'].Values[:,:,0])
            #     SysVar_ProcessEmissions_PrimaryProd_m       = np.einsum('mXt,tm->Xtm'   ,RECC_System.ParameterDict['4_PE_ProcessExtensions'].Values[:,:,0,:,mR,mS],RECC_System.FlowDict['F_3_4'].Values[:,:,0])
            #     # Unit: Mt/yr.
                
            #     # J) Calculate emissions from energy supply
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings       = 0.001 * np.einsum('Xnrt,trBn->Xt',RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb) \
            #                                                                    + 0.001 * np.einsum('Xnrt,trNn->Xt',RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb) \
            #                                                                    + 0.001 * np.einsum('Xnt,toNn->Xt', RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg)
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles           = 0.001 * np.einsum('Xnrt,trpn->Xt',RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings_EL    = 0.001 * np.einsum('Xrt,trB->Xt',  RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,0,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb[:,:,:,0]) \
            #                                                                    + 0.001 * np.einsum('Xrt,trN->Xt',  RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,0,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb[:,:,:,0]) \
            #                                                                    + 0.001 * np.einsum('Xt,toN->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,0,mS,mR,0,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg[:,:,:,0]) # electricity only
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles_EL        = 0.001 * np.einsum('Xrt,trp->Xt',  RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,0,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav[:,:,:,0])   # electricity only
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings_Ot    = SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings - SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings_EL
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles_Ot        = SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles     - SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles_EL
            #     if Nr > 1:
            #         SysVar_IndirectGHGEms_EnergySupply_PrimaryProd             = 0.001 * np.einsum('Xnt,tmn->Xt',  RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_PrimaryProd)
            #         SysVar_IndirectGHGEms_EnergySupply_PrimaryProd_m           = 0.001 * np.einsum('Xnt,tmn->Xtm', RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_PrimaryProd)
            #         SysVar_IndirectGHGEms_EnergySupply_Manufacturing           = 0.001 * np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_Manufacturing)
            #         SysVar_IndirectGHGEms_EnergySupply_WasteMgt                = 0.001 * np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_WasteMgt)
            #         SysVar_IndirectGHGEms_EnergySupply_Remelting               = 0.001 * np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_Remelting)
            #         SysVar_IndirectGHGEms_EnergySupply_Remelting_m             = 0.001 * np.einsum('Xnt,tnm->Xtm', RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_Remelting_m)
            #         SysVar_IndirectGHGEms_EnergySupply_WasteToEnergy           = -0.001* np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergySavings_WasteToEnerg)
            #     else:
            #         SysVar_IndirectGHGEms_EnergySupply_PrimaryProd             = 0.001 * np.einsum('Xnt,tmn->Xt',  RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_PrimaryProd)
            #         SysVar_IndirectGHGEms_EnergySupply_PrimaryProd_m           = 0.001 * np.einsum('Xnt,tmn->Xtm', RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_PrimaryProd)
            #         SysVar_IndirectGHGEms_EnergySupply_Manufacturing           = 0.001 * np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_Manufacturing)
            #         SysVar_IndirectGHGEms_EnergySupply_WasteMgt                = 0.001 * np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_WasteMgt)
            #         SysVar_IndirectGHGEms_EnergySupply_Remelting               = 0.001 * np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_Remelting)
            #         SysVar_IndirectGHGEms_EnergySupply_Remelting_m             = 0.001 * np.einsum('Xnt,tnm->Xtm', RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_Remelting_m)
            #         SysVar_IndirectGHGEms_EnergySupply_WasteToEnergy           = -0.001* np.einsum('Xnt,tn->Xt',   RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,0,:],SysVar_EnergySavings_WasteToEnerg)
            #     # Unit: Mt/yr.
                
            #     # Calculate emissions by energy carrier:
            #     SysVar_DirectEmissions_UsePhase_Vehicles_n                     = 0.001 * np.einsum('Xn,trpn->Xtrn',  RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
            #     SysVar_DirectEmissions_UsePhase_ResBuildings_n                 = 0.001 * np.einsum('Xn,trBn->Xtrn',  RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_ResBuildings_n     = 0.001 * np.einsum('Xnrt,trBn->Xtrn',RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb) 
            #     SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles_n         = 0.001 * np.einsum('Xnrt,trpn->Xtrn',RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply'].Values[:,:,mS,mR,:,:],SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)
                
            #     # K) Calculate emissions benefits
            #     if ScriptConfig['ScrapExportRecyclingCredit'] == 'True':
            #         SysVar_EnergyDemand_RecyclingCredit                = -1 * 1000 * np.einsum('mnt,tm->tmn' ,RECC_System.ParameterDict['4_EI_ProcessEnergyIntensity'].Values[:,:,:,0,mR],RECC_System.FlowDict['F_12_21'].Values[:,-1,:,0])
            #         SysVar_DirectEmissions_RecyclingCredit             = -1 * 0.001 * np.einsum('Xn,tmn->Xt' ,RECC_System.ParameterDict['6_PR_DirectEmissions'].Values,SysVar_EnergyDemand_RecyclingCredit)
            #         SysVar_ProcessEmissions_RecyclingCredit            = -1 * np.einsum('mXt,tm->Xt'         ,RECC_System.ParameterDict['4_PE_ProcessExtensions'].Values[:,:,0,:,mR,mS],RECC_System.FlowDict['F_12_21'].Values[:,-1,:,0])
            #         SysVar_IndirectGHGEms_EnergySupply_RecyclingCredit = -1 * 0.001 * np.einsum('Xnt,tmn->Xt',RECC_System.ParameterDict['4_PE_GHGIntensityEnergySupply_World'].Values[:,:,mS,mR,0,:],SysVar_EnergyDemand_RecyclingCredit)
            #     else:
            #         SysVar_EnergyDemand_RecyclingCredit                = np.zeros((Nt,Nm,Nn))
            #         SysVar_DirectEmissions_RecyclingCredit             = np.zeros((NX,Nt))
            #         SysVar_ProcessEmissions_RecyclingCredit            = np.zeros((NX,Nt))
            #         SysVar_IndirectGHGEms_EnergySupply_RecyclingCredit = np.zeros((NX,Nt))
                    
            #     # L) GWP_bio calculation, using the original RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values
            #     # and NOT the extended lifetime.
            #     SysVar_GHGEms_GWP_bio_r = np.zeros((NX,Nt,Nr))
            #     SysVar_GHGEms_GWP_bio_o = np.zeros((NX,Nt))
            #     if 'reb' in SectorList:
            #         for ntt in range(0,Nt):
            #             for nrr in range(0,Nr):
            #                 for nBB in range(0,NB):
            #                     #total carbon in wood (carbon content ca. 0.5)
            #                     mass_C  = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] *12/44 * RECC_System.FlowDict['F_6_7_Nr'].Values[ntt,nrr,Sector_reb_rge[nBB],9,0]
            #                     GWP_bio = RECC_System.ParameterDict['6_MIP_GWP_Bio'].Values[int(np.floor(RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_resbuildings'].Values[nBB,nrr,ntt+SwitchTime-1]))]
            #                     SysVar_GHGEms_GWP_bio_r[0,ntt,nrr] += 44/12 * mass_C * GWP_bio # convert from C to CO2
            #     if 'nrb' in SectorList:
            #         for ntt in range(0,Nt):
            #             for nrr in range(0,Nr):
            #                 for nNN in range(0,NN):
            #                     #total carbon in wood (carbon content ca. 0.5)
            #                     mass_C  = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] *12/44 * RECC_System.FlowDict['F_6_7_Nr'].Values[ntt,nrr,Sector_nrb_rge[nNN],9,0]
            #                     GWP_bio = RECC_System.ParameterDict['6_MIP_GWP_Bio'].Values[int(np.floor(RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_NonResbuildings'].Values[nNN,nrr,ntt+SwitchTime-1]))]
            #                     SysVar_GHGEms_GWP_bio_r[0,ntt,nrr] += 44/12 * mass_C * GWP_bio # convert from C to CO2                        
            #     if 'nrbg' in SectorList:
            #         for ntt in range(0,Nt):
            #             for nNN in range(0,NN):
            #                 #total carbon in wood (carbon content ca. 0.5)
            #                 mass_C  = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] *12/44 * RECC_System.FlowDict['F_6_7_No'].Values[ntt,0,Sector_nrbg_rge_reg[nNN],9,0]
            #                 GWP_bio = RECC_System.ParameterDict['6_MIP_GWP_Bio'].Values[int(np.floor(RECC_System.ParameterDict['3_LT_RECC_ProductLifetime_nonresbuildings_g'].Values[nNN,0,ntt+SwitchTime-1]))]
            #                 SysVar_GHGEms_GWP_bio_o[0,ntt] += 44/12 * mass_C * GWP_bio # convert from C to CO2                        
            #     SysVar_GHGEms_GWP_bio = np.einsum('Xtr->Xt',SysVar_GHGEms_GWP_bio_r) + SysVar_GHGEms_GWP_bio_o
            #     # not used anymore! Have time and process-explicit carbon flow and stock accounting now.
                
            #     SysVar_CO2UptakeEmissions_Forests = np.zeros((NX,Nt,Nm))
            #     SysVar_CO2UptakeEmissions_Forests[CO2_loc,:,Wood_loc] = -1 * 44/12 * RECC_System.FlowDict['F_0_1'].Values[:,Carbon_loc] # negative sign because emissions are measured in X_0 direction.
            #     SysVar_CO2UptakeEmissions_Forests[:,0,:] = 0
            #     # F_0_1 has a large negative initial value due to the boundary conditions. The value for year 0 is set to 0 in the results of the calculations using this system variable, as year 0 is for initialisation only.
                
            #     # M) Calculate emissions of system, by process group, INCLUDING GWPbio and credits
            #     # Number indicates the process number of the ODYM-RECC system definition
            #     # 'd' behind the number indicates direct, 'i' indirect emissions of that process.
            #     SysVar_GHGEms_UsePhase_7d              = np.einsum('XtrB->Xt',SysVar_DirectEmissions_UsePhase_Buildings) + np.einsum('XtrN->Xt',SysVar_DirectEmissions_UsePhase_NRBuildgs) + np.einsum('XtoN->Xt',SysVar_DirectEmissions_UsePhase_NRBuildgs_g) + np.einsum('Xtrp->Xt',SysVar_DirectEmissions_UsePhase_Vehicles)
            #     SysVar_GHGEms_UsePhase_7i_Scope2_El    = SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings_EL + SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles_EL
            #     SysVar_GHGEms_UsePhase_7i_OtherIndir   = SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings_Ot + SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles_Ot
            #     SysVar_GHGEms_PrimaryMaterial_3di      = np.einsum('Xtm->Xt',SysVar_DirectEmissions_PrimaryProd) + SysVar_ProcessEmissions_PrimaryProd + SysVar_IndirectGHGEms_EnergySupply_PrimaryProd
            #     SysVar_GHGEms_PrimaryMaterial_3di_m    = SysVar_DirectEmissions_PrimaryProd + SysVar_ProcessEmissions_PrimaryProd_m + SysVar_IndirectGHGEms_EnergySupply_PrimaryProd_m
            #     SysVar_GHGEms_Manufacturing_5di        = SysVar_DirectEmissions_Manufacturing + SysVar_IndirectGHGEms_EnergySupply_Manufacturing
            #     SysVar_GHGEms_WasteMgtRemelting_9di    = SysVar_DirectEmissions_WasteMgt + SysVar_DirectEmissions_Remelting + SysVar_IndirectGHGEms_EnergySupply_WasteMgt + SysVar_IndirectGHGEms_EnergySupply_Remelting
            #     SysVar_GHGEms_MaterialCycle_5di_9di    = SysVar_GHGEms_Manufacturing_5di + SysVar_GHGEms_WasteMgtRemelting_9di
                                                     
            #     SysVar_GHGEms_RecyclingCredit          = SysVar_DirectEmissions_RecyclingCredit + SysVar_ProcessEmissions_RecyclingCredit + SysVar_IndirectGHGEms_EnergySupply_RecyclingCredit
            #     SysVar_GHGEms_EnergyRecoveryWaste_9di  = np.zeros((NX,Nt))
            #     SysVar_GHGEms_EnergyRecoveryWaste_9di[CO2_loc,:] = BiogenicCO2WasteCombustion[t,mS,mR].copy()
            #     SysVar_GHGEms_EnergyRecoveryWaste_9di += SysVar_IndirectGHGEms_EnergySupply_WasteToEnergy
            #     # Calculate total emissions of system
            #     SysVar_GHGEms_OtherThanUsePhaseDirect  = SysVar_GHGEms_UsePhase_7i_Scope2_El + SysVar_GHGEms_UsePhase_7i_OtherIndir + SysVar_GHGEms_PrimaryMaterial_3di + SysVar_GHGEms_MaterialCycle_5di_9di
            #     SysVar_TotalGHGEms_3579di              = SysVar_GHGEms_UsePhase_7d + SysVar_GHGEms_OtherThanUsePhaseDirect + SysVar_GHGEms_EnergyRecoveryWaste_9di + np.einsum('Xtm->Xt',SysVar_CO2UptakeEmissions_Forests)
        
            #     SysVar_GHGEms_Materials_3di_9di        = SysVar_GHGEms_PrimaryMaterial_3di + SysVar_GHGEms_WasteMgtRemelting_9di
            #     # Unit: Mt/yr.
                
            #     # N) Calculate indicators
            #     #SysVar_TotalGHGCosts                   = np.einsum('t,Xt->Xt',RECC_System.ParameterDict['3_PR_RECC_CO2Price_SSP_32R'].Values[mR,:,m_reg_o,mS],SysVar_TotalGHGEms_3579di)
            #     # Unit: million $ / yr.
                
            #     # O) Compile results
            #     # Emissions breakdown by system processes
            #     GWP_System_3579di[:,mS,mR]                  = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_TotalGHGEms_3579di)[GWP100_loc,:].copy()
            #     GWP_UsePhase_7d[:,mS,mR]                    = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_UsePhase_7d)[GWP100_loc,:].copy()
            #     GWP_UsePhase_7i_Scope2_El[:,mS,mR]          = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_UsePhase_7i_Scope2_El)[GWP100_loc,:].copy()
            #     GWP_UsePhase_7i_OtherIndir[:,mS,mR]         = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_UsePhase_7i_OtherIndir)[GWP100_loc,:].copy()
            #     GWP_MaterialCycle_5di_9di[:,mS,mR]          = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_MaterialCycle_5di_9di)[GWP100_loc,:].copy()
            #     GWP_RecyclingCredit[:,mS,mR]                = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_RecyclingCredit)[GWP100_loc,:].copy()
            #     GWP_ForestCO2Uptake[:,mS,mR]                = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,np.einsum('Xtm->Xt',SysVar_CO2UptakeEmissions_Forests))[GWP100_loc,:].copy()
            #     GWP_EnergyRecoveryWasteWood[:,mS,mR]        = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_EnergyRecoveryWaste_9di)[GWP100_loc,:].copy()
            #     GWP_OtherThanUsePhaseDirect[:,mS,mR]        = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_OtherThanUsePhaseDirect)[GWP100_loc,:].copy() # all non use-phase processes
            #     GWP_Materials_3di_9di[:,mS,mR]              = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_Materials_3di_9di)[GWP100_loc,:].copy()
            #     GWP_Vehicles_Direct[:,:,mS,mR]              = np.einsum('xX,Xtrp->xtr'  ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_DirectEmissions_UsePhase_Vehicles)[GWP100_loc,:,:].copy()
            #     GWP_ReBuildgs_Direct[:,:,mS,mR]             = np.einsum('xX,XtrB->xtr'  ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_DirectEmissions_UsePhase_Buildings)[GWP100_loc,:,:].copy()
            #     GWP_NRBuildgs_Direct[:,:,mS,mR]             = np.einsum('xX,XtrN->xtr'  ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_DirectEmissions_UsePhase_NRBuildgs)[GWP100_loc,:,:].copy()
            #     GWP_NRBuildgs_Direct_g[:,mS,mR]             = np.einsum('xX,XtoN->xt'   ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_DirectEmissions_UsePhase_NRBuildgs_g)[GWP100_loc,:].copy()
            #     GWP_PrimaryMaterial_3di[:,mS,mR]            = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_PrimaryMaterial_3di)[GWP100_loc,:].copy()
            #     GWP_PrimaryMaterial_3di_m[:,:,mS,mR]        = np.einsum('xX,Xtm->xtm'   ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_PrimaryMaterial_3di_m)[GWP100_loc,:,:].copy()
            #     GWP_Manufact_5di_all[:,mS,mR]               = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_Manufacturing_5di)[GWP100_loc,:].copy()
            #     GWP_WasteMgt_9di_all[:,mS,mR]               = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_GHGEms_WasteMgtRemelting_9di)[GWP100_loc,:].copy()
            #     # other emissions breakdown
            #     GWP_SecondaryMetal_di_m[:,:,mS,mR]          = np.einsum('xX,Xtm->xtm'   ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_DirectEmissions_Remelting_m + SysVar_IndirectGHGEms_EnergySupply_Remelting_m)[GWP100_loc,:,:].copy()
            #     GWP_Vehicles_indir[:,mS,mR]                 = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles)[GWP100_loc,:].copy()
            #     GWP_AllBuildings_indir[:,mS,mR]             = np.einsum('xX,Xt->xt'     ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_IndirectGHGEms_EnergySupply_UsePhase_AllBuildings)[GWP100_loc,:].copy()
            #     GWP_ByEnergyCarrier_UsePhase_d[:,:,:,mS,mR] = np.einsum('xX,Xtrn->xtrn' ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_DirectEmissions_UsePhase_Vehicles_n + SysVar_DirectEmissions_UsePhase_ResBuildings_n)[GWP100_loc,:,:,:].copy()
            #     GWP_ByEnergyCarrier_UsePhase_i[:,:,:,mS,mR] = np.einsum('xX,Xtrn->xtrn' ,RECC_System.ParameterDict['6_MIP_CharacterisationFactors'].Values,SysVar_IndirectGHGEms_EnergySupply_UsePhase_Vehicles_n + SysVar_IndirectGHGEms_EnergySupply_UsePhase_ResBuildings_n)[GWP100_loc,:,:,:].copy()
            #     GWP_WoodCycle[:,mS,mR]                      = GWP_ForestCO2Uptake[:,mS,mR].copy() + GWP_EnergyRecoveryWasteWood[:,mS,mR].copy() # net wood use emissions, not exported.
            # Mass flows
            if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
                 Material_Inflow[:,Sector_32reg_rge,:,mS,mR]                = np.einsum('trgm->gtm',RECC_System.FlowDict['F_6_7_Nr'].Values[:,:,:,:,0]).copy()
           
                 Material_Outflow[:,Sector_32reg_rge,:,mS,mR]                = np.einsum('trgm->gtm',RECC_System.FlowDict['F_7_8_Nr'].Values[:,:,:,:,:,0].sum(axis=1))

            if 'ind' in SectorList:
                Material_Inflow[:,Sector_11reg_rge,:,mS,mR] = np.einsum('tlLm->Ltm',RECC_System.FlowDict['F_6_7_Nl'].Values[:,:,:,:,0]).copy()
          
                Material_Outflow[:,Sector_11reg_rge,:,mS,mR] = np.einsum('tlLm->Ltm',RECC_System.FlowDict['F_7_8_Nl'].Values[:,:,:,:,:,0].sum(axis=1))
  
          
            if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
                Material_Inflow[:,Sector_1reg_rge,:,mS,mR]  = np.einsum('toOm->Otm',RECC_System.FlowDict['F_6_7_No'].Values[:,:,:,:,0]).copy()    
                
                Material_Outflow[:,Sector_1reg_rge,:,mS,mR]  = np.einsum('toOm->Otm',RECC_System.FlowDict['F_7_8_No'].Values[:,:,:,:,:,0].sum(axis=1))   

                
                
                
            Scrap_Outflow[:,:,mS,mR]                    = np.einsum('trw->tw',RECC_System.FlowDict['F_9_10'].Values[:,:,:,0]).copy() + np.einsum('trw->tw',RECC_System.FlowDict['F_9_10_Nl'].Values[:,:,:,0]).copy() + np.einsum('trw->tw',RECC_System.FlowDict['F_9_10_No'].Values[:,:,:,0]).copy()
            PrimaryProduction[:,:,mS,mR]                = RECC_System.FlowDict['F_3_4'].Values[:,:,0].copy()
            SecondaryProduct[:,:,mS,mR]                 = RECC_System.FlowDict['F_9_12'].Values[:,0,:,0].copy()
            SecondaryExport[:,:,mS,mR]                  = RECC_System.FlowDict['F_12_21'].Values[:,0,:,0].copy() 
            RenovationMaterialInflow_7[:,:,mS,mR]       = np.einsum('tcrgm->tm',F_6_7_ren[:,:,:,:,:,0]).copy()
            FabricationScrap[:,:,mS,mR]                 = RECC_System.FlowDict['F_5_10'].Values[:,0,:,0].copy()
            ReUse_Materials[:,:,mS,mR]                  = np.einsum('tcrgm->tm',RECC_System.FlowDict['F_8_17_Nr'].Values[:,:,:,:,:,0]) + np.einsum('tclLm->tm',RECC_System.FlowDict['F_8_17_Nl'].Values[:,:,:,:,:,0]) + np.einsum('tcoOm->tm',RECC_System.FlowDict['F_8_17_No'].Values[:,:,:,:,:,0])
        #    Carbon_Wood_Inflow[:,mS,mR]                 = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44 * (np.einsum('trg->t', RECC_System.FlowDict['F_6_7_Nr'].Values[:,:,:,Wood_loc,0]).copy() + np.einsum('tlL->t', RECC_System.FlowDict['F_6_7_Nl'].Values[:,:,:,Wood_loc,0]).copy() + np.einsum('toO->t', RECC_System.FlowDict['F_6_7_No'].Values[:,:,:,Wood_loc,0]).copy())
         #   Carbon_Wood_Outflow[:,mS,mR]                = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44 * (np.einsum('tcrg->t',RECC_System.FlowDict['F_7_8'].Values[:,:,:,:,Wood_loc,0]).copy() + np.einsum('tclL->t',RECC_System.FlowDict['F_7_8_Nl'].Values[:,:,:,:,Wood_loc,0]).copy() + np.einsum('tcoO->t',RECC_System.FlowDict['F_7_8_No'].Values[:,:,:,:,Wood_loc,0]).copy())
          #  Carbon_Wood_Stock[:,mS,mR]                  = RECC_System.ParameterDict['3_MC_CO2FromWoodCombustion'].Values[0,Wood_loc] * 12/44 * (np.einsum('tcrg->t',RECC_System.StockDict['S_7'].Values[:,:,:,:,Wood_loc,0]).copy() + np.einsum('tclL->t',RECC_System.StockDict['S_7_Nl'].Values[:,:,:,:,Wood_loc,0]).copy() + np.einsum('tcoO->t',RECC_System.StockDict['S_7_No'].Values[:,:,:,:,Wood_loc,0]).copy())
            SecondaryProducts_tot[:,:,mS,mR]                  =  RECC_System.FlowDict['F_9_12'].Values[:,0,:,0].copy() + ReUse_Materials[:,:,mS,mR] 
           

           

           
     
            # if 'C ' in Element_List:  
           
            #      # Energy flows
            #     EnergyCons_UP_Vh[:,mS,mR]                   = np.einsum('trpn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav).copy()
            #     EnergyCons_UP_Bd[:,mS,mR]                   = np.einsum('trBn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb).copy() + np.einsum('trNn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb).copy() + np.einsum('toNn->t',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg).copy()
            #     EnergyCons_Mn[:,mS,mR]                      = SysVar_EnergyDemand_Manufacturing.sum(axis =1).copy()
            #     EnergyCons_Wm[:,mS,mR]                      = SysVar_EnergyDemand_WasteMgt.sum(axis =1).copy() +  SysVar_EnergyDemand_Remelting.sum(axis =1).copy()
            #     EnergyCons_UP_Service[:,:,Service_Drivg,mS,mR] = SysVar_EnergyDemand_UsePhase_ByService_pav[:,:,Service_Drivg].copy()
            #     EnergyCons_UP_Service[:,:,Heating_loc,mS,mR]= SysVar_EnergyDemand_UsePhase_ByService_reb[:,:,Heating_loc].copy()
            #     EnergyCons_UP_Service[:,:,Cooling_loc,mS,mR]= SysVar_EnergyDemand_UsePhase_ByService_reb[:,:,Cooling_loc].copy()
            #     EnergyCons_UP_Service[:,:,DomstHW_loc,mS,mR]= SysVar_EnergyDemand_UsePhase_ByService_reb[:,:,DomstHW_loc].copy()
            #     EnergyCons_UP_total[:,:,mS,mR]              = SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_all.copy()
            #     EnergyCons_total[:,:,mS,mR]                 = SysVar_TotalEnergyDemand.copy()
            # # Service flows
            Passenger_km[:,mS,mR]                       = np.einsum('rt,tr->t',RECC_System.ParameterDict['1_F_Function_Future'].Values[Sector_pav_loc,:,:,mS],RECC_System.ParameterDict['2_P_RECC_Population_SSP_32R'].Values[0,:,:,mS]).copy()
            # Hop over to save memory:
            # Vehicle_km[:,mS,mR]                         = np.einsum('tcpr->t',SysVar_StockServiceProvision_UsePhase_pav[:,:,:,:,Service_Driving])
            Vehicle_km[:,mS,mR]                         = np.einsum('rt,tcpr->t', RECC_System.ParameterDict['3_IO_Vehicles_UsePhase_eff' ].Values[Service_Drivg,:,:,mS],   Stock_Detail_UsePhase_p)
            Service_IO_ResBuildings[:,:,mS,mR]          = SysVar_StockServiceProvision_UsePhase_reb_agg.copy()
      #      Service_IO_NonResBuildings[:,:,mS,mR]       = SysVar_StockServiceProvision_UsePhase_nrb_agg.copy()
            # Parameters        
     
            if 'C ' in Element_List:  
                Vehicle_FuelEff[:,:,:,mS,mR]                = np.einsum('tpnr->tpr',RECC_System.ParameterDict['3_EI_Products_UsePhase_passvehicles'].Values[SwitchTime-1::,:,Service_Drivg,:,:,mS])
                ResBuildng_EnergyCons[:,:,:,mS,mR]          = np.einsum('VtBnr->tBr',RECC_System.ParameterDict['3_EI_Products_UsePhase_resbuildings'].Values[SwitchTime-1::,:,Service_Reb,:,:,mS])
           #     GWP_bio_Credit[:,mS,mR]                     = SysVar_GHGEms_GWP_bio[0,:].copy()
            # Product flows
            EoL_Products_for_WasteMgt[:,:,mS,mR]        = np.einsum('trgm->tg', RECC_System.FlowDict['F_8_9'].Values[:,:,:,:,0]).copy()
            Outflow_Products_Usephase_all[:,:,mS,mR]    = np.einsum('tcrgm->tg',RECC_System.FlowDict['F_7_8'].Values[:,:,:,:,:,0]).copy()
            if 'ind' in SectorList:
                EoL_Products_for_WasteMgt[:,Sector_11reg_rge,mS,mR]        = np.einsum('tlLm->tL', RECC_System.FlowDict['F_8_9_Nl'].Values[:,:,:,:,0]).copy()
                Outflow_Products_Usephase_all[:,Sector_11reg_rge,mS,mR]    = np.einsum('tclLm->tL',RECC_System.FlowDict['F_7_8_Nl'].Values[:,:,:,:,:,0]).copy()   
            if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
                EoL_Products_for_WasteMgt[:,Sector_1reg_rge,mS,mR]         = np.einsum('toOm->tO', RECC_System.FlowDict['F_8_9_No'].Values[:,:,:,:,0]).copy()
                Outflow_Products_Usephase_all[:,Sector_1reg_rge,mS,mR]     = np.einsum('tcoOm->tO',RECC_System.FlowDict['F_7_8_No'].Values[:,:,:,:,:,0]).copy()               
            Outflow_Materials_Usephase_all[:,:,mS,mR]   = np.einsum('tcrgm->tm',RECC_System.FlowDict['F_7_8'].Values[:,:,:,:,:,0]).copy() + np.einsum('tclLm->tm',RECC_System.FlowDict['F_7_8_Nl'].Values[:,:,:,:,:,0]).copy() + np.einsum('tcoOm->tm',RECC_System.FlowDict['F_7_8_No'].Values[:,:,:,:,:,0]).copy()
  #          WasteMgtLosses_To_Landfill[:,:,mS,mR]       = RECC_System.FlowDict['F_9_13'].Values.copy()
            StockCurves_Mat[:,:,mS,mR]                  = np.einsum('tcrgm->tm',RECC_System.StockDict['S_7'].Values[:,:,:,:,:,0]).copy() + np.einsum('tclLm->tm',RECC_System.StockDict['S_7_Nl'].Values[:,:,:,:,:,0]).copy() + np.einsum('tcoOm->tm',RECC_System.StockDict['S_7_No'].Values[:,:,:,:,:,0]).copy()
            
    #        # Diagnostics:
    #        Aa = np.einsum('trpn->trp',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav)   # Total use phase energy demand, pav
    #        Aa = np.einsum('trBn->trB',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb)   # Total use phase energy demand, reb
    #        Aa = np.einsum('trNn->trN',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb)   # Total use phase energy demand, nrb
    #        Aa = np.einsum('toNn->toN',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrbg)  # Total use phase energy demand, nrb
    #        Aa = np.einsum('tme->tme',Element_Material_Composition[:,:,:,mS,mR])           # Element composition over years
    #        Aa = np.einsum('tme->tme',Element_Material_Composition_raw[:,:,:,mS,mR])       # Element composition over years, with zero entries
    #        Aa = np.einsum('tgm->tgm',Manufacturing_Output[:,:,:,mS,mR])                   # Manufacturing_output_by_material
            
    # #        # Extract calibration for SSP1:
    #         if mS == 1:
    #             E_Calib_Vehicles  = np.einsum('trgn->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_pav[:,:,:,0:7])
    #             E_Calib_Buildings = np.einsum('trgn->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_reb[:,:,:,0:7])
    # #             E_Calib_NRBuildgs = np.einsum('trgn->tr',SysVar_EnergyDemand_UsePhase_ByEnergyCarrier_nrb[:,:,:,0:7])
            
            # Determine exit flags            
            ExitFlags['Positive_Inflow_F6_7_R32_SSP_'  + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_6_7_Nr'].Values.min(),0, IsClose_Remainder_Small)
            ExitFlags['Positive_Outflow_F7_8_R32_SSP_' + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_7_8'].Values.min(),0, IsClose_Remainder_Small)  
            ExitFlags['Positive_Inflow_F8_9_R32_SSP_'  + str(mS) + '_RCP_' + str(mR)] = np.isclose(RECC_System.FlowDict['F_8_9'].Values.min(),0, IsClose_Remainder_Small)
            
            # del RECC_System # Delete system when done, clear memory.
            '''                
            Emissions scopes reported:
            << Emissions account item >>                     << ODYM-RECC variable >>       // << Result file label (text) >>
            System, all processes:                              GWP_System_3579di           // GHG emissions, system-wide _3579di   
            is composed of:
            (i) Operation
            Use phase only:                                     GWP_UsePhase_7d             // GHG emissions, use phase _7d
            Use phase, electricity scope2:                      GWP_UsePhase_7i_Scope2_El   // GHG emissions, use phase scope 2 (electricity) _7i
            Use phase, indirect, rest:                          GWP_UsePhase_7i_OtherIndir  // GHG emissions, use phase other indirect (non-el.) _7i
            (ii) Material production and manufacturing
            Material production:                                GWP_Materials_3di_9di       // GHG emissions, material cycle industries and their energy supply _3di_9di
            Manufacturing:                                      GWP_Manufact_5di_all        // GHG emissions, manufacturing _5i, all
            (iii) Energy recovery, forest carbon uptake
            Energy recovery wood waste                          GWP_EnergyRecoveryWasteWood // GHG emissions, energy recovery from waste wood (biogenic C plus energy substitution within System)
            Forest CO2 uptake                                   GWP_ForestCO2Uptake         // GHG sequestration by forests (w. neg. sign)
            (iv) Reported extra, not part of System emissions
            RecyclingCredit                                     GWP_RecyclingCredit         // GHG emissions, energy recovery from waste wood (biogenic C plus energy substitution within System)
            '''        
            
    #            # DIAGNOSTICS
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_5_6',0)
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_12_5',0)
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_5_10',0)
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_4_5',0)
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_17_6',0)
    #            
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_12_5',0)
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_12_21',0)
    #            a,b,c = RECC_System.Check_If_All_Chem_Elements_Are_present('F_9_12',0)
            
    ##############################################################
    #   Section 7) Export and plot results, save, and close      #
    ##############################################################
    Mylog.info('## 5 - Evaluate results, save, and close')
    
    
    Mylog.info('### 5.0 - Check data and results for boundary constraints and plausibility. Exit flags.')
    Mylog.info('Model input')          
    
    ExitFlags['3_SHA_LightWeighting_Vehicles_min']             = ParameterDict['3_SHA_LightWeighting_Vehicles'].Values.min() >= 0
    ExitFlags['3_SHA_LightWeighting_Vehicles_max']             = ParameterDict['3_SHA_LightWeighting_Vehicles'].Values.max()/100 <= 1 # unit is % not 1!
    ExitFlags['3_SHA_DownSizing_Vehicles_min']                 = ParameterDict['3_SHA_DownSizing_Vehicles'].Values.min() >= 0
    ExitFlags['3_SHA_DownSizing_Vehicles_max']                 = ParameterDict['3_SHA_DownSizing_Vehicles'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_Vehicles_min']                  = ParameterDict['3_SHA_TypeSplit_Vehicles'].Values.min() >= 0
    ExitFlags['3_SHA_TypeSplit_Vehicles_max']                  = ParameterDict['3_SHA_TypeSplit_Vehicles'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_Vehicles_sum']                  = np.isclose(ParameterDict['3_SHA_TypeSplit_Vehicles'].Values.sum(),Nr*NR*Nt, IsClose_Remainder_Large)
    ExitFlags['3_SHA_TypeSplit_Buildings_min']                 = ParameterDict['3_SHA_TypeSplit_Buildings'].Values.min() >= 0
    ExitFlags['3_SHA_TypeSplit_Buildings_max']                 = ParameterDict['3_SHA_TypeSplit_Buildings'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_Buildings_sum']                 = np.isclose(ParameterDict['3_SHA_TypeSplit_Buildings'].Values.sum(),Nr*Nt*NS, IsClose_Remainder_Large)
    ExitFlags['3_SHA_TypeSplit_NonResBuildings_min']           = ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values.min() >= 0
    ExitFlags['3_SHA_TypeSplit_NonResBuildings_max']           = ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values.max() <= 1
    ExitFlags['3_SHA_TypeSplit_NonResBuildings_sum']           = np.isclose(ParameterDict['3_SHA_TypeSplit_NonResBuildings'].Values.sum(),Nr*Nt*NS, IsClose_Remainder_Large)
    ExitFlags['LTE_Renovation_Consistency']                    = bool(ScriptConfig['Include_REStrategy_LifeTimeExtension']) & bool(ScriptConfig['Include_Renovation_reb']) & bool(ScriptConfig['Include_Renovation_nrb'])
    ExitFlags['Secondary_Material_Flows_Positive']             = SecondaryProduct.min() >= 0
    
    
    Mylog.info('Model exit flags:')
    for key in ExitFlags:
        Mylog.info(key + ': ' + str(ExitFlags[key]))
        
        
    
    Mylog.info('Model output')
               
    Mylog.info('### 5.1 - Create plots and include in logfiles')
    Mylog.info('Plot and export results')
    
    book = openpyxl.Workbook()
    ws1 = book.active
    ws1.title = 'Cover'
    ws1.cell(row=3, column=2).value = 'ScriptConfig'
    ws1.cell(row=3, column=2).font = openpyxl.styles.Font(bold=True)
    m = 4
    for x in sorted(ScriptConfig.keys()):
        ws1.cell(row=m, column=2).value = x
        ws1.cell(row=m, column=3).value = ScriptConfig[x]
        m +=1
    
    ws2 = book.create_sheet('Model_Results')
    ColLabels = ['Indicator','Unit','Region','System_location','RE scen','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws2.cell(row=1, column=m+1).value = ColLabels[m]
        ws2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws2.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    
    # GHG overview, bulk materials
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_System_3579di,2,len(ColLabels),'GHG emissions, system-wide _3579di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'all processes','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_PrimaryMaterial_3di,newrowoffset,len(ColLabels),'GHG emissions, primary material production _3di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'Process and direct emissions in process 3 and related energy supply','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,8,:,:],newrowoffset,len(ColLabels),'Cement production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,0:4,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Primary steel production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct.sum(axis=1),newrowoffset,len(ColLabels),'Secondary materials, total','Mt / yr',ScriptConfig['RegionalScope'],'F_10_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction.sum(axis=1),newrowoffset,len(ColLabels),'Primary materials, total','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # final material consumption, fab scrap
    for m in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Material_Inflow[:,:,m,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Final consumption of materials: ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[m],'Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,0:4,:,:]),newrowoffset,len(ColLabels),'Final consumption of materials: iron&steel (4 groups)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tgmSR->tSR',Material_Inflow[:,:,4:6,:,:]),newrowoffset,len(ColLabels),'Final consumption of materials: aluminium (2 groups)','Mt / yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    for m in range(0,Nw):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,FabricationScrap[:,m,:,:],newrowoffset,len(ColLabels),'Fabrication scrap: ' + IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items[m],'Mt / yr',ScriptConfig['RegionalScope'],'F_5_10','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # secondary materials
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,0:4,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Secondary steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,4:6,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Secondary Al','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,6,:,:],newrowoffset,len(ColLabels),'Secondary copper','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_UsePhase_7d,newrowoffset,len(ColLabels),'GHG emissions, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_OtherThanUsePhaseDirect,newrowoffset,len(ColLabels),'GHG emissions, other than use phase direct: all industries and energy supply','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'F_9_12','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # GHG emissions, detail
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_Vehicles_Direct.sum(axis =1),newrowoffset,len(ColLabels),'GHG emissions, vehicles, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_7_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_ReBuildgs_Direct.sum(axis =1),newrowoffset,len(ColLabels),'GHG emissions, res. buildings, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_7_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_NRBuildgs_Direct.sum(axis =1),newrowoffset,len(ColLabels),'GHG emissions, non-res. buildings, use phase _7d','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_7_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_Vehicles_indir,newrowoffset,len(ColLabels),'GHG emissions, vehicles, energy supply _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_AllBuildings_indir,newrowoffset,len(ColLabels),'GHG emissions, res+non-res buildings, energy supply _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_Manufact_5di_all,newrowoffset,len(ColLabels),'GHG emissions, manufacturing _5i, all','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_5_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_WasteMgt_9di_all,newrowoffset,len(ColLabels),'GHG emissions, waste mgt. and remelting _9di, all','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_9_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,4:6,:,:].sum(axis=1),newrowoffset,len(ColLabels),'Primary Al production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,6,:,:],newrowoffset,len(ColLabels),'Primary Cu production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_Materials_3di_9di,newrowoffset,len(ColLabels),'GHG emissions, material cycle industries and their energy supply _3di_9di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_3_0 and related energy supply emissions','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # energy flows
    for nn in range(0,Nn):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_total[:,nn,:,:],newrowoffset,len(ColLabels),'energy consumption, system-wide: ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[nn],'Tt / yr',ScriptConfig['RegionalScope'],'F_15_x','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for nn in range(0,Nn):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_total[:,nn,:,:],newrowoffset,len(ColLabels),'energy consumption, use phase: ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[nn],'Tt / yr',ScriptConfig['RegionalScope'],'F_15_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_Vh,newrowoffset,len(ColLabels),'Energy cons., use phase, vehicles','TJ/yr',ScriptConfig['RegionalScope'],'E_16_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_Bd,newrowoffset,len(ColLabels),'Energy cons., use phase, res+non-res buildings','TJ/yr',ScriptConfig['RegionalScope'],'E_16_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_Mn,newrowoffset,len(ColLabels),'Energy cons., manufacturing','TJ/yr',ScriptConfig['RegionalScope'],'E_16_5','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_Wm,newrowoffset,len(ColLabels),'Energy cons., waste mgt. and remelting','TJ/yr',ScriptConfig['RegionalScope'],'E_16_9','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # stocks
    if 'pav' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Totl[:,Sector_pav_loc,:,:],newrowoffset,len(ColLabels),'In-use stock, pass. vehicles','million units',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Totl[:,Sector_reb_loc,:,:],newrowoffset,len(ColLabels),'In-use stock, res. buildings','million m2',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Totl[:,Sector_nrb_loc,:,:],newrowoffset,len(ColLabels),'In-use stock, nonres. buildings','million m2',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Prod[:,mg,:,:],newrowoffset,len(ColLabels),'In-use stock, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million, Buildings: million m2',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Mat[:,mm,:,:],newrowoffset,len(ColLabels),'In-use stock, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,StockCurves_Mat.sum(axis=1),newrowoffset,len(ColLabels),'In-use stock, all materials','Mt',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    #per capita stocks per sector
    for mr in range(0,Nr):
        for mG in range(0,NG):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,pCStocksCurves[:,mG,mr,:,:],newrowoffset,len(ColLabels),'per capita in-use stock, ' + IndexTable.Classification[IndexTable.index.get_loc('Sectors')].Items[mG],'vehicles: cars per person, buildings: m2 per person',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'S_7 (part, per capita)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    #population
    for mr in range(0,Nr):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Population[:,mr,:,:],newrowoffset,len(ColLabels),'Population','million',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'P (population)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    #UsingLessMaterialByDesign and Mat subst. shares
    # a) pass. vehicles:
    if 'pav' in SectorList:
        for mr in range(0,Nr):
            for ms in range(0,Ns): # vehicle downsizing parameter is modified by script, result exported here.
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tS,R->tSR',ParameterDict['3_SHA_DownSizing_Vehicles'].Values[ms,mr,:,:],np.ones((NR))),newrowoffset,len(ColLabels),'Segment split of newly registered pass. vehicles, ' +IndexTable.Classification[IndexTable.index.get_loc('Car_segments')].Items[ms],'1',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
            for mp in range(0,len(Sector_pav_rge)): # vehicle lightweighting parameter is modified by script, result exported here.
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tS,R->tSR',ParameterDict['3_SHA_LightWeighting_Vehicles'].Values[mp,mr,:,:],np.ones((NR))),newrowoffset,len(ColLabels), 'Share of light-weighted cars in newly registered ' +IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_pav_rge[mp]],'%',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)      
            for mp in range(0,len(Sector_pav_rge)): # export vehicle type split
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('Rt,S->tSR',ParameterDict['3_SHA_TypeSplit_Vehicles'].Values[Sector_pav_loc,mr,:,mp,:],np.ones((NS))),newrowoffset,len(ColLabels), 'Type split of newly registered cars, ' +IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_pav_rge[mp]],'%',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)                      
    # b) res. buildings   
    if 'reb' in SectorList:     
        for mr in range(0,Nr): # building downsizing parameter is modified by script, result exported here.
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tS,R->tSR',ParameterDict['3_SHA_DownSizing_Buildings'].Values[0,mr,:,:],np.ones((NR))),newrowoffset,len(ColLabels),'Share of newly built downsized res. buildings','%',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
            for mB in range(0,len(Sector_reb_rge)): # building lightweighting parameter is modified by script, result exported here.
                newrowoffset = msf.xlsxExportAdd_tAB(ws2,np.einsum('tS,R->tSR',ParameterDict['3_SHA_LightWeighting_Buildings'].Values[mB,mr,:,:],np.ones((NR))),newrowoffset,len(ColLabels),'Share of newly built light-weighted ' +IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_reb_rge[mB]],'%',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    #passenger and vehicle km, building service
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Passenger_km,newrowoffset,len(ColLabels),'passenger-km supplied by pass. vehicles','million km/yr',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Vehicle_km,newrowoffset,len(ColLabels),'vehicle-km driven by pass. vehicles','million km/yr',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_ResBuildings[:,Heating_loc,:,:],newrowoffset,len(ColLabels),'Total heated floor space, res. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_ResBuildings[:,Cooling_loc,:,:],newrowoffset,len(ColLabels),'Total cooled floor space, res. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_NonResBuildings[:,Heating_loc,:,:],newrowoffset,len(ColLabels),'Total heated floor space, nonres. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Service_IO_NonResBuildings[:,Cooling_loc,:,:],newrowoffset,len(ColLabels),'Total cooled floor space, nonres. buildings','million m2',ScriptConfig['RegionalScope'],'P7 (use phase)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Use phase indirect GHG, primary prodution GHG, material cycle and recycling credit
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_UsePhase_7i_Scope2_El,newrowoffset,len(ColLabels),'GHG emissions, use phase scope 2 (electricity) _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (electricity, for use phase energy)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_UsePhase_7i_OtherIndir,newrowoffset,len(ColLabels),'GHG emissions, use phase other indirect (non-el.) _7i','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_15_0 (other than el., for use phase energy)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_PrimaryMaterial_3di,newrowoffset,len(ColLabels),'GHG emissions, primary material production (redundant) _3di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_3_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_MaterialCycle_5di_9di,newrowoffset,len(ColLabels),'GHG emissions, manufact, wast mgt., remelting and indirect _5di_9di','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'E_9_0 + E_15_0 (part, for energy supply waste mgt.)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_RecyclingCredit,newrowoffset,len(ColLabels),'GHG emissions, recycling credits','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'outside system','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_ForestCO2Uptake,newrowoffset,len(ColLabels),'GHG sequestration by forests (w. neg. sign)','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'Process 1','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_EnergyRecoveryWasteWood,newrowoffset,len(ColLabels),'GHG emissions, energy recovery from waste wood (biogenic C plus energy substitution within System)','Mt of CO2-eq / yr',ScriptConfig['RegionalScope'],'waste mgt. and energy supply','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Primary and secondary material production, if not included above already
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,0,:,:], newrowoffset,len(ColLabels),'Primary construction grade steel production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,1,:,:], newrowoffset,len(ColLabels),'Primary automotive steel production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,2,:,:], newrowoffset,len(ColLabels),'Primary stainless production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,3,:,:], newrowoffset,len(ColLabels),'Primary cast iron production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,4,:,:], newrowoffset,len(ColLabels),'Primary wrought Al production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,5,:,:], newrowoffset,len(ColLabels),'Primary cast Al production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,7,:,:], newrowoffset,len(ColLabels),'Primary plastics production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,9,:,:], newrowoffset,len(ColLabels),'Wood, from forests','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,PrimaryProduction[:,10,:,:],newrowoffset,len(ColLabels),'Primary zinc production','Mt / yr',ScriptConfig['RegionalScope'],'F_3_4 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,0,:,:],  newrowoffset,len(ColLabels),'Secondary construction steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,1,:,:],  newrowoffset,len(ColLabels),'Secondary automotive steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,2,:,:],  newrowoffset,len(ColLabels),'Secondary stainless steel','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,3,:,:],  newrowoffset,len(ColLabels),'Secondary cast iron','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,4,:,:],  newrowoffset,len(ColLabels),'Secondary wrought Al','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,5,:,:],  newrowoffset,len(ColLabels),'Secondary cast Al','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,7,:,:],  newrowoffset,len(ColLabels),'Secondary plastics','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,9,:,:],  newrowoffset,len(ColLabels),'Recycled wood','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,10,:,:], newrowoffset,len(ColLabels),'Recycled zinc','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryProduct[:,11,:,:], newrowoffset,len(ColLabels),'Recycled concrete','Mt / yr',ScriptConfig['RegionalScope'],'F_9_12 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # GHG of primary and secondary material production
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_PrimaryMaterial_3di_m[:,mm,:,:],newrowoffset,len(ColLabels),'GHG emissions, production of primary _3di_' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'E_3_0 (part) and associated em. in E_15_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_SecondaryMetal_di_m[:,mm,:,:],newrowoffset,len(ColLabels),'GHG emissions, production of secondary _di_' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'E_9_0 (part) and associated em. in E_15_0','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # inflow and outflow of commodities
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,mg,:,:],newrowoffset,len(ColLabels),'final consumption (use phase inflow), ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'pav' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,Sector_pav_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all drive technologies together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,Sector_reb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all res. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Inflow_Prod[:,Sector_nrb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'final consumption (use phase inflow), all nonres. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_6_7','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,mg,:,:],newrowoffset,len(ColLabels),'EoL products (use phase outflow), ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'pav' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_pav_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'EoL products (use phase outflow), all drive technologies together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    if 'reb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_reb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'EoL products (use phase outflow), all res. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    if 'nrb' in SectorList:
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Prod[:,Sector_nrb_rge,:,:].sum(axis=1),newrowoffset,len(ColLabels),'EoL products (use phase outflow), all nonres. building types together','Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    # Material reuse
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,ReUse_Materials[:,mm,:,:],newrowoffset,len(ColLabels),'ReUse of materials in products, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'F_17_6','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # carbon in wood inflow, stock, and outflow
    #newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Inflow, newrowoffset,len(ColLabels),'Carbon in wood and wood products, final consumption/inflow','Mt/yr',ScriptConfig['RegionalScope'],'F_6_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    # newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Outflow,newrowoffset,len(ColLabels),'Carbon in wood and wood products, EoL flows, outflow use phase','Mt/yr',ScriptConfig['RegionalScope'],'F_7_8 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    #newrowoffset = msf.xlsxExportAdd_tAB(ws2,Carbon_Wood_Stock,  newrowoffset,len(ColLabels),'Carbon in wood and wood products, in-use stock','Mt',ScriptConfig['RegionalScope'],'S_7 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)    
    # specific energy consumption of vehicles and buildings
    for mr in range(0,Nr):
        for mp in range(0,len(Sector_pav_rge)):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,Vehicle_FuelEff[:,mp,mr,:,:],newrowoffset,len(ColLabels),'specific energy consumption, driving, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_pav_rge[mp]],'MJ/km',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mr in range(0,Nr):
        for mB in range(0,len(Sector_reb_rge)):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,ResBuildng_EnergyCons[:,mB,mr,:,:],newrowoffset,len(ColLabels),'specific energy consumption, heating/cooling/DHW, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[Sector_reb_rge[mB]],'MJ/m2',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # specific energy consumption of vehicles and residential buildings
    for mr in range(0,Nr):
        for mV in range(0,NV):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,EnergyCons_UP_Service[:,mr,mV,:,:],newrowoffset,len(ColLabels),'Total use phase energy consumption, ' + IndexTable.Classification[IndexTable.index.get_loc('ServiceType')].Items[mV],'TJ/yr',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # GWP by energy carrier, vehicles and residential buildings
    for mr in range(0,Nr):
        for mn in range(0,Nn):
            newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_ByEnergyCarrier_UsePhase_d[:,mr,mn,:,:] + GWP_ByEnergyCarrier_UsePhase_i[:,mr,mn,:,:],newrowoffset,len(ColLabels),'GWP by energy carrier, use phase direct + indirect, all sectors covered by model run, ' + IndexTable.Classification[IndexTable.index.get_loc('Energy')].Items[mn],'Mt/yr',IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items[mr],'use phase and scope 2','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # carbon credit from longterm biomass storage
    # newrowoffset = msf.xlsxExportAdd_tAB(ws2,GWP_bio_Credit,newrowoffset,len(ColLabels),'GWP_bio_usephase','Mt / yr',ScriptConfig['RegionalScope'],'use phase','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)        
    # Excess secondary material export    
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,0,:,:], newrowoffset,len(ColLabels),'Export of Secondary construction steel','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,1,:,:], newrowoffset,len(ColLabels),'Export of Secondary automotive steel','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,2,:,:], newrowoffset,len(ColLabels),'Export of Secondary stainless steel','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,3,:,:], newrowoffset,len(ColLabels),'Export of Secondary cast iron','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,4,:,:], newrowoffset,len(ColLabels),'Export of Secondary wrought Al','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,5,:,:], newrowoffset,len(ColLabels),'Export of Secondary cast Al','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,6,:,:], newrowoffset,len(ColLabels),'Export of Secondary copper','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,7,:,:], newrowoffset,len(ColLabels),'Export of Secondary plastics','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,8,:,:], newrowoffset,len(ColLabels),'Export of Secondary cement','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,9,:,:], newrowoffset,len(ColLabels),'Export of Recycled wood','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,10,:,:],newrowoffset,len(ColLabels),'Export of Recycled zinc','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,11,:,:],newrowoffset,len(ColLabels),'Export of Recycled concrete','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,SecondaryExport[:,12,:,:],newrowoffset,len(ColLabels),'Export of Secondary concrete aggregates','Mt / yr',ScriptConfig['RegionalScope'],'F_12_21 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # manufacturing output by materials
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,0,:,:].sum(axis=1), newrowoffset,len(ColLabels),'construction steel in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,1,:,:].sum(axis=1), newrowoffset,len(ColLabels),'automotive steel in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,2,:,:].sum(axis=1), newrowoffset,len(ColLabels),'stainless steel in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,3,:,:].sum(axis=1), newrowoffset,len(ColLabels),'cast iron in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,4,:,:].sum(axis=1), newrowoffset,len(ColLabels),'wrought Al in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,5,:,:].sum(axis=1), newrowoffset,len(ColLabels),'cast Al in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,6,:,:].sum(axis=1), newrowoffset,len(ColLabels),'copper in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,7,:,:].sum(axis=1), newrowoffset,len(ColLabels),'plastics in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,8,:,:].sum(axis=1), newrowoffset,len(ColLabels),'cement in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,9,:,:].sum(axis=1), newrowoffset,len(ColLabels),'wood in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,10,:,:].sum(axis=1),newrowoffset,len(ColLabels),'zinc in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,11,:,:].sum(axis=1),newrowoffset,len(ColLabels),'concrete in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    newrowoffset = msf.xlsxExportAdd_tAB(ws2,Manufacturing_Output[:,:,12,:,:].sum(axis=1),newrowoffset,len(ColLabels),'concrete aggregates in manufactured goods','Mt / yr',ScriptConfig['RegionalScope'],'F_5_6 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # postconsumer scrap
    for m in range(0,Nw):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Scrap_Outflow[:,m,:,:],newrowoffset,len(ColLabels),'Postconsumer scrap: ' + IndexTable.Classification[IndexTable.index.get_loc('Waste_Scrap')].Items[m],'Mt / yr',ScriptConfig['RegionalScope'],'F_9_10 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # EoL Products to waste mgt.
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,EoL_Products_for_WasteMgt[:,mg,:,:],newrowoffset,len(ColLabels),'EoL Products to waste mgt., ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_8_9 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Outflow of products from use phase        
    for mg in range(0,Ng):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Products_Usephase_all[:,mg,:,:],newrowoffset,len(ColLabels),'Outflow of products from use phase, ' + IndexTable.Classification[IndexTable.index.get_loc('Good')].Items[mg],'Vehicles: million/yr, Buildings: million m2/yr',ScriptConfig['RegionalScope'],'F_7_8 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,Outflow_Materials_Usephase_all[:,mm,:,:],newrowoffset,len(ColLabels),'Outflow of materials from use phase, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'F_7_8 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # Losses from waste mgt.        
    # for me in range(0,Ne):    
    #     newrowoffset = msf.xlsxExportAdd_tAB(ws2,WasteMgtLosses_To_Landfill[:,me,:,:],newrowoffset,len(ColLabels),'Waste mgt and remelting losses, ' + IndexTable.Classification[IndexTable.index.get_loc('Element')].Items[me],'Mt/yr',ScriptConfig['RegionalScope'],'F_9_13 (part)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    # # Renovation material inflow:
    for mm in range(0,Nm):
        newrowoffset = msf.xlsxExportAdd_tAB(ws2,RenovationMaterialInflow_7[:,mm,:,:],newrowoffset,len(ColLabels),'Inflow of renovation material into use phase, ' + IndexTable.Classification[IndexTable.index.get_loc('Engineering materials')].Items[mm],'Mt/yr',ScriptConfig['RegionalScope'],'F_6_7 (part: renovation inflow)','Cf. Cover sheet',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        
        
        
        
    #### save copper results in seperate work sheet
        
    ws_Cu = book.create_sheet('Model_Results_Cu_Losses_Stocks')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
     
    
    
    # save losses in results sheet
    if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
        for mm in range(0,NO):
            newrowoffset_Cu_Obs = msf.xlsxExportAdd_tAB(ws_Cu,ObsoleteStocks[:,mm,:,:],2+mm*6,len(ColLabels),'Obsolete Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
          
        for mh in range(0,NO):
            newrowoffset_Cu_Hyb = msf.xlsxExportAdd_tAB(ws_Cu,HybernatingStocks[:,mh,:,:],mh*6+newrowoffset_Cu_Obs,len(ColLabels),'Hybernating Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mh],'S_19',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for ms in range(0,NO):
            newrowoffset_Cu_Sort = msf.xlsxExportAdd_tAB(ws_Cu,SortingLosses[:,ms,:,:],ms*6+newrowoffset_Cu_Hyb,len(ColLabels),'Sorting Losses','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[ms],'S_13',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
         
        for mr in range(0,NO):
            newrowoffset_Cu_Rec = msf.xlsxExportAdd_tAB(ws_Cu,RecoveryLosses[:,:,:],mr*6+newrowoffset_Cu_Sort,len(ColLabels),'Recovery Losses','Mt','Recovery losses','S_23','All Products',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
              
    
        
    if 'ind' in SectorList:
        for mm in range(0,NI):
            newrowoffset_Cu_Obs = msf.xlsxExportAdd_tAB(ws_Cu,ObsoleteStocks[:,mm,:,:],2+mm*6,len(ColLabels),'Obsolete Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
          
        for mh in range(0,NI):
            newrowoffset_Cu_Hyb = msf.xlsxExportAdd_tAB(ws_Cu,HybernatingStocks[:,mh,:,:],mh*6+newrowoffset_Cu_Obs,len(ColLabels),'Hybernating Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mh],'S_19',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for ms in range(0,NI):
            newrowoffset_Cu_Sort = msf.xlsxExportAdd_tAB(ws_Cu,SortingLosses[:,ms,:,:],ms*6+newrowoffset_Cu_Hyb,len(ColLabels),'Sorting Losses','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[ms],'S_13',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
        
        for mr in range(0,NI):
            newrowoffset_Cu_Rec = msf.xlsxExportAdd_tAB(ws_Cu,RecoveryLosses[:,:,:],mr*6+newrowoffset_Cu_Sort,len(ColLabels),'Recovery Losses','Mt',ScriptConfig['RegionalScope'],'Recovery losses stocks','S_23',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
              
    
    if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
        for mm in range(0,NT):
            newrowoffset_Cu_Obs = msf.xlsxExportAdd_tAB(ws_Cu,ObsoleteStocks[:,mm,:,:],2+mm*6,len(ColLabels),'Obsolete Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
          
        for mh in range(0,NT):
            newrowoffset_Cu_Hyb = msf.xlsxExportAdd_tAB(ws_Cu,HybernatingStocks[:,mh,:,:],mh*6+newrowoffset_Cu_Obs,len(ColLabels),'Hybernating Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mh],'S_19',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for ms in range(0,NT):
            newrowoffset_Cu_Sort = msf.xlsxExportAdd_tAB(ws_Cu,SortingLosses[:,ms,:,:],ms*6+newrowoffset_Cu_Hyb,len(ColLabels),'Sorting Losses','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[ms],'S_13',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for mr in range(0,NT):
            newrowoffset_Cu_Rec = msf.xlsxExportAdd_tAB(ws_Cu,RecoveryLosses[:,:,:],mr*6+newrowoffset_Cu_Sort,len(ColLabels),'Recovery Losses','Mt',ScriptConfig['RegionalScope'],'Recovery losses','S_23',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
              
    
        
    
    ws_Cu_2 = book.create_sheet('Model_Results_Services')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_2.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_2.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_2.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    
    if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
        for mo in range(0,NO):
            newrowoffset_Cu_Stocks_tot = msf.xlsxExportAdd_tAB(ws_Cu_2,StockCurves_Prod[:,mo,:,:],2+mo*6,len(ColLabels),'Total stock services','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mo],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    if 'ind' in SectorList:
        for mI in range(0,NI):
            newrowoffset_Cu_Stocks_tot = msf.xlsxExportAdd_tAB(ws_Cu_2,StockCurves_Prod[:,mI,:,:],2+mI*6,len(ColLabels),'Total stock services','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items[mI],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
        for mI in range(0,NT):
            newrowoffset_Cu_Stocks_tot = msf.xlsxExportAdd_tAB(ws_Cu_2,StockCurves_Prod[:,mI,:,:],2+mI*6,len(ColLabels),'Total stock services','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mI],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    
        
    
    ws_Cu_3 = book.create_sheet('Model_Results_Material_Inflow')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_3.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_3.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_3.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_3.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
        
        
    
    
    if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
        for mo in range(0,NO):
            newrowoffset_Cu_Stocks_tot = msf.xlsxExportAdd_tAB(ws_Cu_3,Material_Inflow[:,mo,6,:,:],2+mo*6,len(ColLabels),'Material Inflow','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mo],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    if 'ind' in SectorList:
        for mI in range(0,NI):
            newrowoffset_Cu_Stocks_tot = msf.xlsxExportAdd_tAB(ws_Cu_3,Material_Inflow[:,mI,6,:,:],2+mI*6,len(ColLabels),'Material Inflow','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items[mI],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
        for mI in range(0,NT):
            newrowoffset_Cu_Stocks_tot = msf.xlsxExportAdd_tAB(ws_Cu_3,Material_Inflow[:,mI,6,:,:],2+mI*6,len(ColLabels),'Material Inflow','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mI],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
      
    
    ws_Cu_4 = book.create_sheet('Secondary Cu')
    ColLabels = ['Indicator','Unit','Region','Figure','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_4.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_4.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_4.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_4.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    
    
    if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
        newrowoffset_Cu_Secondary_tot = msf.xlsxExportAdd_tAB(ws_Cu_4,SecondaryProducts_tot[:,6,:,:],2,len(ColLabels),'Secondary Copper','Mt / yr',ScriptConfig['RegionalScope'],'Figure unclear','F_9_12 (part)',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    
    
    
    if 'ind' in SectorList:
        newrowoffset_Cu_Secondary_tot = msf.xlsxExportAdd_tAB(ws_Cu_4,SecondaryProducts_tot[:,6,:,:],2,len(ColLabels),'Secondary Copper','Mt / yr',ScriptConfig['RegionalScope'],'Figure unclear','F_9_12 (part)',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    
    if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
        newrowoffset_Cu_Secondary_tot = msf.xlsxExportAdd_tAB(ws_Cu_4,SecondaryProducts_tot[:,6,:,:],2,len(ColLabels),'Secondary Copper','Mt / yr',ScriptConfig['RegionalScope'],'Figure unclear','F_9_12 (part)',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    
    
    ws_Cu_5 = book.create_sheet('Model_Results_Cu_Losses_Flows')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_5.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_5.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_5.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_5.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


    # save losses in results sheet
    if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
        for mm in range(0,NO):
            newrowoffset_Cu_Obs = msf.xlsxExportAdd_tAB(ws_Cu_5, ObsoleteFlows[:,mm,:,:],2+mm*6,len(ColLabels),'Obsolete Flows','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
          
        for mh in range(0,NO):
            newrowoffset_Cu_Hyb = msf.xlsxExportAdd_tAB(ws_Cu_5,HybernatingFlows[:,mh,:,:],mh*6+newrowoffset_Cu_Obs,len(ColLabels),'Hybernating Flows','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mh],'S_19',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for ms in range(0,NO):
            newrowoffset_Cu_Sort = msf.xlsxExportAdd_tAB(ws_Cu_5,SortingFlows[:,ms,:,:] ,ms*6+newrowoffset_Cu_Hyb,len(ColLabels),'Sorting Flows','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[ms],'S_13',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for mr in range(0,NO):
            newrowoffset_Cu_Rec = msf.xlsxExportAdd_tAB(ws_Cu_5,RecoveryFlows[:,:,:],mr*6+newrowoffset_Cu_Sort,len(ColLabels),'Recovery Losses','Mt',ScriptConfig['RegionalScope'],'Recovery losses stocks','S_23',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
              
    
        
    if 'ind' in SectorList:
        for mm in range(0,NI):
            newrowoffset_Cu_Obs = msf.xlsxExportAdd_tAB(ws_Cu_5,ObsoleteFlows[:,mm,:,:],2+mm*6,len(ColLabels),'Obsolete Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
          
        for mh in range(0,NI):
            newrowoffset_Cu_Hyb = msf.xlsxExportAdd_tAB(ws_Cu_5,HybernatingFlows[:,mh,:,:],mh*6+newrowoffset_Cu_Obs,len(ColLabels),'Hybernating Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mh],'S_19',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for ms in range(0,NI):
            newrowoffset_Cu_Sort = msf.xlsxExportAdd_tAB(ws_Cu_5,SortingFlows[:,ms,:,:],ms*6+newrowoffset_Cu_Hyb,len(ColLabels),'Sorting Losses','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[ms],'S_13',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
       
        for mr in range(0,NI):
            newrowoffset_Cu_Rec = msf.xlsxExportAdd_tAB(ws_Cu_5,RecoveryFlows[:,:,:],mr*6+newrowoffset_Cu_Sort,len(ColLabels),'Recovery Losses','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mr],'S_23',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
              
    
    if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
        for mm in range(0,NT):
            newrowoffset_Cu_Obs = msf.xlsxExportAdd_tAB(ws_Cu_5,ObsoleteFlows[:,mm,:,:],2+mm*6,len(ColLabels),'Obsolete Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
          
        for mh in range(0,NT):
            newrowoffset_Cu_Hyb = msf.xlsxExportAdd_tAB(ws_Cu_5,HybernatingFlows[:,mh,:,:],mh*6+newrowoffset_Cu_Obs,len(ColLabels),'Hybernating Stocks','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mh],'S_19',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
            
        for ms in range(0,NT):
            newrowoffset_Cu_Sort = msf.xlsxExportAdd_tAB(ws_Cu_5,SortingFlows[:,ms,:,:],ms*6+newrowoffset_Cu_Hyb,len(ColLabels),'Sorting Losses','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[ms],'S_13',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
           
        for mr in range(0,NO):
            newrowoffset_Cu_Rec = msf.xlsxExportAdd_tAB(ws_Cu_5,RecoveryFlows[:,:,:],mr*6+newrowoffset_Cu_Sort,len(ColLabels),'Recovery Losses','Mt',ScriptConfig['RegionalScope'],'Recovery losses stocks','S_23',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)

   

    
   
    
    
    ws_Cu_6 = book.create_sheet('Lifetime')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_6.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_6.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_6.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_6.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


    if 'nrbg' in SectorList:
        Par_RECC_ProductLifetime_nrbg_safe = np.einsum('prc,SR->prcSR', Par_RECC_ProductLifetime_nrbg,np.ones((NS,NR)))        
        for mm in range(0,NN):
                newrowoffset_Cu_lifetime_nrbg = msf.xlsxExportAdd_tAB(ws_Cu_6, Par_RECC_ProductLifetime_nrbg_safe[mm,0,SwitchTime-1:,:,:],2+mm*6,len(ColLabels),'Average Lifetime','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
       
    else:
        newrowoffset_Cu_lifetime_nrbg = 2

    if 'app' in SectorList:
        
        
        Par_RECC_ProductLifetime_app_safe = np.einsum('prc,SR->prcSR', Par_RECC_ProductLifetime_app,np.ones((NS,NR)))
        for mm in range(0,Na):
                 newrowoffset_Cu_lifetime_app = msf.xlsxExportAdd_tAB(ws_Cu_6, Par_RECC_ProductLifetime_app_safe[mm,0,SwitchTime-1:,:,:],mm*6+newrowoffset_Cu_lifetime_nrbg,len(ColLabels),'Average Lifetime','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
                 print(newrowoffset_Cu_lifetime_app)

    else:
        newrowoffset_Cu_lifetime_app = 2 

    if 'inf' in SectorList:
        Par_RECC_ProductLifetime_i_safe = np.einsum('prc,SR->prcSR', Par_RECC_ProductLifetime_inf,np.ones((NS,NR)))        
        for mm in range(0,Ni):
                 newrowoffset_Cu_lifetime_o = msf.xlsxExportAdd_tAB(ws_Cu_6, Par_RECC_ProductLifetime_i_safe[mm,0,SwitchTime-1:,:,:],mm*6+newrowoffset_Cu_lifetime_app,len(ColLabels),'Average Lifetime','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('i')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
                

    else:
        newrowoffset_Cu_lifetime_o = 2 


    if 'otv' in SectorList:
        Par_RECC_ProductLifetime_otv_safe = np.einsum('prc,SR->prcSR', Par_RECC_ProductLifetime_otv,np.ones((NS,NR)))        
        for mm in range(0,Nv):
                 newrowoffset_Cu_lifetime_otv = msf.xlsxExportAdd_tAB(ws_Cu_6, Par_RECC_ProductLifetime_otv_safe[mm,0,SwitchTime-1:,:,:],mm*6+newrowoffset_Cu_lifetime_o,len(ColLabels),'Average Lifetime','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('v')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
                

    if 'pav' in SectorList:
        Par_RECC_ProductLifetime_p_safe_regions = np.einsum('prc,SR->prcSR', Par_RECC_ProductLifetime_p,np.ones((NS,NR)))
     
        Par_RECC_ProductLifetime_p_safe = Par_RECC_ProductLifetime_p_safe_regions.mean(axis=1)
     
        for mm in range (0,Np):
             newrowoffset_Cu_lifetime_pav = msf.xlsxExportAdd_tAB(ws_Cu_6, Par_RECC_ProductLifetime_p_safe[mm,SwitchTime-1:,:,:],mm*6+2,len(ColLabels),'Average Lifetime','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items[mm],'Lifetime',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)     
    
        
       
    else:    
        newrowoffset_Cu_lifetime_pav = 2
       
    if 'reb' in SectorList:        
        Par_RECC_ProductLifetime_B_safe_regions = np.einsum('prc,SR->prcSR', Par_RECC_ProductLifetime_B,np.ones((NS,NR)))
        Par_RECC_ProductLifetime_B_safe = Par_RECC_ProductLifetime_B_safe_regions.mean(axis=1)

        
        for mm in range (0,NB):
             newrowoffset_Cu_lifetime_r = msf.xlsxExportAdd_tAB(ws_Cu_6, Par_RECC_ProductLifetime_B_safe[mm,SwitchTime-1:,:,:],mm*6+newrowoffset_Cu_lifetime_pav,len(ColLabels),'Average Lifetime','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items[mm],'Lifetime',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)     
             
     
    if  'ind' in SectorList:
        Par_RECC_ProductLifetime_I_safe_regions = np.einsum('prc,SR->prcSR', Par_RECC_ProductLifetime_ind,np.ones((NS,NR)))
        Par_RECC_ProductLifetime_I_safe=  Par_RECC_ProductLifetime_I_safe_regions.mean(axis=1)

       
        for mm in range (0,NI):
             newrowoffset_Cu_lifetime = msf.xlsxExportAdd_tAB(ws_Cu_6, Par_RECC_ProductLifetime_I_safe[mm,SwitchTime-1:,:,:],mm*6+2,len(ColLabels),'Average Lifetime','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items[mm],'Lifetime',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)     
   


 
    
    ws_Cu_7 = book.create_sheet('Collection rates')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_7.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_7.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_7.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_7.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


        Par_RECC_CollectionRate_safe = np.einsum('gt,SR->gtSR', Par_RECC_CollectionRate[:,:],np.ones((NS,NR)))        
        for mm in range(0,Ng):
                newrowoffset_Cu_collRate = msf.xlsxExportAdd_tAB(ws_Cu_7, Par_RECC_CollectionRate_safe[mm,:,:,:],2+mm*6,len(ColLabels),'Collection Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('g')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     

    


    
    ws_Cu_8 = book.create_sheet('Sorting rates cu')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_8.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_8.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_8.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_8.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


    if 'app' in SectorList or 'inf' in SectorList or 'nrbg' in SectorList or 'otv' in SectorList:

        Par_RECC_EoL_RR_No_safe = np.einsum('trT,SR->TtSR', Par_RECC_EoL_RR_No[:,:,6,:,5],np.ones((NS,NR)))        
        for mm in range(0,Ng):
                newrowoffset_Cu_sortingRate = msf.xlsxExportAdd_tAB(ws_Cu_8, Par_RECC_EoL_RR_No_safe[mm,:,:,:],2+mm*6,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     


    if 'ind' in SectorList:

        Par_RECC_EoL_RR_Nl_safe_regions = np.einsum('trT,SR->rTtSR', Par_RECC_EoL_RR_Nl[:,:,6,:,5],np.ones((NS,NR)))        
        Par_RECC_EoL_RR_Nl_safe = Par_RECC_EoL_RR_Nl_safe_regions.mean(axis=0)
        
        for mm in range(0,Ng):
                newrowoffset_Cu_sortingRate = msf.xlsxExportAdd_tAB(ws_Cu_8, Par_RECC_EoL_RR_Nl_safe[mm,:,:,:],2+mm*6,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     




    if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:

        Par_RECC_EoL_RR_Nr_safe_regions = np.einsum('trT,SR->rTtSR', Par_RECC_EoL_RR_Nr[:,:,6,:,5],np.ones((NS,NR)))        
        Par_RECC_EoL_RR_Nr_safe = Par_RECC_EoL_RR_Nr_safe_regions.mean(axis=0)
        
        
       
        for mm in range(0,Ng):
                newrowoffset_Cu_sortingRate = msf.xlsxExportAdd_tAB(ws_Cu_8, Par_RECC_EoL_RR_Nr_safe[mm,:,:,:],2+mm*6,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     

    ws_Cu_9 = book.create_sheet('Reuse rates')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_9.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_9.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_9.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_9.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


    

    if 'nrbg' in SectorList:
        ReUseFactor_tNSR_safe = np.einsum('tNS,R->tNSR', ReUseFactor_tmNrS[:,6,:,0,:],np.ones((NR)))        
        for mm in range(0,NN):
                newrowoffset_Cu_ReUseFactor_N = msf.xlsxExportAdd_tAB(ws_Cu_9, ReUseFactor_tNSR_safe[:,mm,:,:],2+mm*6,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
    else: 
        newrowoffset_Cu_ReUseFactor_N = 2

    if 'app' in SectorList:

        ReUseFactor_taSR_safe = np.einsum('taS,R->taSR', ReUseFactor_tmaS[:,6,:,:],np.ones((NR)))        
        for mm in range(0,Na):
                newrowoffset_Cu_ReUseFactor_a = msf.xlsxExportAdd_tAB(ws_Cu_9, ReUseFactor_taSR_safe[:,mm,:,:],mm*6+newrowoffset_Cu_ReUseFactor_N,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
        
    else: 
        newrowoffset_Cu_ReUseFactor_a = 2     
        
    if 'inf' in SectorList:
        
        ReUseFactor_tiSR_safe = np.einsum('taS,R->taSR', ReUseFactor_tmiS[:,6,:,:],np.ones((NR)))        
        for mm in range(0,Ni):
                newrowoffset_Cu_ReUseFactor = msf.xlsxExportAdd_tAB(ws_Cu_9, ReUseFactor_tiSR_safe[:,mm,:,:],mm*6+newrowoffset_Cu_ReUseFactor_a,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('i')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
        
    else: 
        newrowoffset_Cu_ReUseFactor = 2     
        
    if 'otv' in SectorList:
        
        ReUseFactor_tvSR_safe = np.einsum('taS,R->taSR', ReUseFactor_tmvS[:,6,:,:],np.ones((NR)))        
        for mm in range(0,Nv):
                newrowoffset_Cu_ReUseFactor = msf.xlsxExportAdd_tAB(ws_Cu_9, ReUseFactor_tvSR_safe[:,mm,:,:],mm*6+newrowoffset_Cu_ReUseFactor,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('v')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
       

    if 'pav' in SectorList:

        ReUseFactor_tmprS_safe_regions = np.einsum('tprS,R->tprSR', ReUseFactor_tmprS[:,6,:,:,:],np.ones((NR)))        
        ReUseFactor_tmprS_safe = ReUseFactor_tmprS_safe_regions.mean(axis=2)
        
        
       
        for mm in range(0,Np):
                newrowoffset_Cu_ReUseFactor_p = msf.xlsxExportAdd_tAB(ws_Cu_9, ReUseFactor_tmprS_safe[:,mm,:,:],2+mm*6,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
 
    else: 
        newrowoffset_Cu_ReUseFactor_p = 2             
        
    
    if 'reb' in SectorList:      
        
        ReUseFactor_tmBrS_safe_regions = np.einsum('tprS,R->tprSR', ReUseFactor_tmBrS[:,6,:,:,:],np.ones((NR)))        
        ReUseFactor_tmBrS_safe = ReUseFactor_tmBrS_safe_regions.mean(axis=2)
        
        
       
        for mm in range(0,NB):
                newrowoffset_Cu_ReUseFactor_B = msf.xlsxExportAdd_tAB(ws_Cu_9, ReUseFactor_tmBrS_safe[:,mm,:,:],mm*6+newrowoffset_Cu_ReUseFactor_p,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
       
        


    if 'ind' in SectorList:

        ReUseFactor_tmIS_safe = np.einsum('tIS,R->tISR', ReUseFactor_tmIS[:,6,:,:],np.ones((NR)))        
         
        for mm in range(0,Ng):
                newrowoffset_Cu_sortingRate = msf.xlsxExportAdd_tAB(ws_Cu_9, ReUseFactor_tmIS_safe[:,mm,:,:],2+mm*6,len(ColLabels),'Sorting Rate','years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     






    
    ws_Cu_10 = book.create_sheet('Model_Results_Cu_Stocks')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_10.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_10.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_10.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_10.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


    if 'nrbg' in SectorList:
        for mm in range(0,NN):
                newrowoffset_Cu_Stock_N = msf.xlsxExportAdd_tAB(ws_Cu_10, Cu_Stocks_No[:,mm,:,:],2+mm*6,len(ColLabels),'Cu Stock','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('N')].Items[mm],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
    else: 
        newrowoffset_Cu_Stock_N = 2

    if 'app' in SectorList:

        for mm in range(0,Na):
                newrowoffset_Cu_Stock_a = msf.xlsxExportAdd_tAB(ws_Cu_10, Cu_Stocks_No[:,mm,:,:],mm*6+newrowoffset_Cu_Stock_N,len(ColLabels),'Cu Stock','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('a')].Items[mm],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
        
    else: 
        newrowoffset_Cu_Stock_a = 2     
        
    if 'inf' in SectorList:
        
        for mm in range(0,Ni):
                newrowoffset_Cu_Stock_i = msf.xlsxExportAdd_tAB(ws_Cu_10, Cu_Stocks_No[:,mm,:,:],mm*6+newrowoffset_Cu_Stock_a,len(ColLabels),'Cu Stock','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('i')].Items[mm],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
    else: 
        newrowoffset_Cu_Stock_i = 2     
        
    if 'otv' in SectorList:
        
        for mm in range(0,Nv):
                newrowoffset_Cu_Stock_v = msf.xlsxExportAdd_tAB(ws_Cu_10, Cu_Stocks_No[:,mm,:,:],mm*6+newrowoffset_Cu_Stock_i,len(ColLabels),'Cu Stock','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('v')].Items[mm],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
           
  

    if 'pav' in SectorList:

       
        
       
        for mm in range(0,Np):
                newrowoffset_Cu_Stock_p = msf.xlsxExportAdd_tAB(ws_Cu_10, Cu_Stocks_Nr[:,mm,:,:],2+mm*6,len(ColLabels),'Cu Stock','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('p')].Items[mm],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
 
    else: 
        newrowoffset_Cu_Stock_p = 2             
        
    
    if 'reb' in SectorList:      
              
        
       
        for mm in range(0,NB):
                newrowoffset_Cu_Stock_B = msf.xlsxExportAdd_tAB(ws_Cu_10, Cu_Stocks_Nr[:,mm,:,:],mm*6+newrowoffset_Cu_Stock_p,len(ColLabels),'Cu Stock','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('B')].Items[mm],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     
       
        


    if 'ind' in SectorList:

         
        for mm in range(0,Ng):
                newrowoffset_Cu_Stock_I = msf.xlsxExportAdd_tAB(ws_Cu_10, Cu_Stocks_Nl[:,mm,:,:],2+mm*6,len(ColLabels),'Cu Stock','Mt',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('L')].Items[mm],'S_7',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     


 
    
    
    
    
    
    







    ws_Cu_11 = book.create_sheet('Ind stock')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_11.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_11.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+90):
        ws_Cu_11.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Cohort')].Items[n-m-1+70])
        ws_Cu_11.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


    if 'ind' in SectorList:
        for l in range(0,Nl):
            for mm in range(0,NI):
                newrowoffset_Cu_sortingRate = msf.xlsxExportAdd_tAB(ws_Cu_11, StockCurves_UsePhase_industry_NS_NR[70:,mm,l,:,:] ,2+mm*6+(NI*6)*l,len(ColLabels),IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('l')].Items[l],'years',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items[mm],'S_11',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
     


    ws_Cu_12 = book.create_sheet('Outflow')
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    newrowoffset_outflow = 2
  
    ColLabels = ['Indicator','Unit','Region','Product group','System_location','SocEc scen','ClimPol scen']
    for m in range(0,len(ColLabels)):
        ws_Cu_12.cell(row=1, column=m+1).value = ColLabels[m]
        ws_Cu_12.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
    for n in range(m+1,m+1+Nt):
        ws_Cu_12.cell(row=1, column=n+1).value = int(IndexTable.Classification[IndexTable.index.get_loc('Time')].Items[n-m-1])
        ws_Cu_12.cell(row=1, column=m+1).font  = openpyxl.styles.Font(bold=True)
  


    if 'app' in SectorList or 'nrbg' in SectorList or 'inf' in SectorList or 'otv' in SectorList:
            for mo in range(0,NO):
                newrowoffset_Cu_Outflow_o = msf.xlsxExportAdd_tAB(ws_Cu_12,Material_Outflow[:,mo,6,:,:],2+mo*6,len(ColLabels),'Material Outflow','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('O')].Items[mo],'F_7_8',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    if 'ind' in SectorList:
            for mI in range(0,NI):
                newrowoffset_Cu_Outflow_l = msf.xlsxExportAdd_tAB(ws_Cu_12,Material_Outflow[:,mI,6,:,:],2+mI*6,len(ColLabels),'Material Outflow','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('I')].Items[mI],'F_7_8',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
    if 'pav' in SectorList or 'reb' in SectorList or 'nrb' in SectorList:
            for mI in range(0,NT):
                newrowoffset_Cu_Outflow_r = msf.xlsxExportAdd_tAB(ws_Cu_12,Material_Outflow[:,mI,6,:,:],2+mI*6,len(ColLabels),'Material Outflow','Individual',ScriptConfig['RegionalScope'],IndexTable.Classification[IndexTable.set_index('IndexLetter').index.get_loc('T')].Items[mI],'F_7_8',IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items,IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items)
    
      
    



    # Post calibration 2015 parameter values
    Calib_Result_workbook = openpyxl.Workbook()
    S1 = Calib_Result_workbook.active
    S1.title = 'Cover'
    
    
    if 'pav' in SectorList:
        pav_Sheet = Calib_Result_workbook.create_sheet('passenger vehicles')
        pav_Sheet.cell(1,2).value = '2015 post calibration values, by model region'
        pav_Sheet.cell(1,2).font  = openpyxl.styles.Font(bold=True)
        pav_Sheet.cell(2,2).value = 'region'
        pav_Sheet.cell(2,2).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
            pav_Sheet.cell(m+1,2).value = Rname
            pav_Sheet.cell(m+1,2).font  = openpyxl.styles.Font(bold=True)
            m+=1
        # pC stock values
        pav_Sheet.cell(2,3).value = '2015 per capita stock values, total (all segments and drive technologies), by model region. Unit: 1 (veh. per person).'
        pav_Sheet.cell(2,3).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
            pav_Sheet.cell(m+1,3).value = TotalStockCurves_UsePhase_p_pC[0,m-2]
            m+=1
        # passenger-km
        pav_Sheet.cell(2,4).value = '2015 annual passenger kilometrage, by model region. Unit: km/yr.'
        pav_Sheet.cell(2,4).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
            pav_Sheet.cell(m+1,4).value = Total_Service_pav_tr_pC[0,m-2]
            m+=1
        # vehicle km
        pav_Sheet.cell(2,5).value = '2015 annual vehicle kilometrage, by model region. Unit: km/yr. Value for SSP1.'
        pav_Sheet.cell(2,5).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
            pav_Sheet.cell(m+1,5).value = ParameterDict['3_IO_Vehicles_UsePhase_eff'].Values[Service_Drivg,m-2,0,1]
            m+=1
        # vehicle occupancy rate
        pav_Sheet.cell(2,6).value = '2015 average vehicle occupancy rate, across all segments and drive technologies, by model region. Unit: km/yr. Value for SSP1.'
        pav_Sheet.cell(2,6).font  = openpyxl.styles.Font(bold=True)
        m=2
        # for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
        #     pav_Sheet.write(m,5,label = ParameterDict['6_MIP_VehicleOccupancyRate'].Values[Sector_pav_loc,m-2,0,1])
        #     m+=1
        # energy consumption, use phase
        pav_Sheet.cell(2,7).value = '2015 use phase energy consumption, across all segments and drive technologies, by model region. Unit: TJ/yr. Value for SSP1.'
        pav_Sheet.cell(2,7).font  = openpyxl.styles.Font(bold=True)
        m=2
        # for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
        #     pav_Sheet.cell(m+1,7).value = E_Calib_Vehicles[0,m-2]
        #     m+=1
            
    if 'reb' in SectorList:
        reb_Sheet = Calib_Result_workbook.create_sheet('residential buildings')
        reb_Sheet.cell(1,2).value = '2015 post calibration values, by model region'
        reb_Sheet.cell(1,2).font  = openpyxl.styles.Font(bold=True)
        reb_Sheet.cell(2,2).value = 'region'
        reb_Sheet.cell(2,2).font  = openpyxl.styles.Font(bold=True)
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
            reb_Sheet.cell(m+1,2).value = Rname
            reb_Sheet.cell(m+1,2).font  = openpyxl.styles.Font(bold=True)        
            m+=1
        # pC stock values
        reb_Sheet.cell(2,3).value = '2015 per capita stock values, total (all building types and energy standards), by model region. Unit: m2 per person.'
        reb_Sheet.cell(2,3).font  = openpyxl.styles.Font(bold=True)      
        m=2
        for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
            reb_Sheet.cell(m+1,3).value = TotalStockCurves_UsePhase_B_pC[0,m-2]
            m+=1
        # energy consumption, use phase
        reb_Sheet.cell(2,4).value = '2015 use phase energy consumption, across all building types and energy standards, by model region. Unit: TJ/yr. Value for SSP1.'
        reb_Sheet.cell(2,4).font  = openpyxl.styles.Font(bold=True)      
        m=2
      #  for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
      #      reb_Sheet.cell(m+1,4).value = E_Calib_Buildings[0,m-2]
      #      m+=1            
    
    # if 'nrb' in SectorList:
    #     nrb_Sheet = Calib_Result_workbook.create_sheet('nonres. buildings')
    #     nrb_Sheet.cell(1,2).value = '2015 post calibration values, by model region'
    #     nrb_Sheet.cell(1,2).font  = openpyxl.styles.Font(bold=True)  
    #     nrb_Sheet.cell(2,2).value = 'region'
    #     nrb_Sheet.cell(2,2).font  = openpyxl.styles.Font(bold=True)
    #     m=2
    #     for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
    #         nrb_Sheet.cell(m+1,2).value = Rname
    #         m+=1
    #     # pC stock values
    #     nrb_Sheet.cell(2,3).value = '2015 per capita stock values, total (all building types and energy standards), by model region. Unit: m2 per person.'
    #     nrb_Sheet.cell(2,3).font  = openpyxl.styles.Font(bold=True)
    #     m=2
    #     for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
    #         nrb_Sheet.cell(m+1,3).value = TotalStockCurves_UsePhase_N_pC[0,m-2]
    #         m+=1
    #     # energy consumption, use phase
    #     nrb_Sheet.cell(2,4).value = '2015 use phase energy consumption, across all building types and energy standards, by model region. Unit: TJ/yr. Value for SSP1.'
    #     nrb_Sheet.cell(2,4).font  = openpyxl.styles.Font(bold=True)
    #     m=2
    #     for Rname in IndexTable.Classification[IndexTable.index.get_loc('Region32')].Items:
    #         nrb_Sheet.cell(m+1,4).value = E_Calib_NRBuildgs[0,m-2]
    #         m+=1     
    
    # PLOT
    MyColorCycle = pylab.cm.Paired(np.arange(0,1,0.2))
    #linewidth = [1.2,2.4,1.2,1.2,1.2]
    linewidth  = [1.2,2,1.2]
    linewidth2 = [1.2,2,1.2]
    
    Figurecounter = 1
    LegendItems_SSP    = IndexTable.Classification[IndexTable.index.get_loc('Scenario')].Items
    #LegendItems_RCP    = IndexTable.Classification[IndexTable.index.get_loc('Scenario_RCP')].Items
    LegendItems_SSP_RE = ['LED, no EST', 'LED, 2°C ES', 'SSP1, no EST', 'SSP1, 2°C ES', 'SSP2, no EST', 'SSP2, 2°C ES']
    LegendItems_SSP_UP = ['Use Phase, SSP1, no EST', 'Rest of system GHG, SSP1, no EST','Use Phase, SSP1, 2°C ES', 'Rest of system GHG, SSP1, 2°C ES']
    ColorOrder         = [1,0,3]
    

    # Plot implementation curves
    # fig1, ax1 = plt.subplots()
    # ax1.set_prop_cycle('color', MyColorCycle)
    # ProxyHandlesList = []
    # for m in range(0,NR):
    plt.title('Strategy scale up in the SSP2 Baseline scenario', fontsize = 12) 
    plt.plot(np.arange(Model_Time_Start,Model_Time_End +1), ParameterDict['3_SHA_RECC_REStrategyScaleUp'].Values[0,0,:,2]*100) # world region, SSP1
    # #plt_lgd  = plt.legend(LegendItems_RCP,shadow = False, prop={'size':9}, loc = 'upper right' ,bbox_to_anchor=(1.20, 1))
    plt.ylabel('Potential seized, %.', fontsize = 12) 
    plt.xlabel('year', fontsize = 12) 
    # if ScriptConfig['UseGivenPlotBoundaries'] == True:    
    plt.axis([2015, 2050, 0, 110])
    # plt.show()
    # fig_name = 'ImplementationCurves_' + IndexTable.Classification[IndexTable.index.get_loc('Region')].Items[0]
    # # include figure in logfile:
    # fig_name = 'Figure ' + str(Figurecounter) + '_' + fig_name + '_' + ScriptConfig['RegionalScope'] + '.png'
    # fig1.savefig(os.path.join(ProjectSpecs_Path_Result, fig_name), dpi=DPI_RES, bbox_inches='tight')
    # Mylog.info('![%s](%s){ width=850px }' % (fig_name, fig_name))
    # Figurecounter += 1

    
    # Plot system emissions, by process, stacked.
    # Area plot, stacked, GHG emissions, material production, waste mgt, remelting, etc.
    MyColorCycle = pylab.cm.gist_earth(np.arange(0,1,0.155)) # select 12 colors from the 'Set1' color map.            
    #grey0_9      = np.array([0.9,0.9,0.9,1])
    
    SSPScens   = ['LED','SSP1','SSP2']
    RCPScens   = ['No climate policy','RCP2.6 energy mix']
    Area       = ['use phase','use phase, scope 2 (el)','use phase, other indirect','primary material product.','manufact. & recycling','total (+ forest & biogen. C)']     
    DataAExp   = np.zeros((NS,NR,Nt,6))
    
    
    
    ### 5.2) Export to Excel
    Mylog.info('### 5.2 - Export to Excel')
    # Export list data
    book.save(os.path.join(ProjectSpecs_Path_Result,'ODYM_RECC_ModelResults_'+ ScriptConfig['Current_UUID'] + '.xlsx'))
    
    # Export other table data
    Calib_Result_workbook.save(os.path.join(ProjectSpecs_Path_Result,'CalibResults.xlsx'))
    
    # Export area plot as multi-index Excel file:
    ColIndex       = [str(mmx) for mmx in  range(2015,2061)]
    RowIndex       = pd.MultiIndex.from_product([['use phase','use phase, scope 2 (el)','use phase, other indirect','primary material product.','manufact. & recycling','total (+ forest & biogen. C)'],['LED','SSP1','SSP2'],['Base','RCP2_6']], names=('System scope','SSP','RCP'))
    DF_GHGA_global = pd.DataFrame(np.einsum('SRts->sSRt',DataAExp).reshape(36,46), index=RowIndex, columns=ColIndex)
    #DF_GHGA_global.to_excel(os.path.join(ProjectSpecs_Path_Result,'GHG_Area_Data.xls'), merge_cells=False)    
    
    ## 5.3) Export as .mat file
    #Mylog.info('### 5.4 - Export to Matlab')
    #Mylog.info('Saving stock data to Matlab.')
    #Filestring_Matlab_out = os.path.join(ProjectSpecs_Path_Result, 'StockData.mat')
    #scipy.io.savemat(Filestring_Matlab_out, mdict={'F_6_7_tgmSR_Mt/yr': Material_Inflow, 'F_9_10_twSR_Mt/yr': Scrap_Outflow})
    
    ### 5.4) Model run is finished. Wrap up.
    Mylog.info('### 5.5 - Finishing')
    Mylog.debug("Converting " + os.path.join(ProjectSpecs_Path_Result, '..', log_filename))
    # everything from here on will not be included in the converted log file
    msf.convert_log(os.path.join(ProjectSpecs_Path_Result, log_filename))
    Mylog.info('Script is finished. Terminating logging process and closing all log files.')
    Time_End = time.time()
    Time_Duration = Time_End - Time_Start
    Mylog.info('End of simulation: ' + time.asctime())
    Mylog.info('Duration of simulation: %.1f seconds.' % Time_Duration)
    
    # remove all handlers from logger
    root = log.getLogger()
    root.handlers = []  # required if you don't want to exit the shell
    log.shutdown()
    
    ### 5.5) Create descriptive folder name and rename result folder
    SectList    = eval(ScriptConfig['SectorSelect'])
    DescrString = '__'
    FirstFlag   = True
    for sect in SectList:
        if FirstFlag is True:
            DescrString += sect
            FirstFlag = False
        else:
            DescrString += '_'
            DescrString += sect
    DescrString += '__'        
    
    ##### here was the RES strategy list appending block, I mived it up to use it already before to decide if a strategy should be applied in the model
            
    
    # Read REStratLit from config file
    REStratList = []
    if ScriptConfig['Include_REStrategy_FabYieldImprovement'] == 'True':
        REStratList.append('FYI')
    if ScriptConfig['Include_REStrategy_FabScrapDiversion'] == 'True':
        REStratList.append('FSD')    
    if ScriptConfig['Include_REStrategy_EoL_RR_Improvement'] == 'True':
        REStratList.append('EoL')
    if ScriptConfig['Include_REStrategy_MaterialSubstitution'] == 'True':
        REStratList.append('MSU')
    if ScriptConfig['Include_REStrategy_UsingLessMaterialByDesign'] == 'True':
        REStratList.append('ULD')
    if ScriptConfig['Include_REStrategy_ReUse'] == 'True':
        REStratList.append('RUS')
    if ScriptConfig['Include_REStrategy_LifeTimeExtension'] == 'True':
        REStratList.append('LTE')
    if ScriptConfig['Include_REStrategy_MoreIntenseUse'] == 'True':
        REStratList.append('MIU')
    if ScriptConfig['Include_REStrategy_CarSharing'] == 'True':
        REStratList.append('CaS')
    if ScriptConfig['Include_REStrategy_RideSharing'] == 'True':
        REStratList.append('RiS')
    if ScriptConfig['IncludeRecycling'] == 'False':
        REStratList.append('NoR')
    if ScriptConfig['No_EE_Improvements'] == 'True':        
        REStratList.append('NoEE')
    if ScriptConfig['Include_RES_ImproveSortingRate'] == 'True':        
        REStratList.append('SOR')
    if ScriptConfig['Include_RES_ImproveCollectionRate'] == 'True':    
        REStratList.append('COL')
    if ScriptConfig['Include_RES_GoodToService'] == 'True':    
        REStratList.append('GtS')
    if ScriptConfig['Include_Scenario'] == 'ambitious':    
        REStratList.append('amb')            
    
    FirstFlag = True
    if len(REStratList) > 0:
        for REStrat in REStratList:
            if FirstFlag is True:
                DescrString += REStrat
                FirstFlag = False
            else:
                DescrString += '_'
                DescrString += REStrat        
    
    ProjectSpecs_Path_Result_New = os.path.join(RECC_Paths.results_path, Name_Scenario + '__' + TimeString + DescrString)
    os.rename(ProjectSpecs_Path_Result,ProjectSpecs_Path_Result_New)
    
    
    
    print('done.')
    
    OutputDict['Name_Scenario'] = Name_Scenario + '__' + TimeString + DescrString # return new scenario folder name to ScenarioControl script
        
        #return OutputDict
                            
         #code for script to be run as standalone function
        #if __name__ == "__main__":
        #    main()
        
        
        # The End.
